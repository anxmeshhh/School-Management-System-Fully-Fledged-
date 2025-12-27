
from django.shortcuts import render
from django.contrib.auth import authenticate, login
from django.http import HttpResponseRedirect
from django.shortcuts import render, redirect
from django.contrib import messages

# Database connection
import pymysql
def get_db_connection():
    return pymysql.connect(
        host="localhost",  # Host
        user="root",  # Username
        password="theanimesh2005",  # Password
        database="school_db",  # Database name
        port=3306  # Port
    )




from django.http import HttpResponse
from django.shortcuts import render
from django.db import connection

def signup_view(request):
    if request.method == "POST":
        # Trim all inputs to prevent whitespace issues
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()

        # Validate required fields
        if not all([username, email, password, confirm_password]):
            return HttpResponse("All fields are required.")

        # Check if passwords match
        if password != confirm_password:
            return HttpResponse("Passwords do not match!")

        try:
            with connection.cursor() as cursor:
                # Check for existing user/email
                cursor.execute("SELECT id FROM users WHERE username = %s OR email = %s", (username, email))
                if cursor.fetchone():
                    return HttpResponse("Username or email already exists!")

                # Insert user into database
                query = "INSERT INTO users (username, email, password) VALUES (%s, %s, %s)"
                cursor.execute(query, (username, email, password))
                connection.commit()

                return HttpResponse("Success")

        except Exception as e:
            connection.rollback()  # Rollback on error
            error_msg = "Database error occurred. Please try again."
            if "Duplicate" in str(e) or "unique" in str(e).lower():
                error_msg = "Username or email already exists!"
            return HttpResponse(error_msg)

    return render(request, "users/index.html")


from django.shortcuts import render
from django.http import HttpResponse
from django.db import connection

from django.db import connection
from django.http import HttpResponse
from django.shortcuts import render

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()  # Trim user input
        password = request.POST.get("password", "").strip()  # Trim user input

        # Optional: Temp debug logs (remove after fixing)
        print(f"DEBUG: Input - Username: '{repr(username)}' (len: {len(username)})")
        print(f"DEBUG: Input - Password: '{repr(password)}' (len: {len(password)})")

        # Check user credentials in MySQL with TRIM for exact match
        with connection.cursor() as cursor:
            # Optional: Temp debug - find similar users (remove after)
            cursor.execute(
                "SELECT id, username, LENGTH(username) FROM users WHERE TRIM(username) LIKE %s", 
                (f"%{username}%",)
            )
            similar_users = cursor.fetchall()
            if similar_users:
                print("DEBUG: Similar users found:")
                for u in similar_users:
                    print(f"  ID: {u[0]}, Username: '{repr(u[1])}' (len: {u[2]})")

            # Exact match query with TRIM
            cursor.execute(
                "SELECT id, username FROM users WHERE TRIM(username) = %s AND TRIM(password) = %s", 
                (username, password)
            )
            user = cursor.fetchone()

        if user:
            # Store trimmed username in session
            clean_username = user[1].strip()
            request.session["user_id"] = user[0]
            request.session["username"] = clean_username
            
            # Optional: Temp success log (remove after)
            print(f"DEBUG: SUCCESS for '{clean_username}'")
            
            return HttpResponse("Success")

        # Optional: Temp failure log (remove after)
        print("DEBUG: No exact match found")
        
        # If credentials are invalid, send error message
        return HttpResponse("Invalid credentials!")

    return render(request, "users/login.html")


def change_password_view(request):
    if request.method == "POST":
        current_username = request.POST.get("current_username")
        new_username = request.POST.get("new_username")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        if new_password != confirm_password:
            return HttpResponse("Passwords do not match!")

        # Check if current user exists
        with connection.cursor() as cursor:
            cursor.execute("SELECT id FROM users WHERE username = %s", (current_username,))
            user = cursor.fetchone()

            if not user:
                return HttpResponse("User not found!")

            # Check if new username already exists (if different from current)
            if new_username != current_username:
                cursor.execute("SELECT id FROM users WHERE username = %s", (new_username,))
                existing_user = cursor.fetchone()
                if existing_user:
                    return HttpResponse("Username already exists!")

            # Update username and/or password
            if new_username != current_username:
                cursor.execute("UPDATE users SET username = %s, password = %s WHERE username = %s", (new_username, new_password, current_username))
            else:
                cursor.execute("UPDATE users SET password = %s WHERE username = %s", (new_password, current_username))
            connection.commit()

        return HttpResponse("Success")

    # For GET requests, redirect to login or handle appropriately
    return redirect('login')


 

from django.shortcuts import render

def dashboard_view(request):
    # Get the username from the session
    username = request.session.get("username", "Guest")  # Default to "Guest" if not logged in

    # Pass the username to the template
    return render(request, "users/dashboard.html", {"username": username})





import uuid
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction, connection
from django.utils import timezone
import os
from django.conf import settings
from django.contrib.auth.models import User
from django.http import JsonResponse

def profile_view(request):
    if "user_id" not in request.session:
        return redirect("/login/")  # Redirect to login if not authenticated

    user_id = request.session["user_id"]  # Get logged-in user's ID

    # Fetch profile picture
    profile_picture = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
            profile_picture_result = cursor.fetchone()
            if profile_picture_result:
                profile_picture = f"{settings.MEDIA_URL}{profile_picture_result[0]}"
                print(f"DEBUG: Found profile picture: {profile_picture}")
    except Exception as e:
        print("Error fetching profile picture:", e)

    if request.method == "POST":
        print(f"DEBUG: POST request received")
        print(f"DEBUG: FILES in request: {request.FILES}")
        print(f"DEBUG: POST data: {list(request.POST.keys())}")
        
        try:
            with transaction.atomic():
                # Handle ONLY profile picture upload
                if 'profile_picture' in request.FILES and request.FILES['profile_picture']:
                    print("DEBUG: Processing profile picture upload")
                    profile_picture_file = request.FILES['profile_picture']
                    print(f"DEBUG: File name: {profile_picture_file.name}, Size: {profile_picture_file.size}")
                    
                    # Validate file type
                    allowed_extensions = ['.png', '.jpg', '.jpeg']
                    file_ext = os.path.splitext(profile_picture_file.name)[1].lower()
                    if file_ext not in allowed_extensions:
                        messages.error(request, "Only PNG, JPG, or JPEG files are allowed.")
                        return redirect('profile_view')

                    # Validate file size (5MB limit)
                    if profile_picture_file.size > 5 * 1024 * 1024:
                        messages.error(request, "File size must be less than 5MB.")
                        return redirect('profile_view')

                    # Generate file path using UUID and user_id
                    filename = f"{uuid.uuid4().hex}_{user_id}{file_ext}"
                    
                    # Create pfpics directory in MEDIA_ROOT
                    pfpics_dir = os.path.join(settings.MEDIA_ROOT, 'pfpics')
                    os.makedirs(pfpics_dir, exist_ok=True)
                    
                    file_path = os.path.join(pfpics_dir, filename)
                    print(f"DEBUG: Saving file to: {file_path}")
                    
                    # Delete old profile picture if exists
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
                            old_pic = cursor.fetchone()
                            if old_pic:
                                old_file_path = os.path.join(settings.MEDIA_ROOT, old_pic[0])
                                if os.path.exists(old_file_path):
                                    os.remove(old_file_path)
                                    print(f"DEBUG: Deleted old file: {old_file_path}")
                    except Exception as e:
                        print(f"Error deleting old profile picture: {e}")

                    # Save new file
                    try:
                        with open(file_path, 'wb+') as destination:
                            for chunk in profile_picture_file.chunks():
                                destination.write(chunk)
                        print(f"DEBUG: File saved successfully to {file_path}")
                    except Exception as e:
                        print(f"ERROR: Failed to save file: {e}")
                        messages.error(request, "Failed to save file.")
                        return redirect('profile_view')

                    # Update or insert profile picture path in database
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("SELECT id FROM profile_pics WHERE user_id = %s", [user_id])
                            existing = cursor.fetchone()
                            
                            db_path = f"pfpics/{filename}"
                            if existing:
                                cursor.execute(
                                    "UPDATE profile_pics SET image_path = %s, uploaded_at = %s WHERE user_id = %s",
                                    [db_path, timezone.now(), user_id]
                                )
                                print(f"DEBUG: Updated existing record with path: {db_path}")
                            else:
                                cursor.execute(
                                    "INSERT INTO profile_pics (user_id, image_path, uploaded_at) VALUES (%s, %s, %s)",
                                    [user_id, db_path, timezone.now()]
                                )
                                print(f"DEBUG: Inserted new record with path: {db_path}")
                    except Exception as e:
                        print(f"ERROR: Database operation failed: {e}")
                        messages.error(request, "Failed to update database.")
                        return redirect('profile_view')
                    
                    messages.success(request, "Profile picture uploaded successfully!")
                    return redirect('profile_view')

                # Handle regular form submission (when no file is uploaded)
                else:
                    print("DEBUG: Processing regular form submission")
                    # Your existing form processing logic here
                    name = request.POST.get("name")
                    admission_number = request.POST.get("admission_number")
                    student_class = request.POST.get("class")
                    section = request.POST.get("section")
                    roll_number = request.POST.get("roll_number")
                    emis = request.POST.get("emis")

                    # Only process if we have actual form data
                    if name or admission_number:
                        # Validate required fields for Page 1
                        if not all([name, admission_number, student_class]):
                            messages.error(request, "Please fill in all required fields.")
                            return redirect('profile_view')

                        # Convert roll_number to integer if provided
                        if roll_number:
                            try:
                                roll_number = int(roll_number)
                            except (ValueError, TypeError):
                                messages.error(request, "Roll number must be a valid integer.")
                                return redirect('profile_view')

                        # Page 2 data
                        gender = request.POST.get("gender")
                        community = request.POST.get("community")
                        tamil_name = request.POST.get("tamil_name")
                        dob = request.POST.get("dob") or None
                        nationality = request.POST.get("nationality")
                        blood_group = request.POST.get("blood_group")
                        mother_tongue = request.POST.get("mother_tongue")
                        caste = request.POST.get("caste")
                        religion = request.POST.get("religion")
                        place_of_birth = request.POST.get("place_of_birth")
                        aadhaar = request.POST.get("aadhaar")
                        disability = request.POST.get("disability")
                        id_mark1 = request.POST.get("id_mark1")
                        id_mark2 = request.POST.get("id_mark2")
                        current_class = request.POST.get("current_class")
                        admission_class = request.POST.get("admission_class")
                        admission_year = request.POST.get("admission_year")
                        admission_date = request.POST.get("admission_date") or None

                        # Page 3 data (Communication Details)
                        email = request.POST.get("email")
                        address = request.POST.get("address")
                        contact = request.POST.get("contact")
                        alt_contact = request.POST.get("alt_contact")
                        country = request.POST.get("country")
                        state = request.POST.get("state")
                        city = request.POST.get("city")
                        pincode = request.POST.get("pincode")
                        status = request.POST.get("status")
                        house = request.POST.get("house")
                        teacher_ward = request.POST.get("teacher_ward")
                        rte = request.POST.get("rte")
                        sports_quota = request.POST.get("sports_quota")
                        prev_school = request.POST.get("prev_school")
                        prev_board = request.POST.get("prev_board")

                        # Page 4 data (Parent & Medical Information)
                        father_name = request.POST.get("father_name")
                        father_name_tamil = request.POST.get("father_name_tamil")
                        mother_name = request.POST.get("mother_name")
                        mother_name_tamil = request.POST.get("mother_name_tamil")
                        father_contact = request.POST.get("father_contact")
                        mother_contact = request.POST.get("mother_contact")
                        father_email = request.POST.get("father_email")
                        mother_email = request.POST.get("mother_email")
                        father_qualification = request.POST.get("father_qualification")
                        mother_qualification = request.POST.get("mother_qualification")
                        father_occupation = request.POST.get("father_occupation")
                        mother_occupation = request.POST.get("mother_occupation")
                        father_income = request.POST.get("father_income")
                        mother_income = request.POST.get("mother_income")
                        guardian_name = request.POST.get("guardian_name")
                        guardian_contact = request.POST.get("guardian_contact")
                        guardian_email = request.POST.get("guardian_email")
                        child_living = request.POST.get("child_living")
                        rights_on_child = request.POST.get("rights_on_child")
                        med_blood_group = request.POST.get("med_blood_group")
                        diseases = request.POST.get("diseases")
                        allergies = request.POST.get("allergies")
                        medicines = request.POST.get("medicines")
                        hospital = request.POST.get("hospital")
                        doctor = request.POST.get("doctor")

                        with connection.cursor() as cursor:
                            # Insert or Update student_page1
                            cursor.execute("""
                                INSERT INTO student_page1 (user_id, name, admission_number, class, section, roll_number, emis)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    name=VALUES(name), admission_number=VALUES(admission_number), 
                                    class=VALUES(class), section=VALUES(section), 
                                    roll_number=VALUES(roll_number), emis=VALUES(emis)
                            """, (user_id, name, admission_number, student_class, section, roll_number, emis))

                            # Insert or Update student_page2
                            cursor.execute("""
                                INSERT INTO student_page2 (user_id, gender, community, tamil_name, dob, nationality, blood_group, 
                                                           mother_tongue, caste, religion, place_of_birth, aadhaar, disability, 
                                                           id_mark1, id_mark2, current_class, admission_class, admission_year, admission_date)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    gender=VALUES(gender), community=VALUES(community), tamil_name=VALUES(tamil_name), dob=VALUES(dob),
                                    nationality=VALUES(nationality), blood_group=VALUES(blood_group), 
                                    mother_tongue=VALUES(mother_tongue), caste=VALUES(caste), religion=VALUES(religion),
                                    place_of_birth=VALUES(place_of_birth), aadhaar=VALUES(aadhaar),
                                    disability=VALUES(disability), id_mark1=VALUES(id_mark1), id_mark2=VALUES(id_mark2),
                                    current_class=VALUES(current_class), admission_class=VALUES(admission_class),
                                    admission_year=VALUES(admission_year), admission_date=VALUES(admission_date)
                            """, (user_id, gender, community, tamil_name, dob, nationality, blood_group, mother_tongue, 
                                  caste, religion, place_of_birth, aadhaar, disability, id_mark1, id_mark2, current_class, 
                                  admission_class, admission_year, admission_date))

                            # Insert or Update student_page3
                            cursor.execute("""
                                INSERT INTO student_page3 (user_id, email, address, contact, alt_contact, country, state, city, pincode, status, 
                                                           house, teacher_ward, rte, sports_quota, prev_school, prev_board)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    email=VALUES(email), address=VALUES(address), contact=VALUES(contact), alt_contact=VALUES(alt_contact),
                                    country=VALUES(country), state=VALUES(state), city=VALUES(city), pincode=VALUES(pincode), 
                                    status=VALUES(status), house=VALUES(house), teacher_ward=VALUES(teacher_ward), rte=VALUES(rte), 
                                    sports_quota=VALUES(sports_quota), prev_school=VALUES(prev_school), prev_board=VALUES(prev_board)
                            """, (user_id, email, address, contact, alt_contact, country, state, city, pincode, status, 
                                  house, teacher_ward, rte, sports_quota, prev_school, prev_board))

                            # Insert or Update student_page4 (Parent & Medical Information)
                            cursor.execute("""
                                INSERT INTO student_page4 (
                                    user_id, father_name, father_name_tamil, mother_name, mother_name_tamil, father_contact, 
                                    mother_contact, father_email, mother_email, father_qualification, mother_qualification, 
                                    father_occupation, mother_occupation, father_income, mother_income, guardian_name, 
                                    guardian_contact, guardian_email, child_living, rights_on_child, med_blood_group, 
                                    diseases, allergies, medicines, hospital, doctor
                                ) 
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    father_name=VALUES(father_name), father_name_tamil=VALUES(father_name_tamil),
                                    mother_name=VALUES(mother_name), mother_name_tamil=VALUES(mother_name_tamil),
                                    father_contact=VALUES(father_contact), mother_contact=VALUES(mother_contact),
                                    father_email=VALUES(father_email), mother_email=VALUES(mother_email),
                                    father_qualification=VALUES(father_qualification), mother_qualification=VALUES(mother_qualification),
                                    father_occupation=VALUES(father_occupation), mother_occupation=VALUES(mother_occupation),
                                    father_income=VALUES(father_income), mother_income=VALUES(mother_income),
                                    guardian_name=VALUES(guardian_name), guardian_contact=VALUES(guardian_contact),
                                    guardian_email=VALUES(guardian_email), child_living=VALUES(child_living),
                                    rights_on_child=VALUES(rights_on_child), med_blood_group=VALUES(med_blood_group),
                                    diseases=VALUES(diseases), allergies=VALUES(allergies), medicines=VALUES(medicines),
                                    hospital=VALUES(hospital), doctor=VALUES(doctor)
                            """, (
                                user_id, father_name, father_name_tamil, mother_name, mother_name_tamil, father_contact,
                                mother_contact, father_email, mother_email, father_qualification, mother_qualification,
                                father_occupation, mother_occupation, father_income, mother_income, guardian_name,
                                guardian_contact, guardian_email, child_living, rights_on_child, med_blood_group,
                                diseases, allergies, medicines, hospital, doctor
                            ))

                        messages.success(request, "Profile updated successfully.")
                        return redirect('profile_view')

        except Exception as e:
            print(f"ERROR: Exception in POST processing: {e}")
            messages.error(request, f"Failed to process request: {str(e)}")
            return redirect('profile_view')

    # Fetch student details for display
    student_data = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    s1.name, s1.admission_number, s1.class, s1.section, s1.roll_number, s1.emis,
                    s2.gender, s2.community, s2.tamil_name, s2.dob, s2.nationality, s2.blood_group, 
                    s2.mother_tongue, s2.caste, s2.religion, s2.place_of_birth, s2.aadhaar, s2.disability,
                    s2.id_mark1, s2.id_mark2, s2.current_class, s2.admission_class, s2.admission_year, s2.admission_date,
                    s3.email, s3.address, s3.contact, s3.alt_contact, s3.country, s3.state, s3.city, s3.pincode,
                    s3.status, s3.house, s3.teacher_ward, s3.rte, s3.sports_quota, s3.prev_school, s3.prev_board,
                    s4.father_name, s4.father_name_tamil, s4.mother_name, s4.mother_name_tamil, 
                    s4.father_contact, s4.mother_contact, s4.father_email, s4.mother_email,
                    s4.father_qualification, s4.mother_qualification, s4.father_occupation, s4.mother_occupation,
                    s4.father_income, s4.mother_income, s4.guardian_name, s4.guardian_contact,
                    s4.guardian_email, s4.child_living, s4.rights_on_child,
                    s4.med_blood_group, s4.diseases, s4.allergies, s4.medicines, 
                    s4.hospital, s4.doctor
                FROM student_page1 s1
                LEFT JOIN student_page2 s2 ON s1.user_id = s2.user_id
                LEFT JOIN student_page3 s3 ON s1.user_id = s3.user_id
                LEFT JOIN student_page4 s4 ON s1.user_id = s4.user_id
                WHERE s1.user_id = %s
            """, [user_id])
            student_data = cursor.fetchone()
    except Exception as e:
        print(f"DEBUG: Error fetching student data: {e}")

    print(f"DEBUG: Rendering template with profile_picture: {profile_picture}")
    return render(request, "users/profile.html", {
        "student_data": student_data,
        "profile_picture": profile_picture,
        "user_id": user_id
    })





from django.http import HttpResponse
from django.shortcuts import render, redirect
import PIL.Image
import PIL.ImageDraw
import PIL.ImageFont
from io import BytesIO
from django.db import connection
import qrcode
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from django.conf import settings

def generate_id_card(request):
    if "user_id" not in request.session:
        return redirect("/login/")

    user_id = request.session["user_id"]

    # Fetch student data
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                sp1.name, 
                sp1.class, 
                sp1.admission_number, 
                sp3.address
            FROM 
                student_page1 sp1
            JOIN 
                student_page3 sp3 ON sp1.user_id = sp3.user_id
            WHERE 
                sp1.user_id = %s
        """, [user_id])
        student_data = cursor.fetchone()

    if not student_data:
        return render(request, "users/profile.html", {"error": "Student data not found."})

    name, student_class, admission_number, address = student_data

    # Fetch profile picture
    profile_picture = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
            profile_picture_result = cursor.fetchone()
            if profile_picture_result:
                profile_picture_path = os.path.join(settings.MEDIA_ROOT, profile_picture_result[0])
                if os.path.exists(profile_picture_path):
                    profile_picture = PIL.Image.open(profile_picture_path).convert("RGB")
                else:
                    print(f"Profile picture file not found at: {profile_picture_path}")
    except Exception as e:
        print(f"Error fetching profile picture: {e}")

    # Load ID card template
    template_path = "users/static/users/images/id_card.jpg"
    template_image = PIL.Image.open(template_path).convert("RGB")
    draw = PIL.ImageDraw.Draw(template_image)

    # Get image dimensions for centering
    img_width, img_height = template_image.size

    # Place profile picture on the ID card (if available)
    if profile_picture:
        # Resize profile picture to fit the placeholder
        photo_size = (250, 250)
        profile_picture = profile_picture.resize(photo_size, PIL.Image.Resampling.LANCZOS)
        photo_x = (img_width - photo_size[0]) // 2  # Center horizontally
        photo_y = 251
        template_image.paste(profile_picture, (photo_x, photo_y))
        start_y = 569  # Adjusted Y position below the photo if image is present
    else:
        start_y = 569  # Adjusted Y position if no image is present (start earlier)

    # Load fonts with larger sizes
    font_path = os.path.join("users", "static", "users", "fonts", "arial.ttf")
    try:
        title_font = PIL.ImageFont.truetype(font_path, 36)  # For "IDENTITY CARD" text
        name_font = PIL.ImageFont.truetype(font_path, 32)  # For student name
        detail_font = PIL.ImageFont.truetype(font_path, 28)  # For other details
    except Exception as e:
        print(f"Font loading error: {e}")
        title_font = PIL.ImageFont.load_default()
        name_font = PIL.ImageFont.load_default()
        detail_font = PIL.ImageFont.load_default()

    # Function to calculate centered text position (pass img_width as parameter)
    def get_centered_x(text, font, img_width):
        text_width = draw.textlength(text, font=font)
        return (img_width - text_width) / 2

    # Position parameters
    line_spacing = 40  # Space between lines

    # Draw centered text, handle None values by defaulting to empty string
    current_y = start_y
    
    # Name (larger font)
    name_text = f"NAME: {(name or '').upper()}"  # Default to empty string if None
    draw.text((get_centered_x(name_text, name_font, img_width), current_y), 
              name_text, font=name_font, fill="black")
    current_y += line_spacing + 10
    
    # Class
    class_text = f"CLASS: {(student_class or '').upper()}"  # Default to empty string if None
    draw.text((get_centered_x(class_text, detail_font, img_width), current_y), 
              class_text, font=detail_font, fill="black")
    current_y += line_spacing + 10
    
    # Admission Number
    adm_text = f"ADMISSION NO: {(admission_number or '').upper()}"  # Default to empty string if None
    draw.text((get_centered_x(adm_text, detail_font, img_width), current_y), 
              adm_text, font=detail_font, fill="black")
    current_y += line_spacing + 10
    
    # Address (might need to handle multiline if too long)
    addr_text = f"ADDRESS: {(address or '').upper()}"  # Default to empty string if None
    draw.text((get_centered_x(addr_text, detail_font, img_width), current_y), 
              addr_text, font=detail_font, fill="black")

    # Generate QR code
    qr = qrcode.make(f"http://yourdomain.com/id_card/{admission_number or 'unknown'}/")  # Fallback for None
    qr_buffer = BytesIO()
    qr.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)
    qr_image = PIL.Image.open(qr_buffer)

    # Resize QR code
    qr_size = (175, 175)
    qr_image = qr_image.resize(qr_size, PIL.Image.Resampling.LANCZOS)

    # Position QR code at the top center
    qr_x = (img_width - qr_size[0]) // 2  # Center horizontally
    qr_y = 800  # Adjusted to top center as in the attached image
    template_image.paste(qr_image, (qr_x, qr_y))

    # Save image to memory buffer
    id_card_buffer = BytesIO()
    template_image.save(id_card_buffer, format="JPEG")
    id_card_buffer.seek(0)

    # Create PDF
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)

    # Get original dimensions of the ID card
    id_image = PIL.Image.open(id_card_buffer)
    orig_width, orig_height = id_image.size

    # Calculate available space on PDF
    margin = 50
    max_width = 612 - 2 * margin
    max_height = 792 - 2 * margin

    # Calculate scaling factor while maintaining aspect ratio
    scale = min(max_width / orig_width, max_height / orig_height)
    scaled_width = orig_width * scale
    scaled_height = orig_height * scale

    # Center the image on the page
    x_pos = (612 - scaled_width) / 2
    y_pos = (792 - scaled_height) / 2

    # Draw ID card with original aspect ratio
    c.drawInlineImage(id_image, x_pos, y_pos, width=scaled_width, height=scaled_height)

    c.showPage()
    c.save()
    pdf_buffer.seek(0)

    # Return PDF response
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ID_Card_{admission_number or "unknown"}.pdf"'
    return response



from django.shortcuts import render, redirect
from django.db import connection
import qrcode
from io import BytesIO
import base64
from datetime import datetime


def qr_page(request):
    if "user_id" not in request.session:
        return redirect("/login/")  # Redirect to login if not authenticated

    user_id = request.session["user_id"]  # Get logged-in user's ID

    # Fetch student data from the database using a JOIN for both tables
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                sp1.name, 
                sp1.class, 
                sp1.admission_number, 
                sp3.address
            FROM 
                student_page1 sp1
            JOIN 
                student_page3 sp3 ON sp1.user_id = sp3.user_id
            WHERE 
                sp1.user_id = %s
        """, [user_id])
        student_data = cursor.fetchone()

    if not student_data:
        return render(request, "users/profile.html", {"error": "Student data not found."})

    # Extract the student data
    name, student_class, admission_number, address = student_data

    # Generate QR code for the ID card URL
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L,
                       box_size=10, border=4)
    qr.add_data(f"http://yourdomain.com/id_card/{admission_number}/")  # Replace with actual URL
    qr.make(fit=True)

    # Create an in-memory image for the QR code
    img = qr.make_image(fill="black", back_color="white")
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)

    # Convert QR code to base64
    qr_code_base64 = base64.b64encode(img_io.read()).decode('utf-8')

    # Get current date for footer
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


    # Pass data to template
    context = {
        "name": name,
        "student_class": student_class,
        "address": address,
        "admission_number": admission_number,
        "qr_code_image": qr_code_base64,
        "date": current_date
    }

    return render(request, 'users/qr_page.html', context)






from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from PIL import Image as PILImage, ImageDraw, ImageFont
import os
import qrcode
from django.db import connection
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.conf import settings
from datetime import datetime


def calculate_age(dob):
    """Calculate age from DOB string in 'YYYY-MM-DD' format."""
    if not dob:
        return "N/A"
    try:
        birth_date = datetime.strptime(dob, '%Y-%m-%d').date()
        today = datetime.now().date()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return str(age)
    except:
        return "N/A"

def bulk_id_card(request):
    if not request.session.get('admin_id'):
        return redirect('admin_login')

    if request.method == 'POST':
        student_class = request.POST.get('class')
        section = request.POST.get('section')

        # Fetch students data including roll_number and DOB for age calculation
        with connection.cursor() as cursor:
            cursor.execute("""
    SELECT
        sp1.user_id,
        sp1.name,
        sp1.class,
        sp1.admission_number,
        sp1.roll_number,
        sp3.address,
        sp4.father_name,
        sp4.father_contact,
        sp4.father_email,
        sp4.mother_name,
        sp4.mother_contact,
        sp4.mother_email,
        sp2.dob
    FROM
        student_page1 sp1
    JOIN
        student_page3 sp3 ON sp1.user_id = sp3.user_id
    LEFT JOIN
        student_page2 sp2 ON sp1.user_id = sp2.user_id
    LEFT JOIN
        student_page4 sp4 ON sp1.user_id = sp4.user_id
    WHERE
        sp1.class = %s AND sp1.section = %s
""", [student_class, section])
            students_data = cursor.fetchall()

        if not students_data:
            return render(request, "users/bulk_id_card.html", {"error": "No students found for the selected class and section."})

        # Create PDF buffer
        pdf_buffer = BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=letter)

        for student_data in students_data:
            (
                user_id, name, student_class, admission_number, roll_number,
                address, father_name, father_contact, father_email,
                mother_name, mother_contact, mother_email, dob
            ) = student_data

            # Fetch profile picture
            profile_picture = None
            try:
                with connection.cursor() as cursor:
                    cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
                    profile_picture_result = cursor.fetchone()
                    if profile_picture_result:
                        profile_picture_path = os.path.join(settings.MEDIA_ROOT, profile_picture_result[0])
                        if os.path.exists(profile_picture_path):
                            profile_picture = PILImage.open(profile_picture_path).convert("RGB")
                        else:
                            print(f"Profile picture file not found at: {profile_picture_path}")
            except Exception as e:
                print(f"Error fetching profile picture for user {user_id}: {e}")

            # Load ID card template (same for front and back)
            template_path = "users/static/users/images/id_card.jpg"
            template_image = PILImage.open(template_path).convert("RGB")
            draw = ImageDraw.Draw(template_image)

            # Get image dimensions for centering
            img_width, img_height = template_image.size

            # --- FRONT SIDE ---
            # Place profile picture (if available)
            if profile_picture:
                photo_size = (250, 250)
                profile_picture = profile_picture.resize(photo_size, PILImage.Resampling.LANCZOS)
                photo_x = (img_width - photo_size[0]) // 2
                photo_y = 251
                template_image.paste(profile_picture, (photo_x, photo_y))
                start_y = 569
            else:
                start_y = 569

            # Load fonts
            font_path = os.path.join("users", "static", "users", "fonts", "arial.ttf")
            try:
                title_font = ImageFont.truetype(font_path, 36)
                name_font = ImageFont.truetype(font_path, 32)
                detail_font = ImageFont.truetype(font_path, 28)
            except Exception as e:
                print(f"Font loading error: {e}")
                title_font = ImageFont.load_default()
                name_font = ImageFont.load_default()
                detail_font = ImageFont.load_default()

            def get_centered_x(text, font, img_width):
                text_width = draw.textlength(text, font=font)
                return (img_width - text_width) / 2

            line_spacing = 40
            current_y = start_y

            # Draw "IDENTITY CARD" title (assuming it's already in template; if not, add it here)
            # draw.text((get_centered_x("IDENTITY CARD", title_font, img_width), 50), "IDENTITY CARD", font=title_font, fill="black")

            # Name
            name_text = f"NAME: {(name or '').upper()}"
            draw.text((get_centered_x(name_text, name_font, img_width), current_y),
                      name_text, font=name_font, fill="black")
            current_y += line_spacing + 10

            # Class
            class_text = f"CLASS: {(student_class or '').upper()}"
            draw.text((get_centered_x(class_text, detail_font, img_width), current_y),
                      class_text, font=detail_font, fill="black")
            current_y += line_spacing + 10

            # Admission Number
            adm_text = f"ADMISSION NO: {(admission_number or '').upper()}"
            draw.text((get_centered_x(adm_text, detail_font, img_width), current_y),
                      adm_text, font=detail_font, fill="black")
            current_y += line_spacing + 10

            # Roll Number (NEW)
            roll_text = f"ROLL NO: {(roll_number or 'N/A')}"
            draw.text((get_centered_x(roll_text, detail_font, img_width), current_y),
                      roll_text, font=detail_font, fill="black")
            current_y += line_spacing + 10

            # Address (moved to back side, so not on front anymore)
            # If you want to keep a short address on front, you can add it back here

            # Generate QR code
            qr = qrcode.make(f"http://yourdomain.com/id_card/{admission_number or 'unknown'}/")
            qr_buffer = BytesIO()
            qr.save(qr_buffer, format="PNG")
            qr_buffer.seek(0)
            qr_image = PILImage.open(qr_buffer)
            qr_size = (175, 175)
            qr_image = qr_image.resize(qr_size, PILImage.Resampling.LANCZOS)
            qr_x = (img_width - qr_size[0]) // 2
            qr_y = 800  # Your original top-center position
            template_image.paste(qr_image, (qr_x, qr_y))

            # Save front side image to buffer
            front_buffer = BytesIO()
            template_image.save(front_buffer, format="JPEG")
            front_buffer.seek(0)
            front_image = PILImage.open(front_buffer)
            orig_width, orig_height = front_image.size

            # Scale and center front side on PDF page
            margin = 50
            max_width = 612 - 2 * margin
            max_height = 792 - 2 * margin
            scale = min(max_width / orig_width, max_height / orig_height)
            scaled_width = orig_width * scale
            scaled_height = orig_height * scale
            x_pos = (612 - scaled_width) / 2
            y_pos = (792 - scaled_height) / 2
            c.drawInlineImage(front_image, x_pos, y_pos, width=scaled_width, height=scaled_height)
            c.showPage()  # End front page

            # --- BACK SIDE (NEW PAGE) ---
            # Reset template for back side
            template_image = PILImage.open(template_path).convert("RGB")
            draw = ImageDraw.Draw(template_image)

            # Clear any unwanted elements (optional: you can paste a plain template if needed)
            # For simplicity, we reuse the same template but draw text in a clean area

            # Back side start Y (adjust as per your template layout)
            back_start_y = 500  # Start higher on back to leave space for title or logo if any

            current_y = back_start_y

            # Title for back side
            back_title = "STUDENT DETAILS"
            draw.text((get_centered_x(back_title, title_font, img_width), current_y),
                      back_title, font=title_font, fill="black")
            current_y += line_spacing + 20

           

            # Father's Name
            father_text = f"FATHER'S NAME: {(father_name or 'N/A').upper()}"
            draw.text((get_centered_x(father_text, detail_font, img_width), current_y),
                      father_text, font=detail_font, fill="black")
            current_y += line_spacing

            # Father's Contact & Email
            father_contact_text = f"FATHER'S CONTACT: {(father_contact or 'N/A')}"
            draw.text((get_centered_x(father_contact_text, detail_font, img_width), current_y),
                      father_contact_text, font=detail_font, fill="black")
            current_y += line_spacing

            father_email_text = f"FATHER'S EMAIL: {(father_email or 'N/A')}"
            draw.text((get_centered_x(father_email_text, detail_font, img_width), current_y),
                      father_email_text, font=detail_font, fill="black")
            current_y += line_spacing + 10

            # Mother's Name
            mother_text = f"MOTHER'S NAME: {(mother_name or 'N/A').upper()}"
            draw.text((get_centered_x(mother_text, detail_font, img_width), current_y),
                      mother_text, font=detail_font, fill="black")
            current_y += line_spacing

            # Mother's Contact & Email
            mother_contact_text = f"MOTHER'S CONTACT: {(mother_contact or 'N/A')}"
            draw.text((get_centered_x(mother_contact_text, detail_font, img_width), current_y),
                      mother_contact_text, font=detail_font, fill="black")
            current_y += line_spacing

            mother_email_text = f"MOTHER'S EMAIL: {(mother_email or 'N/A')}"
            draw.text((get_centered_x(mother_email_text, detail_font, img_width), current_y),
                      mother_email_text, font=detail_font, fill="black")
            current_y += line_spacing + 10

            # Address
            addr_text = f"ADDRESS: {(address or 'N/A').upper()}"
            draw.text((get_centered_x(addr_text, detail_font, img_width), current_y),
                      addr_text, font=detail_font, fill="black")
            current_y += line_spacing + 10

            
            qr = qrcode.make(f"http://yourdomain.com/id_card/{admission_number or 'unknown'}/")
            qr_buffer_back = BytesIO()
            qr.save(qr_buffer_back, format="PNG")
            qr_buffer_back.seek(0)
            qr_image_back = PILImage.open(qr_buffer_back)
            qr_size_back = (175, 175)
            qr_image_back = qr_image_back.resize(qr_size_back, PILImage.Resampling.LANCZOS)
            qr_x_back = (img_width - qr_size_back[0]) // 2
            qr_y_back = 850  # Adjust if needed (e.g., 750 for higher, 850 for lower)
            template_image.paste(qr_image_back, (qr_x_back, qr_y_back))
            # === ADD THIS BLOCK END ===

            # Save back side to buffer
            back_buffer = BytesIO()
            template_image.save(back_buffer, format="JPEG")

            # Save back side to buffer
            back_buffer = BytesIO()
            template_image.save(back_buffer, format="JPEG")
            back_buffer.seek(0)
            back_image = PILImage.open(back_buffer)

            # Reuse same scaling and positioning as front
            back_orig_width, back_orig_height = back_image.size
            back_scale = min(max_width / back_orig_width, max_height / back_orig_height)
            back_scaled_width = back_orig_width * back_scale
            back_scaled_height = back_orig_height * back_scale
            back_x_pos = (612 - back_scaled_width) / 2
            back_y_pos = (792 - back_scaled_height) / 2

            c.drawInlineImage(back_image, back_x_pos, back_y_pos, width=back_scaled_width, height=back_scaled_height)
            c.showPage()  # End back page

        c.save()
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bulk_id_cards_front_back.pdf"'
        return response

    return render(request, "users/bulk_id_card.html")


from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db import connection

from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def admin_login(request):
    if request.method == 'POST':
        email = request.POST.get("email", "").strip()  # Trim user input
        password = request.POST.get("password", "").strip()  # Trim user input

        # Optional: Temp debug logs (remove after fixing)
        print(f"DEBUG: Input - Email: '{repr(email)}' (len: {len(email)})")
        print(f"DEBUG: Input - Password: '{repr(password)}' (len: {len(password)})")

        # Check user credentials in MySQL with TRIM for exact match
        with connection.cursor() as cursor:
            # Optional: Temp debug - find similar admins (remove after)
            cursor.execute(
                "SELECT id, full_name, email, LENGTH(email) FROM admins WHERE TRIM(email) LIKE %s", 
                (f"%{email}%",)
            )
            similar_admins = cursor.fetchall()
            if similar_admins:
                print("DEBUG: Similar admins found:")
                for a in similar_admins:
                    print(f"  ID: {a[0]}, Name: '{a[1]}', Email: '{repr(a[2])}' (len: {a[3]})")

            # Exact match query with TRIM
            cursor.execute(
                "SELECT id, full_name, email, password FROM admins WHERE TRIM(email) = %s AND TRIM(password) = %s", 
                (email, password)
            )
            admin = cursor.fetchone()

        if admin:
            # Store trimmed values in session
            clean_name = admin[1].strip()
            clean_email = admin[2].strip()
            request.session['admin_id'] = admin[0]
            request.session['admin_name'] = clean_name
            request.session['admin_email'] = clean_email  # Optional: Store email too
            
            # Optional: Temp success log (remove after)
            print(f"DEBUG: SUCCESS for '{clean_name}' ({clean_email})")
            
            return HttpResponse("Success")  # Return plain text for AJAX

        # Optional: Temp failure log (remove after)
        print("DEBUG: No exact match found")
        
        return HttpResponse("Invalid credentials")  # Return plain text for AJAX

    return render(request, 'users/admin_login.html')
def admin_change_credentials(request):
    if request.method == 'POST':
        current_email = request.POST.get('current_email')
        new_email = request.POST.get('new_email')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            return HttpResponse("Error: Passwords do not match")

        with connection.cursor() as cursor:
            # Check if current email exists
            cursor.execute("SELECT id FROM admins WHERE email = %s", [current_email])
            admin = cursor.fetchone()
            if not admin:
                return HttpResponse("Error: Current email not found")

            # Check if new email already exists (unless it's the same as current)
            if new_email != current_email:
                cursor.execute("SELECT id FROM admins WHERE email = %s", [new_email])
                existing = cursor.fetchone()
                if existing:
                    return HttpResponse("Error: New email already exists")

            # Update the admin record
            cursor.execute(
                "UPDATE admins SET email = %s, password = %s WHERE email = %s",
                [new_email, new_password, current_email]
            )
            # Commit the transaction (assuming autocommit is off; adjust if needed)
            connection.commit()

        return HttpResponse("Success")
    else:
        return HttpResponse("Error: Invalid request method")

from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db import connection
from django.contrib import messages

def admin_signup(request):
    if request.method == 'POST':
        # Trim all inputs to prevent whitespace issues
        full_name = request.POST.get("name", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()

        # Validate required fields
        if not all([full_name, email, password]):
            messages.error(request, "All fields are required.")
            return render(request, 'users/admin_signup.html')

        # Basic validation (add more as needed, e.g., email format)
        if len(password) < 6:
            messages.error(request, "Password must be at least 6 characters.")
            return render(request, 'users/admin_signup.html')

        try:
            with connection.cursor() as cursor:
                # Check for existing email
                cursor.execute("SELECT id FROM admins WHERE email = %s", (email,))
                if cursor.fetchone():
                    messages.error(request, "Email already exists!")
                    return render(request, 'users/admin_signup.html')

                # Insert admin into database
                query = "INSERT INTO admins (full_name, email, password) VALUES (%s, %s, %s)"
                cursor.execute(query, (full_name, email, password))
                connection.commit()

            messages.success(request, 'Signup successful! Please login.')
            return redirect('admin_login')

        except Exception as e:
            connection.rollback()  # Rollback on error
            error_msg = "Database error occurred. Please try again."
            if "Duplicate" in str(e) or "unique" in str(e).lower():
                error_msg = "Email already exists!"
            messages.error(request, error_msg)
            return render(request, 'users/admin_signup.html')

    return render(request, 'users/admin_signup.html')

def admin_page(request):
    return render(request, 'users/admin_page.html')





import os
from datetime import datetime
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import uuid
from django.db import connection
from django.contrib import messages

# Upload folder path (adjusted to your project structure)
UPLOAD_DIR = os.path.join(settings.BASE_DIR, 'users', 'static', 'uploads')

# Ensure uploads folder exists
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# Normalize string function to handle None and special cases
def normalize_value(value):
    if value is None:
        return ''
    return str(value).strip().lower()

# Admin uploads circular
from django.conf import settings
import os
# Add this
def admin_circular_upload(request):
    CIRCULARS_DIR = os.path.join(settings.MEDIA_ROOT, 'circulars')
    if request.method == 'POST':
        title = request.POST.get('title')
        image = request.FILES.get('image')
        target = request.POST.get('target', 'all')
        class_name = request.POST.get('class') if target == 'specific' else None
        section = request.POST.get('section') if target == 'specific' else None

        if title and image:
            # Validate class and section if specific
            if target == 'specific' and (not class_name or not section):
                messages.error(request, 'Please select both class and section for specific target.')
                return redirect('admin_circular_upload')

            # Normalize class and section
            class_name = normalize_value(class_name)
            section = normalize_value(section)

            # Generate a unique filename to avoid conflicts
            filename = f"{uuid.uuid4().hex}_{image.name}"
            fs = FileSystemStorage(location=CIRCULARS_DIR, base_url='/media/circulars/')
            try:
                filename = fs.save(filename, image)
                full_path = os.path.join(CIRCULARS_DIR, filename)
                if os.path.exists(full_path):
                    print(f"Image saved successfully: {full_path}")
                else:
                    print(f"Image save failed: {full_path}")
                    messages.error(request, 'Failed to save the image.')
                    return redirect('admin_circular_upload')
                image_url = f"/media/circulars/{filename}"  # Use consistent path
                print(f"Generated image_url: {image_url}")
            except Exception as e:
                print(f"Error saving image {filename}: {e}")
                messages.error(request, 'Error saving the image.')
                return redirect('admin_circular_upload')

            # Save circular metadata (title and target) in a unique file
            metadata_file_path = os.path.join(CIRCULARS_DIR, f"{filename}.txt")
            try:
                with open(metadata_file_path, 'w') as f:
                    f.write(f"{title}\n{target}")
                    if target == 'specific':
                        f.write(f"\n{class_name}\n{section}")
                print(f"Saved metadata for {filename}: title={title}, target={target}, class={class_name}, section={section}")
            except Exception as e:
                print(f"Error saving metadata for {filename}: {e}")
                messages.error(request, 'Error saving circular metadata.')
                return redirect('admin_circular_upload')

            messages.success(request, 'Circular uploaded successfully.')
            return redirect('admin_circular_upload')

    # Prepare circulars list for display
    circulars = []
    for file in os.listdir(CIRCULARS_DIR):
        if file.endswith(('.jpg', '.png', '.jpeg', '.webp', '.gif')):
            full_path = os.path.join(CIRCULARS_DIR, file)
            if not os.path.exists(full_path):
                print(f"Image file missing: {full_path}")
                continue

            title_file = f"{file}.txt"
            title_path = os.path.join(CIRCULARS_DIR, title_file)
            title = "Untitled"
            target = "All"
            class_name = ""
            section = ""

            if os.path.exists(title_path):
                try:
                    with open(title_path, 'r') as f:
                        lines = f.readlines()
                        title = lines[0].strip() if lines else "Untitled"
                        target = lines[1].strip() if len(lines) > 1 else "All"
                        if target == 'specific' and len(lines) >= 4:
                            class_name = normalize_value(lines[2])
                            section = normalize_value(lines[3])
                            target = f"Class: {class_name.capitalize()}, Section: {section.capitalize()}"
                except Exception as e:
                    print(f"Error reading metadata from {title_path}: {e}")

            try:
                created_at = datetime.fromtimestamp(os.path.getctime(full_path)).strftime('%Y-%m-%d %H:%M:%S')
                image_url = f"/media/circulars/{file}"  # Consistent path
                print(f"Listing circular: {file}, image_url: {image_url}, full_path: {full_path}")
                circulars.append({
                    'title': title,
                    'image_url': image_url,
                    'date': created_at,
                    'target': target
                })
            except Exception as e:
                print(f"Error processing file {file}: {e}")

    # Filter by date if parameter provided
    date_str = request.GET.get('date')
    date_filter = date_str
    if date_str:
        try:
            filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            filtered_circulars = []
            for circ in circulars:
                circ_date = datetime.strptime(circ['date'], '%Y-%m-%d %H:%M:%S').date()
                if circ_date == filter_date:
                    filtered_circulars.append(circ)
            circulars = filtered_circulars
        except ValueError:
            # Invalid date format, ignore filter
            pass

    # Sort by newest first
    circulars = sorted(circulars, key=lambda x: x['date'], reverse=True)

    # Fetch classes and sections from student_page1
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [normalize_value(row[0]) for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [normalize_value(row[0]) for row in cursor.fetchall()]

    return render(request, 'users/admin_circular_upload.html', {
        'circulars': circulars,
        'classes': classes,
        'sections': sections,
        'date_filter': date_filter
    })

# Student view of circulars
def student_circular(request):
    # Get the logged-in user's user_id from session and fetch class/section from student_page1
    student_class = None
    student_section = None
    error_message = None

    if "user_id" not in request.session:
        error_message = "Please log in to view circulars."
        print("No user_id found in session")
    else:
        user_id = request.session['user_id']
        print(f"Session user_id: {user_id}")
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT class, section FROM student_page1 WHERE user_id = %s",
                    [user_id]
                )
                result = cursor.fetchone()
                if result:
                    student_class, student_section = [normalize_value(r) for r in result]
                    print(f"Student user_id={user_id}: class={student_class}, section={student_section}")
                else:
                    error_message = "No class or section found for your account. Please contact the admin."
                    print(f"No student record found for user_id: {user_id}")
        except Exception as e:
            error_message = "Error fetching your class/section. Please try again later."
            print(f"Error fetching student class/section for user_id={user_id}: {e}")

    # Get filter type from POST request (default to 'all')
    filter_type = request.POST.get('filter_type', 'all')
    print(f"Filter type: {filter_type}")

    circulars = []
    for file in os.listdir(UPLOAD_DIR):
        if file.endswith(('.jpg', '.png', '.jpeg', '.webp', '.gif')):
            full_path = os.path.join(UPLOAD_DIR, file)
            if not os.path.exists(full_path):
                print(f"Image file missing in student view: {full_path}")
                continue

            title_file = f"{file}.txt"
            title_path = os.path.join(UPLOAD_DIR, title_file)
            title = "Untitled"
            target = "all"
            class_name = ""
            section = ""

            if os.path.exists(title_path):
                try:
                    with open(title_path, 'r') as f:
                        lines = f.readlines()
                        title = lines[0].strip() if lines else "Untitled"
                        target = lines[1].strip().lower() if len(lines) > 1 else "all"
                        if target == 'specific' and len(lines) >= 4:
                            class_name = normalize_value(lines[2])
                            section = normalize_value(lines[3])
                        print(f"Circular {file}: title={title}, target={target}, class={class_name}, section={section}")
                except Exception as e:
                    print(f"Error reading metadata from {title_path}: {e}")
                    continue

            # Filter circulars based on filter_type
            include_circular = False
            if filter_type == 'all':
                if target == "all" or (
                    target == "specific" and
                    student_class and student_section and
                    class_name == student_class and section == student_section
                ):
                    include_circular = True
            elif filter_type == 'specific':
                if target == "specific" and student_class and student_section and class_name == student_class and section == student_section:
                    include_circular = True

            if include_circular:
                try:
                    created_at = datetime.fromtimestamp(os.path.getctime(full_path)).strftime('%Y-%m-%d %H:%M:%S')
                    image_url = f"/static/uploads/{file}"  # Consistent path
                    print(f"Included circular: {file}, image_url: {image_url}, full_path: {full_path}")
                    display_target = "All" if target == "all" else f"Class: {class_name.capitalize()}, Section: {section.capitalize()}"
                    circulars.append({
                        'title': title,
                        'image_url': image_url,
                        'date': created_at,
                        'target': display_target
                    })
                except Exception as e:
                    print(f"Error processing file {file}: {e}")

    # Sort by newest first
    circulars = sorted(circulars, key=lambda x: x['date'], reverse=True)
    print(f"Total circulars displayed: {len(circulars)}")

    return render(request, 'users/student_circular.html', {
        'circulars': circulars,
        'student_class': student_class,
        'student_section': student_section,
        'filter_type': filter_type,
        'error_message': error_message
    })

# Teacher uploads circular
from django.conf import settings
import os
 # Add this
def teacher_circular_upload(request):
    CIRCULARS_DIR = os.path.join(settings.MEDIA_ROOT, 'circulars')
    if request.method == 'POST':
        title = request.POST.get('title')
        image = request.FILES.get('image')
        class_name = request.POST.get('class')
        section = request.POST.get('section')

        # Validate inputs
        if not title or not image or not class_name or not section:
            messages.error(request, 'Please provide title, image, class, and section.')
            return redirect('teacher_circular_upload')

        # Normalize class and section
        class_name = normalize_value(class_name)
        section = normalize_value(section)

        # Generate a unique filename to avoid conflicts
        filename = f"{uuid.uuid4().hex}_{image.name}"
        fs = FileSystemStorage(location=CIRCULARS_DIR, base_url='/media/circulars/')
        try:
            filename = fs.save(filename, image)
            full_path = os.path.join(CIRCULARS_DIR, filename)
            if os.path.exists(full_path):
                print(f"Image saved successfully: {full_path}")
            else:
                print(f"Image save failed: {full_path}")
                messages.error(request, 'Failed to save the image.')
                return redirect('teacher_circular_upload')
            image_url = f"/media/circulars/{filename}"  # Consistent path
            print(f"Generated image_url: {image_url}")
        except Exception as e:
            print(f"Error saving image {filename}: {e}")
            messages.error(request, 'Error saving the image.')
            return redirect('teacher_circular_upload')

        # Save circular metadata (title, target, class, section) in a unique file
        metadata_file_path = os.path.join(CIRCULARS_DIR, f"{filename}.txt")
        try:
            with open(metadata_file_path, 'w') as f:
                f.write(f"{title}\nspecific\n{class_name}\n{section}")
            print(f"Saved metadata for {filename}: title={title}, target=specific, class={class_name}, section={section}")
        except Exception as e:
            print(f"Error saving metadata for {filename}: {e}")
            messages.error(request, 'Error saving circular metadata.')
            return redirect('teacher_circular_upload')

        messages.success(request, 'Circular uploaded successfully.')
        return redirect('teacher_circular_upload')

    # Prepare circulars list for display
    circulars = []
    for file in os.listdir(CIRCULARS_DIR):
        if file.endswith(('.jpg', '.png', '.jpeg', '.webp', '.gif')):
            full_path = os.path.join(CIRCULARS_DIR, file)
            if not os.path.exists(full_path):
                print(f"Image file missing: {full_path}")
                continue

            title_file = f"{file}.txt"
            title_path = os.path.join(CIRCULARS_DIR, title_file)
            title = "Untitled"
            target = "All"
            class_name = ""
            section = ""

            if os.path.exists(title_path):
                try:
                    with open(title_path, 'r') as f:
                        lines = f.readlines()
                        title = lines[0].strip() if lines else "Untitled"
                        target = lines[1].strip() if len(lines) > 1 else "All"
                        if target == 'specific' and len(lines) >= 4:
                            class_name = normalize_value(lines[2])
                            section = normalize_value(lines[3])
                            target = f"Class: {class_name.capitalize()}, Section: {section.capitalize()}"
                except Exception as e:
                    print(f"Error reading title from {title_path}: {e}")

            try:
                created_at = datetime.fromtimestamp(os.path.getctime(full_path)).strftime('%Y-%m-%d %H:%M:%S')
                image_url = f"/media/circulars/{file}"  # Consistent path
                print(f"Listing circular: {file}, image_url: {image_url}, full_path: {full_path}")
                circulars.append({
                    'title': title,
                    'image_url': image_url,
                    'date': created_at,
                    'target': target
                })
            except Exception as e:
                print(f"Error processing file {file}: {e}")

    # Sort by newest first
    circulars = sorted(circulars, key=lambda x: x['date'], reverse=True)

    # Fetch classes and sections from student_page1
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [normalize_value(row[0]) for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [normalize_value(row[0]) for row in cursor.fetchall()]

    return render(request, 'users/teacher_circular_upload.html', {
        'circulars': circulars,
        'classes': classes,
        'sections': sections
    })


from datetime import datetime  # Add this import at the top of your file
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

def student_leave(request):
    """Handle student leave request submission and viewing."""
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access the student portal.")
        return redirect("/login/")

    user_id = request.session["user_id"]
    
    if request.method == "POST":
        try:
            form_data = {
                "student_name": request.POST.get("student_name", "").strip(),
                "reg_number": request.POST.get("reg_number", "").strip(),
                "class_number": request.POST.get("class", "").strip(),
                "leave_reason": request.POST.get("leave_reason", "").strip(),
                "leave_start_date": request.POST.get("leave_start_date", ""),
                "leave_end_date": request.POST.get("leave_end_date", ""),
                "leave_duration": request.POST.get("leave_duration", ""),
                "half_day_type": request.POST.get("half_day_type", "")
            }


            required_fields = ["student_name", "reg_number", "class_number", "leave_reason", 
                             "leave_start_date", "leave_end_date", "leave_duration"]
            missing_fields = [field for field in required_fields if not form_data[field]]
            if missing_fields:
                messages.error(request, f"Missing required fields: {', '.join(missing_fields)}")
                return redirect("student_leave")
            
            if form_data["leave_duration"] not in ["full", "half"]:
                messages.error(request, "Invalid leave duration.")
                return redirect("student_leave")
                
            if form_data["leave_duration"] == "half" and not form_data["half_day_type"]:
                messages.error(request, "Please select half day type for half-day leave.")
                return redirect("student_leave")

            try:
                start_date = datetime.strptime(form_data["leave_start_date"], "%Y-%m-%d")
                end_date = datetime.strptime(form_data["leave_end_date"], "%Y-%m-%d")
                if start_date > end_date:
                    messages.error(request, "End date must be on or after start date.")
                    return redirect("student_leave")
            except ValueError:
                messages.error(request, "Invalid date format.")
                return redirect("student_leave")

            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO student_leave_requests 
                    (user_id, student_name, reg_number, class_number, leave_reason,
                    leave_start_date, leave_end_date, leave_duration, half_day_type, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [user_id, form_data["student_name"], form_data["reg_number"], 
                      form_data["class_number"], form_data["leave_reason"], 
                      form_data["leave_start_date"], form_data["leave_end_date"],
                      form_data["leave_duration"], 
                      form_data["half_day_type"] if form_data["leave_duration"] == "half" else None,
                      "Pending"])
                connection.commit()
            messages.success(request, "Leave request submitted successfully.")
        except Exception as e:
            connection.rollback()
            messages.error(request, f"Error submitting leave request: {str(e)}")
        return redirect("student_leave")

    # Fetch leave requests for this student
    leave_requests = []
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                SELECT id, student_name, reg_number, class_number, leave_reason, 
                leave_start_date, leave_end_date, leave_duration, half_day_type, status
                FROM student_leave_requests WHERE user_id = %s
                ORDER BY leave_start_date DESC
            """, [user_id])
            leave_requests = cursor.fetchall()
        except Exception as e:
            messages.error(request, f"Error fetching leave requests: {str(e)}")

    return render(request, "users/student_leave.html", {
        "leave_requests": leave_requests
    })

def download_leave_pdf(request):
    """Generate and download leave request PDF."""
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access this page.")
        return redirect("/login/")

    user_id = request.session["user_id"]
    
    if request.method == "POST":
        try:
            leave_id = request.POST["leave_id"]
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT student_name, reg_number, class_number, leave_reason, leave_start_date,
                    leave_end_date, leave_duration, half_day_type, status
                    FROM student_leave_requests WHERE id = %s AND user_id = %s
                """, [leave_id, user_id])
                record = cursor.fetchone()
            if not record:
                messages.error(request, "Leave request not found.")
                return redirect("student_leave")

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            elements.append(Paragraph("Leave Request Details", styles["Title"]))
            data = [
                ["Field", "Details"],
                ["Student Name", record[0]],
                ["Registration Number", record[1]],
                ["Class", record[2]],
                ["Leave Reason", record[3]],
                ["Start Date", str(record[4])],
                ["End Date", str(record[5])],
                ["Duration", f"{record[6]}{' (' + record[7] + ')' if record[6] == 'half' and record[7] else ''}"],
                ["Status", record[8]]
            ]
            table = Table(data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            doc.build(elements)
            buffer.seek(0)
            return HttpResponse(buffer, content_type="application/pdf", headers={
                "Content-Disposition": f"attachment;filename=leave_request_{leave_id}.pdf"
            })
        except Exception as e:
            messages.error(request, f"Error generating PDF: {str(e)}")
            return redirect("student_leave")
    return redirect("student_leave")





from django.http import JsonResponse
from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages


def admin_accept_portal(request):
    """
    Enhanced Admin Leave Request Portal with TWO SECTIONS:
    1. Pending Leave Requests - Can approve/reject
    2. Request History - View approved/rejected requests (read-only)
    
    NO DATABASE CHANGES REQUIRED - Uses existing table structure
    """
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    # Handle POST requests for approving/rejecting leave
    if request.method == 'POST':
        action = request.POST.get('action')
        leave_id = request.POST.get('leave_id')

        if action and leave_id:
            if action == 'approve':
                new_status = 'Approved'
            elif action == 'reject':
                new_status = 'Rejected'
            else:
                messages.error(request, 'Invalid action')
                return redirect('admin_accept_portal')

            try:
                with connection.cursor() as cursor:
                    # Only update status - NO DATABASE STRUCTURE CHANGES
                    cursor.execute("""
                        UPDATE student_leave_requests
                        SET status = %s
                        WHERE id = %s AND status = 'Pending'
                    """, [new_status, leave_id])
                    
                    if cursor.rowcount == 0:
                        messages.error(request, 'Cannot update: Request already processed or not found.')
                    else:
                        messages.success(request, f'Leave request {new_status.lower()} successfully.')
            except Exception as e:
                print(f"Error updating leave request: {str(e)}")
                messages.error(request, 'Error updating leave request.')
        else:
            messages.error(request, 'Leave ID or action missing.')

        return redirect('admin_accept_portal')

    # GET request - Fetch leave requests separated into two sections
    
    # SECTION 1: Fetch PENDING leave requests only
    pending_requests = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    id, 
                    student_name, 
                    reg_number, 
                    class_number, 
                    section,
                    leave_reason, 
                    leave_start_date, 
                    leave_end_date, 
                    leave_duration, 
                    status
                FROM student_leave_requests
                WHERE status = 'Pending'
                ORDER BY leave_start_date DESC, id DESC
            """)
            pending_requests = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching pending requests: {str(e)}")
        messages.error(request, 'Error fetching pending leave requests.')
    
    # SECTION 2: Fetch APPROVED and REJECTED leave requests (History)
    history_requests = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    id, 
                    student_name, 
                    reg_number, 
                    class_number, 
                    section,
                    leave_reason, 
                    leave_start_date, 
                    leave_end_date, 
                    leave_duration, 
                    status,
                    NULL as processed_date
                FROM student_leave_requests
                WHERE status IN ('Approved', 'Rejected')
                ORDER BY id DESC
            """)
            history_requests = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching history requests: {str(e)}")
        messages.error(request, 'Error fetching leave request history.')
    
    # Calculate statistics
    pending_count = len(pending_requests)
    
    # Count approved and rejected requests
    approved_count = 0
    rejected_count = 0
    try:
        with connection.cursor() as cursor:
            # Count all approved requests
            cursor.execute("""
                SELECT COUNT(*) 
                FROM student_leave_requests
                WHERE status = 'Approved'
            """)
            result = cursor.fetchone()
            approved_count = result[0] if result else 0
            
            # Count all rejected requests
            cursor.execute("""
                SELECT COUNT(*) 
                FROM student_leave_requests
                WHERE status = 'Rejected'
            """)
            result = cursor.fetchone()
            rejected_count = result[0] if result else 0
    except Exception as e:
        print(f"Error calculating stats: {str(e)}")
    
    context = {
        'pending_requests': pending_requests,
        'history_requests': history_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
    }
    
    return render(request, 'users/admin_accept_portal.html', context)


def leave_get_students(request):
    """
    AJAX endpoint to fetch student names for filter dropdown
    Used by both pending and history sections
    NO DATABASE CHANGES - Uses existing student_page1 table
    """
    if not request.session.get('admin_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    class_num = request.GET.get('class')
    section = request.GET.get('section')

    if not class_num or not section:
        return JsonResponse({'students': []})

    try:
        with connection.cursor() as cursor:
            # Fetch from student_page1 table (existing table)
            cursor.execute("""
                SELECT DISTINCT name 
                FROM student_page1 
                WHERE class = %s 
                  AND section = %s 
                  AND name IS NOT NULL
                ORDER BY name
            """, [class_num, section])
            
            students = [row[0] for row in cursor.fetchall()]
            
        return JsonResponse({'students': students})
        
    except Exception as e:
        print("leave_get_students error:", str(e))
        return JsonResponse({'error': str(e)}, status=500)



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from django.http import JsonResponse

def teacher_accept_portal(request):
    """
    Enhanced View for Teacher Leave Request Portal with TWO SECTIONS:
    1. Pending Leave Requests - Can approve/reject
    2. Request History - View approved/rejected requests (read-only)
    
    NO DATABASE CHANGES REQUIRED - Uses existing table structure
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')

    # Handle POST requests for approving/rejecting leave
    if request.method == 'POST':
        action = request.POST.get('action')
        leave_id = request.POST.get('leave_id')

        if action and leave_id:
            if action not in ['approve', 'reject']:
                messages.error(request, 'Invalid action.')
                return redirect('teacher_accept_portal')

            new_status = 'Approved' if action == 'approve' else 'Rejected'

            try:
                with connection.cursor() as cursor:
                    # Only update status - NO DATABASE STRUCTURE CHANGES
                    cursor.execute("""
                        UPDATE student_leave_requests
                        SET status = %s
                        WHERE id = %s AND status = 'Pending'
                    """, [new_status, leave_id])

                    if cursor.rowcount == 0:
                        messages.error(request, 'Cannot update: Request already processed or not found.')
                    else:
                        action_text = 'approved' if action == 'approve' else 'rejected'
                        messages.success(request, f'Leave request {action_text} successfully.')
            except Exception as e:
                print(f"Error updating leave request: {str(e)}")
                messages.error(request, 'Error updating leave request. Please try again.')
        else:
            messages.error(request, 'Invalid request: Missing action or leave ID.')

        return redirect('teacher_accept_portal')

    # GET request - Fetch leave requests separated into two sections
    
    # SECTION 1: Fetch PENDING leave requests only
    pending_requests = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    id, 
                    student_name, 
                    reg_number, 
                    class_number, 
                    section,
                    leave_reason, 
                    leave_start_date, 
                    leave_end_date, 
                    leave_duration, 
                    status
                FROM student_leave_requests
                WHERE status = 'Pending'
                ORDER BY leave_start_date DESC, id DESC
            """)
            pending_requests = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching pending requests: {str(e)}")
        messages.error(request, 'Error fetching pending leave requests.')
    
    # SECTION 2: Fetch APPROVED and REJECTED leave requests (History)
    history_requests = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    id, 
                    student_name, 
                    reg_number, 
                    class_number, 
                    section,
                    leave_reason, 
                    leave_start_date, 
                    leave_end_date, 
                    leave_duration, 
                    status,
                    NULL as processed_date
                FROM student_leave_requests
                WHERE status IN ('Approved', 'Rejected')
                ORDER BY id DESC
            """)
            history_requests = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching history requests: {str(e)}")
        messages.error(request, 'Error fetching leave request history.')
    
    # Calculate statistics
    pending_count = len(pending_requests)
    
    # Count approved and rejected requests
    approved_count = 0
    rejected_count = 0
    try:
        with connection.cursor() as cursor:
            # Count all approved requests
            cursor.execute("""
                SELECT COUNT(*) 
                FROM student_leave_requests
                WHERE status = 'Approved'
            """)
            result = cursor.fetchone()
            approved_count = result[0] if result else 0
            
            # Count all rejected requests
            cursor.execute("""
                SELECT COUNT(*) 
                FROM student_leave_requests
                WHERE status = 'Rejected'
            """)
            result = cursor.fetchone()
            rejected_count = result[0] if result else 0
    except Exception as e:
        print(f"Error calculating stats: {str(e)}")
    
    context = {
        'pending_requests': pending_requests,
        'history_requests': history_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
    }
    
    return render(request, 'users/teacher_accept_portal.html', context)


def teacher_leave_get_students(request):
    """
    AJAX endpoint to fetch student names for filter dropdown
    Used by both pending and history sections
    NO DATABASE CHANGES - Uses existing student_page1 table
    """
    if not request.session.get('teacher_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    class_num = request.GET.get('class')
    section = request.GET.get('section')

    if not class_num or not section:
        return JsonResponse({'students': []})

    try:
        with connection.cursor() as cursor:
            # Fetch from student_page1 table (existing table)
            cursor.execute("""
                SELECT DISTINCT name 
                FROM student_page1 
                WHERE class = %s 
                  AND section = %s 
                  AND name IS NOT NULL
                ORDER BY name
            """, [class_num, section])
            
            students = [row[0] for row in cursor.fetchall()]
            
        return JsonResponse({'students': students})
        
    except Exception as e:
        print(f"teacher_leave_get_students error: {str(e)}")
        return JsonResponse({'error': 'Server error'}, status=500)



import os
import uuid
from django.shortcuts import render, redirect
from django.conf import settings
from django.contrib import messages
from django.db import connection
from datetime import datetime
from django.http import FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.validators import FileExtensionValidator
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage

# Admin uploads study material
@csrf_exempt
def admin_study_materials_upload(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    # Fetch distinct classes and sections for dropdowns
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT class FROM student_page1")
            classes = [row[0] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
            sections = [row[0] for row in cursor.fetchall()]
    except Exception as e:
        messages.error(request, f"Error fetching classes/sections: {str(e)}")
        classes = []
        sections = []

    if request.method == 'POST':
        title = request.POST.get("title")
        uploaded_file = request.FILES.get("file")
        selected_class = request.POST.get("class")
        selected_section = request.POST.get("section")

        if not all([title, uploaded_file, selected_class, selected_section]):
            messages.error(request, "Title, file, class, and section are required.")
            return redirect("admin_study_materials_upload")

        validator = FileExtensionValidator(allowed_extensions=['pdf'])
        try:
            validator(uploaded_file)
        except ValidationError:
            messages.error(request, "Only PDF files are allowed.")
            return redirect("admin_study_materials_upload")

        filename = f"{uuid.uuid4().hex}_{uploaded_file.name}"
        file_path = os.path.join(settings.MEDIA_ROOT, 'study_materials', filename)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        try:
            with open(file_path, 'wb+') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)
        except Exception as e:
            messages.error(request, f"Error saving file: {str(e)}")
            return redirect("admin_study_materials_upload")

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO study_materials (title, file_path, upload_date, class, section)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [title, f"study_materials/{filename}", datetime.now(), selected_class, selected_section]
                )
            messages.success(request, "Study material uploaded successfully!")
        except Exception as e:
            messages.error(request, f"Error saving to database: {str(e)}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return redirect("admin_study_materials_upload")

        return redirect("admin_study_materials_upload")

    # Get filter values
    selected_class = request.GET.get('class', '')
    selected_section = request.GET.get('section', '')

    try:
        with connection.cursor() as cursor:
            query = """
                SELECT title, file_path, upload_date, class, section
                FROM study_materials
                {where_clause}
                ORDER BY upload_date DESC
            """
            params = []
            where_clause = ""
            if selected_class and selected_section:
                where_clause = "WHERE class = %s AND section = %s"
                params = [selected_class, selected_section]

            cursor.execute(query.format(where_clause=where_clause), params)
            study_materials = [
                {
                    "title": r[0],
                    "file_path": r[1],
                    "upload_date": r[2],
                    "class": r[3],
                    "section": r[4]
                } for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving study materials: {str(e)}")
        study_materials = []

    return render(request, "users/admin_study_materials.html", {
        "study_materials": study_materials,
        "media_url": settings.MEDIA_URL,
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section
    })

# Teacher uploads study material
@csrf_exempt
def teacher_study_materials_upload(request):
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')

    # Fetch distinct classes and sections for dropdowns
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT class FROM student_page1")
            classes = [row[0] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
            sections = [row[0] for row in cursor.fetchall()]
    except Exception as e:
        messages.error(request, f"Error fetching classes/sections: {str(e)}")
        classes = []
        sections = []

    if request.method == 'POST':
        title = request.POST.get("title")
        uploaded_file = request.FILES.get("file")
        selected_class = request.POST.get("class")
        selected_section = request.POST.get("section")

        if not all([title, uploaded_file, selected_class, selected_section]):
            messages.error(request, "Title, file, class, and section are required.")
            return redirect("teacher_study_materials_upload")

        validator = FileExtensionValidator(allowed_extensions=['pdf'])
        try:
            validator(uploaded_file)
        except ValidationError:
            messages.error(request, "Only PDF files are allowed.")
            return redirect("teacher_study_materials_upload")

        filename = f"{uuid.uuid4().hex}_{uploaded_file.name}"
        file_path = os.path.join(settings.MEDIA_ROOT, 'study_materials', filename)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        try:
            with open(file_path, 'wb+') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)
        except Exception as e:
            messages.error(request, f"Error saving file: {str(e)}")
            return redirect("teacher_study_materials_upload")

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO study_materials (title, file_path, upload_date, class, section)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [title, f"study_materials/{filename}", datetime.now(), selected_class, selected_section]
                )
            messages.success(request, "Study material uploaded successfully!")
        except Exception as e:
            messages.error(request, f"Error saving to database: {str(e)}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return redirect("teacher_study_materials_upload")

        return redirect("teacher_study_materials_upload")

    # Get filter values
    selected_class = request.GET.get('class', '')
    selected_section = request.GET.get('section', '')

    try:
        with connection.cursor() as cursor:
            query = """
                SELECT title, file_path, upload_date, class, section
                FROM study_materials
                {where_clause}
                ORDER BY upload_date DESC
            """
            params = []
            where_clause = ""
            if selected_class and selected_section:
                where_clause = "WHERE class = %s AND section = %s"
                params = [selected_class, selected_section]

            cursor.execute(query.format(where_clause=where_clause), params)
            study_materials = [
                {
                    "title": r[0],
                    "file_path": r[1],
                    "upload_date": r[2],
                    "class": r[3],
                    "section": r[4]
                } for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving study materials: {str(e)}")
        study_materials = []

    return render(request, "users/teacher_study_materials.html", {
        "study_materials": study_materials,
        "media_url": settings.MEDIA_URL,
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section
    })

# Display study materials to students
def study_materials(request):
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access the student portal.")
        return redirect("/login/")

    user_id = request.session["user_id"]
    student_class = None
    student_section = None

    # Fetch student's class and section
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT class, section FROM student_page1 WHERE user_id = %s",
                [user_id]
            )
            result = cursor.fetchone()
            if result:
                student_class, student_section = result
            else:
                messages.error(request, "No class or section found for your account. Please contact the admin.")
                return redirect("/homework/")
    except Exception as e:
        messages.error(request, f"Error fetching class/section: {str(e)}")
        return redirect("/homework/")

    try:
        with connection.cursor() as cursor:
            query = """
                SELECT title, file_path, upload_date, class, section
                FROM study_materials
                WHERE class = %s AND section = %s
                ORDER BY upload_date DESC
            """
            cursor.execute(query, [student_class, student_section])
            study_materials = [
                {
                    "title": r[0],
                    "file_path": r[1],
                    "upload_date": r[2],
                    "class": r[3],
                    "section": r[4]
                } for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving study materials: {str(e)}")
        study_materials = []

    return render(request, "users/study_materials.html", {
        "study_materials": study_materials,
        "media_url": settings.MEDIA_URL,
        "student_class": student_class,
        "student_section": student_section
    })

# Serve PDF files
def serve_pdf(request, file_path):
    full_path = os.path.join(settings.MEDIA_ROOT, file_path.lstrip('/media/'))
    if not os.path.exists(full_path):
        raise Http404("File not found")
    
    try:
        response = FileResponse(open(full_path, 'rb'), content_type='application/pdf')
        if request.GET.get('download') == 'true':
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
        return response
    except Exception as e:
        raise Http404(f"Error serving file: {str(e)}")

# Student homework submission
import os
import uuid
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db import connection
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.core.validators import FileExtensionValidator
from django.core.exceptions import ValidationError

# Student homework submission view
def homework_view(request):
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access the student portal.")
        return redirect("/login/")

    user_id = request.session["user_id"]
    student_class = None
    student_section = None

    # Fetch student's class and section
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT class, section FROM student_page1 WHERE user_id = %s",
                [user_id]
            )
            result = cursor.fetchone()
            if result:
                student_class, student_section = result
            else:
                messages.error(request, "No class or section found for your account. Please contact the admin.")
                return redirect("/homework/")
    except Exception as e:
        messages.error(request, f"Error fetching class/section: {str(e)}")
        return redirect("/homework/")

    if request.method == "POST":
        title = request.POST.get("title")
        submission_date = request.POST.get("submission_date")
        uploaded_file = request.FILES.get("file")

        if not all([title, submission_date, uploaded_file]):
            messages.error(request, "All fields are required.")
            return redirect("/homework/")

        # Validate file is a PDF
        validator = FileExtensionValidator(allowed_extensions=['pdf'])
        try:
            validator(uploaded_file)
        except ValidationError:
            messages.error(request, "Only PDF files are allowed.")
            return redirect("/homework/")

        try:
            # Save file with unique filename
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads'))
            filename = f"{uuid.uuid4().hex}_{uploaded_file.name}"
            filename = fs.save(filename, uploaded_file)
            file_path = f"uploads/{filename}"

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO homework (user_id, title, submission_date, file_path, class, section)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    [user_id, title, submission_date, file_path, student_class, student_section]
                )
            messages.success(request, "Homework submitted successfully!")
        except Exception as e:
            messages.error(request, f"Error submitting homework: {str(e)}")
        return redirect("/homework/")

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT h.title, h.submission_date, h.file_path
                FROM homework h
                WHERE h.user_id = %s
                ORDER BY h.submission_date DESC
                """,
                [user_id]
            )
            homework_list = [
                {"title": r[0], "submission_date": r[1], "file_path": r[2]}
                for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving homework: {str(e)}")
        homework_list = []

    return render(request, "users/homework.html", {
        "homework_list": homework_list,
        "student_class": student_class,
        "student_section": student_section
    })

# Admin homework panel
@csrf_exempt
def admin_homework_panel(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    # Fetch distinct classes and sections
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT class FROM student_page1")
            classes = [row[0] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
            sections = [row[0] for row in cursor.fetchall()]
    except Exception as e:
        messages.error(request, f"Error fetching classes/sections: {str(e)}")
        classes = []
        sections = []

    # Get filter values
    selected_class = request.POST.get('class', '')
    selected_section = request.POST.get('section', '')

    try:
        with connection.cursor() as cursor:
            query = """
                SELECT h.user_id, h.title, h.submission_date, h.file_path, h.class, h.section, s.name
                FROM homework h
                JOIN student_page1 s ON h.user_id = s.user_id
                {where_clause}
                ORDER BY h.submission_date DESC
            """
            params = []
            where_clause = ""
            if selected_class and selected_section:
                where_clause = "WHERE h.class = %s AND h.section = %s"
                params = [selected_class, selected_section]

            cursor.execute(query.format(where_clause=where_clause), params)
            homework_list = [
                {
                    "user_id": r[0],
                    "title": r[1],
                    "submission_date": r[2],
                    "file_path": r[3],
                    "class": r[4],
                    "section": r[5],
                    "student_name": r[6]
                } for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving homework submissions: {str(e)}")
        homework_list = []

    return render(request, "users/admin_homework_panel.html", {
        "homework_list": homework_list,
        "media_url": settings.MEDIA_URL,
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section
    })

# Teacher homework panel
import os
import uuid
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from datetime import datetime


def normalize_value(value):
    """Normalize string values for consistency"""
    if value:
        return str(value).strip().lower()
    return value


from django.http import HttpResponse, Http404
from django.conf import settings
from django.shortcuts import redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
import os
import mimetypes

def serve_protected_file(request, file_path, file_type='homework'):
    """
    Generic protected file serving view that handles both teacher uploads and student submissions.
    Detects file type and serves appropriately with view/download options.
    
    Args:
        file_path: Relative path to the file
        file_type: Type of file ('homework' for teacher uploads, 'submission' for student submissions)
    """
    # Check if teacher is logged in
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this file.')
        return redirect('teacher_login')

    # Security: Normalize path to prevent directory traversal
    file_path = os.path.normpath(file_path).lstrip('/')
    
    # Determine base directory based on file type
    if file_type == 'submission':
        # Student submissions are in media root directly
        full_path = os.path.join(settings.MEDIA_ROOT, file_path)
    else:
        # Teacher homework files
        full_path = os.path.join(settings.MEDIA_ROOT, 'teacher_homework', file_path)

    # Check if file exists
    if not os.path.exists(full_path):
        messages.error(request, 'File not found.')
        return redirect('teacher_homework_panel')

    # Extra security: ensure file is inside MEDIA_ROOT
    if not os.path.abspath(full_path).startswith(os.path.abspath(settings.MEDIA_ROOT)):
        messages.error(request, 'Access denied.')
        return redirect('teacher_homework_panel')

    # Detect MIME type
    mime_type, _ = mimetypes.guess_type(full_path)
    file_ext = os.path.splitext(full_path)[1].lower()
    
    # Fallback MIME types for common file extensions
    mime_type_map = {
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.doc': 'application/msword',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp',
        '.txt': 'text/plain',
        '.zip': 'application/zip',
    }
    
    if mime_type is None:
        mime_type = mime_type_map.get(file_ext, 'application/octet-stream')

    # Read and serve file
    try:
        with open(full_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=mime_type)
            filename = os.path.basename(full_path)

            # Check if download is explicitly requested
            force_download = request.GET.get('download', '').lower() == 'true'
            
            if force_download:
                # Force download
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            else:
                # View inline in browser (default behavior)
                # Images, PDFs, and text files will display in browser
                # Other files will prompt download based on browser settings
                viewable_types = [
                    'image/jpeg', 'image/png', 'image/gif', 'image/webp',
                    'application/pdf', 'text/plain'
                ]
                
                if mime_type in viewable_types:
                    response['Content-Disposition'] = f'inline; filename="{filename}"'
                else:
                    # For non-viewable types (DOCX, XLSX), still try inline
                    # The browser will decide whether to display or download
                    response['Content-Disposition'] = f'inline; filename="{filename}"'

            return response
    except Exception as e:
        messages.error(request, f'Error serving file: {str(e)}')
        return redirect('teacher_homework_panel')


@csrf_exempt
def teacher_homework_panel(request):
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')

    teacher_id = request.session.get('teacher_id')
    HOMEWORK_DIR = os.path.join(settings.MEDIA_ROOT, 'teacher_homework')
    
    # Create homework directory if it doesn't exist
    if not os.path.exists(HOMEWORK_DIR):
        os.makedirs(HOMEWORK_DIR)

    # Handle homework upload
    if request.method == 'POST' and request.POST.get('upload_homework'):
        title = request.POST.get('homework_title', '').strip()
        description = request.POST.get('homework_description', '').strip()
        subject = request.POST.get('subject', '').strip()
        due_date = request.POST.get('due_date', '')
        target = request.POST.get('target', 'all')
        class_name = request.POST.get('class') if target == 'specific' else None
        section = request.POST.get('section') if target == 'specific' else None
        homework_file = request.FILES.get('homework_file')

        # Validation
        if not title:
            messages.error(request, 'Homework title is required.')
            return redirect('teacher_homework_panel')
        
        if not homework_file:
            messages.error(request, 'Please upload a file.')
            return redirect('teacher_homework_panel')

        # Validate class and section if specific
        if target == 'specific' and (not class_name or not section):
            messages.error(request, 'Please select both class and section for specific target.')
            return redirect('teacher_homework_panel')

        # Normalize class and section
        class_name = normalize_value(class_name)
        section = normalize_value(section)

        # Validate file type
        allowed_extensions = ['.pdf', '.docx', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.gif', '.webp']
        file_ext = os.path.splitext(homework_file.name)[1].lower()
        
        if file_ext not in allowed_extensions:
            messages.error(request, 'Invalid file type. Allowed: PDF, DOCX, XLSX, XLS, and images.')
            return redirect('teacher_homework_panel')

        # Validate file size (max 10MB)
        if homework_file.size > 10 * 1024 * 1024:
            messages.error(request, 'File size must be less than 10MB.')
            return redirect('teacher_homework_panel')

        # Generate unique filename
        filename = f"{uuid.uuid4().hex}_{homework_file.name}"
        fs = FileSystemStorage(location=HOMEWORK_DIR, base_url='/media/teacher_homework/')
        
        try:
            filename = fs.save(filename, homework_file)
            full_path = os.path.join(HOMEWORK_DIR, filename)
            
            if not os.path.exists(full_path):
                messages.error(request, 'Failed to save the file.')
                return redirect('teacher_homework_panel')
            
            file_url = f"/media/teacher_homework/{filename}"
        except Exception as e:
            messages.error(request, f'Error saving the file: {str(e)}')
            return redirect('teacher_homework_panel')

        # Save homework metadata in a unique file
        metadata_file_path = os.path.join(HOMEWORK_DIR, f"{filename}.txt")
        try:
            with open(metadata_file_path, 'w', encoding='utf-8') as f:
                f.write(f"{title}\n")
                f.write(f"{description}\n")
                f.write(f"{subject}\n")
                f.write(f"{due_date}\n")
                f.write(f"{target}\n")
                f.write(f"{teacher_id}\n")
                if target == 'specific':
                    f.write(f"{class_name}\n")
                    f.write(f"{section}\n")
        except Exception as e:
            messages.error(request, f'Error saving homework metadata: {str(e)}')
            return redirect('teacher_homework_panel')

        messages.success(request, f'Homework "{title}" uploaded successfully!')
        return redirect('teacher_homework_panel')

    # Prepare uploaded homework list for display
    uploaded_homework = []
    if os.path.exists(HOMEWORK_DIR):
        print(f"DEBUG: Checking homework directory: {HOMEWORK_DIR}")
        print(f"DEBUG: Current teacher_id: {teacher_id}")
        
        for file in os.listdir(HOMEWORK_DIR):
            # Skip metadata files
            if file.endswith('.txt'):
                continue
            
            print(f"DEBUG: Processing file: {file}")
            
            # Check for valid file extensions
            if not any(file.lower().endswith(ext) for ext in ['.pdf', '.docx', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.gif', '.webp']):
                print(f"DEBUG: Skipping {file} - invalid extension")
                continue
            
            full_path = os.path.join(HOMEWORK_DIR, file)
            if not os.path.exists(full_path):
                print(f"DEBUG: Skipping {file} - file doesn't exist")
                continue

            metadata_file = f"{file}.txt"
            metadata_path = os.path.join(HOMEWORK_DIR, metadata_file)
            
            title = "Untitled"
            description = ""
            subject = ""
            due_date = ""
            target = "All Classes"
            file_teacher_id = ""
            
            if os.path.exists(metadata_path):
                try:
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                        title = lines[0].strip() if len(lines) > 0 else "Untitled"
                        description = lines[1].strip() if len(lines) > 1 else ""
                        subject = lines[2].strip() if len(lines) > 2 else ""
                        due_date = lines[3].strip() if len(lines) > 3 else ""
                        target_type = lines[4].strip() if len(lines) > 4 else "all"
                        file_teacher_id = lines[5].strip() if len(lines) > 5 else ""
                        
                        print(f"DEBUG: File {file} - teacher_id from metadata: {file_teacher_id}")
                        
                        if target_type == 'specific' and len(lines) >= 8:
                            class_name = lines[6].strip()
                            section = lines[7].strip()
                            target = f"Class: {class_name.capitalize()}, Section: {section.capitalize()}"
                        else:
                            target = "All Classes"
                except Exception as e:
                    print(f"Error reading metadata from {metadata_path}: {e}")
            else:
                print(f"DEBUG: No metadata file found for {file}")
            
            # Only show homework uploaded by this teacher
            if file_teacher_id != str(teacher_id):
                print(f"DEBUG: Skipping {file} - teacher mismatch ({file_teacher_id} != {teacher_id})")
                continue

            print(f"DEBUG: Adding {file} to uploaded_homework list")

            try:
                file_ext = os.path.splitext(file)[1].lower()
                upload_date = datetime.fromtimestamp(os.path.getctime(full_path)).strftime('%B %d, %Y at %I:%M %p')
                
                # Use protected file serving URL
                file_url = f"/serve-file/homework/{file}"
                
                # Format due date if exists
                formatted_due_date = None
                if due_date:
                    try:
                        formatted_due_date = datetime.strptime(due_date, '%Y-%m-%d').strftime('%B %d, %Y')
                    except:
                        formatted_due_date = due_date
                
                uploaded_homework.append({
                    'title': title,
                    'description': description if description else None,
                    'subject': subject if subject else None,
                    'due_date': formatted_due_date,
                    'target': target,
                    'file_path': file,
                    'file_type': file_ext,
                    'date': upload_date,
                    'file_url': file_url
                })
            except Exception as e:
                print(f"Error processing file {file}: {e}")
        
        print(f"DEBUG: Total uploaded_homework items: {len(uploaded_homework)}")

    # Sort by newest first
    uploaded_homework = sorted(uploaded_homework, key=lambda x: x['date'], reverse=True)

    # Get filter values for student submissions
    selected_class = ''
    selected_section = ''
    
    if request.method == 'POST' and request.POST.get('filter_homework'):
        selected_class = normalize_value(request.POST.get('class', ''))
        selected_section = normalize_value(request.POST.get('section', ''))

    # Fetch student homework submissions
    homework_list = []
    try:
        with connection.cursor() as cursor:
            query = """
                SELECT h.user_id, h.title, h.submission_date, h.file_path, 
                       h.class, h.section, s.name
                FROM homework h
                JOIN student_page1 s ON h.user_id = s.user_id
                {where_clause}
                ORDER BY h.submission_date DESC
            """
            params = []
            where_clause = ""
            
            if selected_class and selected_section:
                where_clause = "WHERE h.class = %s AND h.section = %s"
                params = [selected_class, selected_section]
            elif selected_class:
                where_clause = "WHERE h.class = %s"
                params = [selected_class]
            elif selected_section:
                where_clause = "WHERE h.section = %s"
                params = [selected_section]

            cursor.execute(query.format(where_clause=where_clause), params)
            homework_list = [
                {
                    "user_id": r[0],
                    "title": r[1],
                    "submission_date": r[2].strftime('%B %d, %Y at %I:%M %p') if r[2] else 'N/A',
                    "file_path": r[3],
                    "class": r[4],
                    "section": r[5],
                    "student_name": r[6],
                    # Add protected URLs for student submissions
                    "view_url": f"/serve-file/submission/{r[3]}",
                    "download_url": f"/serve-file/submission/{r[3]}?download=true"
                } for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving homework submissions: {str(e)}")

    # Fetch classes and sections from student_page1
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT class FROM student_page1 ORDER BY class")
            classes = [normalize_value(row[0]) for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL ORDER BY section")
            sections = [normalize_value(row[0]) for row in cursor.fetchall()]
    except Exception as e:
        messages.error(request, f"Error fetching classes/sections: {str(e)}")
        classes = []
        sections = []

    # Default subjects list (can be customized)
    subjects = ['Mathematics', 'English', 'Science', 'Social Studies', 'Hindi', 
                'Computer Science', 'Physics', 'Chemistry', 'Biology', 'History', 
                'Geography', 'Economics', 'Accountancy']

    return render(request, "users/teacher_homework_panel.html", {
        "homework_list": homework_list,
        "uploaded_homework": uploaded_homework,
        "media_url": settings.MEDIA_URL,
        "classes": classes,
        "sections": sections,
        "subjects": subjects,
        "selected_class": selected_class,
        "selected_section": selected_section
    })


def teacher_view(request):
    return render(request, "users/teacher.html")

def fees(request):
    return render(request, "users/fees.html")


from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def view_edit_class(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    # Fetch all class-section pairs from admin_student_classes for current admin
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, class, section 
            FROM admin_student_classes 
            WHERE admin_id = %s
            ORDER BY class DESC, section DESC
        """, [admin_id])
        classes = cursor.fetchall()

    # Format as "class-section" for display
    class_list = [{'id': row[0], 'class_name': f"{row[1]}-{row[2]}"} for row in classes]

    return render(request, 'users/view_edit_class.html', {
        'classes': class_list,
        'total_classes': len(class_list)
    })

def add_class(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    if request.method == 'POST':
        class_name = request.POST.get('class_name')
        if class_name:
            try:
                class_part, section = class_name.split('-')
                
                with connection.cursor() as cursor:
                    # Check if this exact class-section combo already exists for this admin
                    cursor.execute("""
                        SELECT COUNT(*) FROM admin_student_classes
                        WHERE admin_id = %s AND class = %s AND section = %s
                    """, [admin_id, class_part, section])
                    exists = cursor.fetchone()[0]

                    if exists:
                        messages.error(request, f'You already created class {class_name}.')
                    else:
                        # Insert new class-section for this admin
                        cursor.execute("""
                            INSERT INTO admin_student_classes (admin_id, class, section)
                            VALUES (%s, %s, %s)
                        """, [admin_id, class_part, section])
                        
                        messages.success(request, f'Class {class_name} added successfully.')
                        return redirect('view_edit_class')
                        
            except ValueError:
                messages.error(request, 'Class name must be in format "Class-Section" (e.g., 2-A).')
            except Exception as e:
                messages.error(request, f'Error adding class: {str(e)}')
        else:
            messages.error(request, 'Class name cannot be empty.')
    
    return render(request, 'users/add_update_class.html', {'title': 'Add New Class'})

def update_class(request, class_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    # Fetch the class-section pair from admin_student_classes
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, class, section 
            FROM admin_student_classes 
            WHERE id = %s AND admin_id = %s
        """, [class_id, admin_id])
        class_data = cursor.fetchone()
        if not class_data:
            messages.error(request, 'Class not found or you don\'t have permission.')
            return redirect('view_edit_class')

    if request.method == 'POST':
        new_class_name = request.POST.get('class_name')
        if new_class_name:
            try:
                new_class, new_section = new_class_name.split('-')
                with connection.cursor() as cursor:
                    # Check if the new class-section combo already exists for this admin
                    cursor.execute("""
                        SELECT COUNT(*) FROM admin_student_classes
                        WHERE admin_id = %s AND class = %s AND section = %s AND id != %s
                    """, [admin_id, new_class, new_section, class_id])
                    exists = cursor.fetchone()[0]

                    if exists:
                        messages.error(request, f'You already have class {new_class_name}.')
                    else:
                        # Update the record in admin_student_classes
                        cursor.execute("""
                            UPDATE admin_student_classes
                            SET class = %s, section = %s
                            WHERE id = %s AND admin_id = %s
                        """, [new_class, new_section, class_id, admin_id])
                        messages.success(request, f'Class updated to {new_class_name} successfully.')
                        return redirect('view_edit_class')
            except ValueError:
                messages.error(request, 'Class name must be in format "Class-Section" (e.g., 2-A).')
            except Exception as e:
                messages.error(request, f'Error updating class: {str(e)}')
        else:
            messages.error(request, 'Class name cannot be empty.')
    return render(request, 'users/add_update_class.html', {
        'title': 'Update Class',
        'class_name': f"{class_data[1]}-{class_data[2]}"
    })

def delete_class(request, class_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    try:
        with connection.cursor() as cursor:
            # Delete the record from admin_student_classes
            cursor.execute("""
                DELETE FROM admin_student_classes 
                WHERE id = %s AND admin_id = %s
            """, [class_id, admin_id])
        messages.success(request, 'Class deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error deleting class: {str(e)}')
    return redirect('view_edit_class')


# NEW: Duplicate Class Function
def duplicate_class(request):
    """Duplicate an existing class with a new section name"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    admin_id = request.session['admin_id']
    
    if request.method == 'POST':
        source_class_id = request.POST.get('source_class_id')
        new_class_name = request.POST.get('new_class_name', '').strip()
        
        # Validation
        if not source_class_id or not new_class_name:
            messages.error(request, 'Please provide both source class and new class name.')
            return redirect('view_edit_class')
        
        # Validate format (should be "Class-Section")
        if '-' not in new_class_name:
            messages.error(request, 'Class name must be in format "Class-Section" (e.g., 10-B).')
            return redirect('view_edit_class')
        
        try:
            new_class, new_section = new_class_name.split('-')
            new_class = new_class.strip()
            new_section = new_section.strip().upper()  # Standardize section to uppercase
            
            if not new_class or not new_section:
                messages.error(request, 'Both class and section parts are required (e.g., 10-B).')
                return redirect('view_edit_class')
            
        except ValueError:
            messages.error(request, 'Invalid class name format. Use "Class-Section" (e.g., 10-B).')
            return redirect('view_edit_class')
        
        with connection.cursor() as cursor:
            # Check if source class exists and belongs to this admin
            cursor.execute("""
                SELECT class, section 
                FROM admin_student_classes 
                WHERE id = %s AND admin_id = %s
            """, [source_class_id, admin_id])
            source_class = cursor.fetchone()
            
            if not source_class:
                messages.error(request, 'Source class not found or you don\'t have permission.')
                return redirect('view_edit_class')
            
            # Check if new class-section already exists for this admin
            cursor.execute("""
                SELECT COUNT(*) 
                FROM admin_student_classes 
                WHERE admin_id = %s AND class = %s AND section = %s
            """, [admin_id, new_class, new_section])
            
            if cursor.fetchone()[0] > 0:
                messages.error(request, f'Class "{new_class}-{new_section}" already exists. Please choose a different name.')
                return redirect('view_edit_class')
            
            # Create the duplicate class
            try:
                cursor.execute("""
                    INSERT INTO admin_student_classes (admin_id, class, section)
                    VALUES (%s, %s, %s)
                """, [admin_id, new_class, new_section])
                
                source_class_display = f"{source_class[0]}-{source_class[1]}"
                new_class_display = f"{new_class}-{new_section}"
                
                messages.success(request, f'✅ Class "{new_class_display}" created successfully as a duplicate of "{source_class_display}"!')
                
            except Exception as e:
                messages.error(request, f'Error creating duplicate class: {str(e)}')
                return redirect('view_edit_class')
        
        return redirect('view_edit_class')
    
    # If not POST, redirect back
    messages.warning(request, 'Invalid request method.')
    return redirect('view_edit_class')



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def student_info(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    # Get filter parameters
    class_filter = request.GET.get('class', 'All')
    section_filter = request.GET.get('section', 'All')
    gender_filter = request.GET.get('gender', 'All')

    with connection.cursor() as cursor:
        # Get all unique class-section combinations
        cursor.execute("""
            SELECT DISTINCT class, section 
            FROM student_page1 
            ORDER BY class, section
        """)
        class_sections = cursor.fetchall()

        # Get total count of students
        cursor.execute("SELECT COUNT(*) FROM student_page1")
        total_students = cursor.fetchone()[0]

        # Get student data with filters, including profile picture path
        query = """
            SELECT sp1.id, sp1.name, sp1.admission_number, sp1.class, sp1.section, sp1.roll_number, sp2.gender, pp.image_path
            FROM student_page1 sp1
            LEFT JOIN student_page2 sp2 ON sp1.user_id = sp2.user_id
            LEFT JOIN profile_pics pp ON sp1.user_id = pp.user_id
            WHERE 1=1
        """
        params = []
        
        if class_filter != 'All':
            query += " AND sp1.class = %s"
            params.append(class_filter)
        if section_filter != 'All':
            query += " AND sp1.section = %s"
            params.append(section_filter)
        if gender_filter != 'All':
            query += " AND sp2.gender = %s"
            params.append(gender_filter)
        
        query += " ORDER BY sp1.class, sp1.section, sp1.name"
        cursor.execute(query, params)
        students = cursor.fetchall()

        # Get gender statistics for each class-section
        gender_stats_query = """
            SELECT sp1.class, sp1.section, 
                   COUNT(*) as total,
                   SUM(CASE WHEN sp2.gender = 'Male' THEN 1 ELSE 0 END) as male_count,
                   SUM(CASE WHEN sp2.gender = 'Female' THEN 1 ELSE 0 END) as female_count
            FROM student_page1 sp1
            LEFT JOIN student_page2 sp2 ON sp1.user_id = sp2.user_id
            GROUP BY sp1.class, sp1.section
            ORDER BY sp1.class, sp1.section
        """
        cursor.execute(gender_stats_query)
        gender_stats = cursor.fetchall()
        
        # Convert to dictionary for easy access
        gender_stats_dict = {}
        for stat in gender_stats:
            key = f"{stat[0]}-{stat[1]}"
            gender_stats_dict[key] = {
                'total': stat[2],
                'male': stat[3],
                'female': stat[4]
            }

    # Organize students by class-section
    class_section_groups = {}
    for student in students:
        class_section = f"{student[3]}-{student[4]}"
        if class_section not in class_section_groups:
            class_section_groups[class_section] = {
                'count': 0,
                'male_count': 0,
                'female_count': 0,
                'students': []
            }
        class_section_groups[class_section]['count'] += 1
        if student[6] == 'Male':
            class_section_groups[class_section]['male_count'] += 1
        elif student[6] == 'Female':
            class_section_groups[class_section]['female_count'] += 1
            
        class_section_groups[class_section]['students'].append({
            'id': student[0],
            'name': student[1],
            'admission_number': student[2],
            'class': student[3],
            'section': student[4],
            'roll_number': student[5],
            'gender': student[6],
            'image_path': student[7]  # Add image path to student data
        })

    context = {
        'class_section_groups': class_section_groups,
        'total_students': total_students,
        'class_options': sorted(list(set([cs[0] for cs in class_sections]))) + ['All'],
        'section_options': sorted(list(set([cs[1] for cs in class_sections]))) + ['All'],
        'gender_options': ['All', 'Male', 'Female'],
        'selected_class': class_filter,
        'selected_section': section_filter,
        'selected_gender': gender_filter,
        'gender_stats': gender_stats_dict
    }

    return render(request, 'users/student_info.html', context)


from django.contrib import messages
from django.db import connection
from django.shortcuts import render, redirect

def add_student(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        try:
            # Required fields
            name = request.POST.get('name', '').strip()
            admission_number = request.POST.get('admission_number', '').strip()
            class_section = request.POST.get('class_section', '').strip()
            roll_number = request.POST.get('roll_number', '').strip()
            emis = request.POST.get('emis', '').strip()
            email = request.POST.get('email', '').strip()

            # Validate required fields
            if not all([name, admission_number, class_section, roll_number, emis, email]):
                missing = [field for field, value in [
                    ('name', name),
                    ('admission_number', admission_number),
                    ('class_section', class_section),
                    ('roll_number', roll_number),
                    ('emis', emis),
                    ('email', email)
                ] if not value]
                messages.error(request, f'Missing required fields: {", ".join(missing)}')
                return render(request, 'users/add_student.html', {
                    'title': 'Add New Student',
                    # Pass all POST data to pre-populate form
                    **request.POST.dict(),
                    'gender_options': ['Male', 'Female', 'Other'],
                    'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                    'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                    'teacher_ward_options': ['yes', 'no'],
                    'rte_options': ['yes', 'no'],
                    'sports_quota_options': ['yes', 'no']
                })

            # Split class and section
            try:
                class_part, section = class_section.split('-')
            except ValueError:
                messages.error(request, 'Class-Section must be in format "Class-Section" (e.g., 2-A)')
                return render(request, 'users/add_student.html', {
                    'title': 'Add New Student',
                    **request.POST.dict(),
                    'gender_options': ['Male', 'Female', 'Other'],
                    'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                    'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                    'teacher_ward_options': ['yes', 'no'],
                    'rte_options': ['yes', 'no'],
                    'sports_quota_options': ['yes', 'no']
                })

            with connection.cursor() as cursor:
                # Check if admission number already exists
                cursor.execute("SELECT admission_number FROM student_page1 WHERE admission_number = %s", [admission_number])
                if cursor.fetchone():
                    messages.error(request, f'Admission number {admission_number} already exists.')
                    return render(request, 'users/add_student.html', {
                        'title': 'Add New Student',
                        **request.POST.dict(),
                        'gender_options': ['Male', 'Female', 'Other'],
                        'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                        'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                        'teacher_ward_options': ['yes', 'no'],
                        'rte_options': ['yes', 'no'],
                        'sports_quota_options': ['yes', 'no']
                    })

                # Create a user account
                username = f"student_{admission_number}"
                password = admission_number  # Default password (Note: Not hashed for simplicity as per original code)

                cursor.execute("""
                    INSERT INTO users (username, email, password)
                    VALUES (%s, %s, %s)
                """, [username, email, password])

                # Get the new user_id
                cursor.execute("SELECT LAST_INSERT_ID()")
                new_user_id = cursor.fetchone()[0]

                # Insert into student_page1
                cursor.execute("""
                    INSERT INTO student_page1 
                    (user_id, name, admission_number, class, section, roll_number, emis)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, [new_user_id, name, admission_number, class_part, section, roll_number, emis])

                # Insert into student_page2
                cursor.execute("""
                    INSERT INTO student_page2 
                    (user_id, gender, community, tamil_name, dob, nationality, 
                     blood_group, mother_tongue, caste, religion, place_of_birth, 
                     aadhaar, disability, id_mark1, id_mark2, current_class, 
                     admission_class, admission_year, admission_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    new_user_id,
                    request.POST.get('gender', ''),
                    request.POST.get('community', ''),
                    request.POST.get('tamil_name', ''),
                    request.POST.get('dob', None) or None,
                    request.POST.get('nationality', ''),
                    request.POST.get('blood_group', ''),
                    request.POST.get('mother_tongue', ''),
                    request.POST.get('caste', ''),
                    request.POST.get('religion', ''),
                    request.POST.get('place_of_birth', ''),
                    request.POST.get('aadhaar', ''),
                    request.POST.get('disability', ''),
                    request.POST.get('id_mark1', ''),
                    request.POST.get('id_mark2', ''),
                    class_part,
                    request.POST.get('admission_class', class_part),
                    request.POST.get('admission_year', ''),
                    request.POST.get('admission_date', None) or None
                ])

                # Insert into student_page3
                cursor.execute("""
                    INSERT INTO student_page3 
                    (user_id, email, address, contact, alt_contact, country, 
                     state, city, pincode, status, house, teacher_ward, 
                     rte, sports_quota, prev_school, prev_board)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s)
                """, [
                    new_user_id,
                    email,
                    request.POST.get('address', ''),
                    request.POST.get('contact', ''),
                    request.POST.get('alt_contact', ''),
                    request.POST.get('country', ''),
                    request.POST.get('state', ''),
                    request.POST.get('city', ''),
                    request.POST.get('pincode', ''),
                    request.POST.get('status', ''),
                    request.POST.get('house', ''),
                    request.POST.get('teacher_ward', 'no'),
                    request.POST.get('rte', 'no'),
                    request.POST.get('sports_quota', 'no'),
                    request.POST.get('prev_school', ''),
                    request.POST.get('prev_board', '')
                ])

                # Insert into student_page4
                cursor.execute("""
                    INSERT INTO student_page4 
                    (user_id, father_name, father_name_tamil, mother_name, mother_name_tamil,
                     father_contact, mother_contact, father_email, mother_email,
                     father_qualification, mother_qualification, father_occupation,
                     mother_occupation, father_income, mother_income, guardian_name,
                     guardian_contact, guardian_email, child_living, rights_on_child,
                     med_blood_group, diseases, allergies, medicines, hospital, doctor)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s)
                """, [
                    new_user_id,
                    request.POST.get('father_name', ''),
                    request.POST.get('father_name_tamil', ''),
                    request.POST.get('mother_name', ''),
                    request.POST.get('mother_name_tamil', ''),
                    request.POST.get('father_contact', ''),
                    request.POST.get('mother_contact', ''),
                    request.POST.get('father_email', ''),
                    request.POST.get('mother_email', ''),
                    request.POST.get('father_qualification', ''),
                    request.POST.get('mother_qualification', ''),
                    request.POST.get('father_occupation', ''),
                    request.POST.get('mother_occupation', ''),
                    request.POST.get('father_income', ''),
                    request.POST.get('mother_income', ''),
                    request.POST.get('guardian_name', ''),
                    request.POST.get('guardian_contact', ''),
                    request.POST.get('guardian_email', ''),
                    request.POST.get('child_living', ''),
                    request.POST.get('rights_on_child', ''),
                    request.POST.get('med_blood_group', ''),
                    request.POST.get('diseases', ''),
                    request.POST.get('allergies', ''),
                    request.POST.get('medicines', ''),
                    request.POST.get('hospital', ''),
                    request.POST.get('doctor', '')
                ])

                # Insert into admin_student_classes
                cursor.execute("""
                    INSERT INTO admin_student_classes
                    (admin_id, class, section)
                    VALUES (%s, %s, %s)
                """, [new_user_id, class_part, section])

            messages.success(request, f'Student {name} added successfully with Admission Number: {admission_number}')
            return redirect('student_info')

        except Exception as e:
            messages.error(request, f'Error adding student: {str(e)}')
            return render(request, 'users/add_student.html', {
                'title': 'Add New Student',
                **request.POST.dict(),
                'gender_options': ['Male', 'Female', 'Other'],
                'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                'teacher_ward_options': ['yes', 'no'],
                'rte_options': ['yes', 'no'],
                'sports_quota_options': ['yes', 'no']
            })

    # GET request - show empty form
    return render(request, 'users/add_student.html', {
        'title': 'Add New Student',
        'gender_options': ['Male', 'Female', 'Other'],
        'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
        'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
        'teacher_ward_options': ['yes', 'no'],
        'rte_options': ['yes', 'no'],
        'sports_quota_options': ['yes', 'no']
    })

from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages

from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages

def update_student(request, admission_number):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    student_data = None
    try:
        with connection.cursor() as cursor:
            # Get student data from multiple tables
            cursor.execute("""
                SELECT 
                    sp1.id, sp1.user_id, sp1.name, sp1.admission_number, sp1.class, sp1.section, 
                    sp1.roll_number, sp1.emis, sp3.email,
                    sp2.gender, sp2.community, sp2.tamil_name, sp2.dob, sp2.nationality,
                    sp2.blood_group, sp2.mother_tongue, sp2.caste, sp2.religion,
                    sp2.place_of_birth, sp2.aadhaar, sp2.disability, sp2.id_mark1,
                    sp2.id_mark2, sp2.current_class, sp2.admission_class, sp2.admission_year, sp2.admission_date,
                    sp3.address, sp3.contact, sp3.alt_contact, sp3.country, sp3.state,
                    sp3.city, sp3.pincode, sp3.status, sp3.house, sp3.teacher_ward,
                    sp3.rte, sp3.sports_quota, sp3.prev_school, sp3.prev_board,
                    sp4.father_name, sp4.father_name_tamil, sp4.mother_name, 
                    sp4.mother_name_tamil, sp4.father_contact, sp4.mother_contact,
                    sp4.father_email, sp4.mother_email, sp4.father_qualification,
                    sp4.mother_qualification, sp4.father_occupation, sp4.mother_occupation,
                    sp4.father_income, sp4.mother_income, sp4.guardian_name,
                    sp4.guardian_contact, sp4.guardian_email, sp4.child_living,
                    sp4.rights_on_child, sp4.med_blood_group, sp4.diseases,
                    sp4.allergies, sp4.medicines, sp4.hospital, sp4.doctor
                FROM student_page1 sp1
                LEFT JOIN student_page2 sp2 ON sp1.user_id = sp2.user_id
                LEFT JOIN student_page3 sp3 ON sp1.user_id = sp3.user_id
                LEFT JOIN student_page4 sp4 ON sp1.user_id = sp4.user_id
                WHERE sp1.admission_number = %s
            """, [admission_number])
            student_data = cursor.fetchone()

            if not student_data:
                messages.error(request, 'Student not found.')
                return redirect('student_info')

    except Exception as e:
        messages.error(request, f'Error fetching student data: {str(e)}')
        return redirect('student_info')

    if request.method == 'POST':
        try:
            # Required fields
            name = request.POST.get('name', '').strip()
            new_admission_number = request.POST.get('admission_number', '').strip()
            class_section = request.POST.get('class_section', '').strip()
            roll_number = request.POST.get('roll_number', '').strip()
            emis = request.POST.get('emis', '').strip()
            email = request.POST.get('email', '').strip()

            # Validate required fields
            if not all([name, new_admission_number, class_section, roll_number, emis, email]):
                missing = [field for field, value in [
                    ('name', name),
                    ('admission_number', new_admission_number),
                    ('class_section', class_section),
                    ('roll_number', roll_number),
                    ('emis', emis),
                    ('email', email)
                ] if not value]
                messages.error(request, f'Missing required fields: {", ".join(missing)}')
                # Inline context prep for error case
                post_data = dict(request.POST)
                # Map boolean fields from DB robustly (handles int 0/1 or str 'yes'/'no')
                teacher_ward_db = student_data[36] if len(student_data) > 36 else None
                teacher_ward = 'yes' if teacher_ward_db in (1, 'yes', 'Yes', True) else 'no'
                rte_db = student_data[37] if len(student_data) > 37 else None
                rte = 'yes' if rte_db in (1, 'yes', 'Yes', True) else 'no'
                sports_quota_db = student_data[38] if len(student_data) > 38 else None
                sports_quota = 'yes' if sports_quota_db in (1, 'yes', 'Yes', True) else 'no'
                context = {
                    'title': 'Update Student',
                    'admission_number': admission_number,
                    'post_data': post_data,  # For repopulating form on errors
                    'id': student_data[0] if len(student_data) > 0 else '',
                    'user_id': student_data[1] if len(student_data) > 1 else '',
                    'name': student_data[2] if len(student_data) > 2 else '',
                    'class_section': f"{student_data[4]}-{student_data[5]}" if len(student_data) > 5 else '',
                    'roll_number': student_data[6] if len(student_data) > 6 else '',
                    'emis': student_data[7] if len(student_data) > 7 else '',
                    'email': student_data[8] if len(student_data) > 8 else '',
                    'gender': student_data[9] if len(student_data) > 9 else '',
                    'community': student_data[10] if len(student_data) > 10 else '',
                    'tamil_name': student_data[11] if len(student_data) > 11 else '',
                    'dob': student_data[12] if len(student_data) > 12 else '',
                    'nationality': student_data[13] if len(student_data) > 13 else '',
                    'blood_group': student_data[14] if len(student_data) > 14 else '',
                    'mother_tongue': student_data[15] if len(student_data) > 15 else '',
                    'caste': student_data[16] if len(student_data) > 16 else '',
                    'religion': student_data[17] if len(student_data) > 17 else '',
                    'place_of_birth': student_data[18] if len(student_data) > 18 else '',
                    'aadhaar': student_data[19] if len(student_data) > 19 else '',
                    'disability': student_data[20] if len(student_data) > 20 else '',
                    'id_mark1': student_data[21] if len(student_data) > 21 else '',
                    'id_mark2': student_data[22] if len(student_data) > 22 else '',
                    'current_class': student_data[23] if len(student_data) > 23 else '',
                    'admission_class': student_data[24] if len(student_data) > 24 else '',
                    'admission_year': student_data[25] if len(student_data) > 25 else '',
                    'admission_date': student_data[26] if len(student_data) > 26 else '',
                    'address': student_data[27] if len(student_data) > 27 else '',
                    'contact': student_data[28] if len(student_data) > 28 else '',
                    'alt_contact': student_data[29] if len(student_data) > 29 else '',
                    'country': student_data[30] if len(student_data) > 30 else '',
                    'state': student_data[31] if len(student_data) > 31 else '',
                    'city': student_data[32] if len(student_data) > 32 else '',
                    'pincode': student_data[33] if len(student_data) > 33 else '',
                    'status': student_data[34] if len(student_data) > 34 else '',
                    'house': student_data[35] if len(student_data) > 35 else '',
                    'teacher_ward': teacher_ward,
                    'rte': rte,
                    'sports_quota': sports_quota,
                    'prev_school': student_data[39] if len(student_data) > 39 else '',
                    'prev_board': student_data[40] if len(student_data) > 40 else '',
                    'father_name': student_data[41] if len(student_data) > 41 else '',
                    'father_name_tamil': student_data[42] if len(student_data) > 42 else '',
                    'mother_name': student_data[43] if len(student_data) > 43 else '',
                    'mother_name_tamil': student_data[44] if len(student_data) > 44 else '',
                    'father_contact': student_data[45] if len(student_data) > 45 else '',
                    'mother_contact': student_data[46] if len(student_data) > 46 else '',
                    'father_email': student_data[47] if len(student_data) > 47 else '',
                    'mother_email': student_data[48] if len(student_data) > 48 else '',
                    'father_qualification': student_data[49] if len(student_data) > 49 else '',
                    'mother_qualification': student_data[50] if len(student_data) > 50 else '',
                    'father_occupation': student_data[51] if len(student_data) > 51 else '',
                    'mother_occupation': student_data[52] if len(student_data) > 52 else '',
                    'father_income': student_data[53] if len(student_data) > 53 else '',
                    'mother_income': student_data[54] if len(student_data) > 54 else '',
                    'guardian_name': student_data[55] if len(student_data) > 55 else '',
                    'guardian_contact': student_data[56] if len(student_data) > 56 else '',
                    'guardian_email': student_data[57] if len(student_data) > 57 else '',
                    'child_living': student_data[58] if len(student_data) > 58 else '',
                    'rights_on_child': student_data[59] if len(student_data) > 59 else '',
                    'med_blood_group': student_data[60] if len(student_data) > 60 else '',
                    'diseases': student_data[61] if len(student_data) > 61 else '',
                    'allergies': student_data[62] if len(student_data) > 62 else '',
                    'medicines': student_data[63] if len(student_data) > 63 else '',
                    'hospital': student_data[64] if len(student_data) > 64 else '',
                    'doctor': student_data[65] if len(student_data) > 65 else '',
                    'gender_options': ['Male', 'Female', 'Other'],
                    'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                    'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                    'teacher_ward_options': ['yes', 'no'],
                    'rte_options': ['yes', 'no'],
                    'sports_quota_options': ['yes', 'no']
                }
                return render(request, 'users/add_update_student.html', context)

            # Split class and section
            try:
                class_part, section = class_section.split('-')
            except ValueError:
                messages.error(request, 'Class-Section must be in format "Class-Section" (e.g., 2-A)')
                # Inline context prep for error case (same as above)
                post_data = dict(request.POST)
                teacher_ward_db = student_data[36] if len(student_data) > 36 else None
                teacher_ward = 'yes' if teacher_ward_db in (1, 'yes', 'Yes', True) else 'no'
                rte_db = student_data[37] if len(student_data) > 37 else None
                rte = 'yes' if rte_db in (1, 'yes', 'Yes', True) else 'no'
                sports_quota_db = student_data[38] if len(student_data) > 38 else None
                sports_quota = 'yes' if sports_quota_db in (1, 'yes', 'Yes', True) else 'no'
                context = {
                    'title': 'Update Student',
                    'admission_number': admission_number,
                    'post_data': post_data,
                    'id': student_data[0] if len(student_data) > 0 else '',
                    'user_id': student_data[1] if len(student_data) > 1 else '',
                    'name': student_data[2] if len(student_data) > 2 else '',
                    'class_section': f"{student_data[4]}-{student_data[5]}" if len(student_data) > 5 else '',
                    'roll_number': student_data[6] if len(student_data) > 6 else '',
                    'emis': student_data[7] if len(student_data) > 7 else '',
                    'email': student_data[8] if len(student_data) > 8 else '',
                    'gender': student_data[9] if len(student_data) > 9 else '',
                    'community': student_data[10] if len(student_data) > 10 else '',
                    'tamil_name': student_data[11] if len(student_data) > 11 else '',
                    'dob': student_data[12] if len(student_data) > 12 else '',
                    'nationality': student_data[13] if len(student_data) > 13 else '',
                    'blood_group': student_data[14] if len(student_data) > 14 else '',
                    'mother_tongue': student_data[15] if len(student_data) > 15 else '',
                    'caste': student_data[16] if len(student_data) > 16 else '',
                    'religion': student_data[17] if len(student_data) > 17 else '',
                    'place_of_birth': student_data[18] if len(student_data) > 18 else '',
                    'aadhaar': student_data[19] if len(student_data) > 19 else '',
                    'disability': student_data[20] if len(student_data) > 20 else '',
                    'id_mark1': student_data[21] if len(student_data) > 21 else '',
                    'id_mark2': student_data[22] if len(student_data) > 22 else '',
                    'current_class': student_data[23] if len(student_data) > 23 else '',
                    'admission_class': student_data[24] if len(student_data) > 24 else '',
                    'admission_year': student_data[25] if len(student_data) > 25 else '',
                    'admission_date': student_data[26] if len(student_data) > 26 else '',
                    'address': student_data[27] if len(student_data) > 27 else '',
                    'contact': student_data[28] if len(student_data) > 28 else '',
                    'alt_contact': student_data[29] if len(student_data) > 29 else '',
                    'country': student_data[30] if len(student_data) > 30 else '',
                    'state': student_data[31] if len(student_data) > 31 else '',
                    'city': student_data[32] if len(student_data) > 32 else '',
                    'pincode': student_data[33] if len(student_data) > 33 else '',
                    'status': student_data[34] if len(student_data) > 34 else '',
                    'house': student_data[35] if len(student_data) > 35 else '',
                    'teacher_ward': teacher_ward,
                    'rte': rte,
                    'sports_quota': sports_quota,
                    'prev_school': student_data[39] if len(student_data) > 39 else '',
                    'prev_board': student_data[40] if len(student_data) > 40 else '',
                    'father_name': student_data[41] if len(student_data) > 41 else '',
                    'father_name_tamil': student_data[42] if len(student_data) > 42 else '',
                    'mother_name': student_data[43] if len(student_data) > 43 else '',
                    'mother_name_tamil': student_data[44] if len(student_data) > 44 else '',
                    'father_contact': student_data[45] if len(student_data) > 45 else '',
                    'mother_contact': student_data[46] if len(student_data) > 46 else '',
                    'father_email': student_data[47] if len(student_data) > 47 else '',
                    'mother_email': student_data[48] if len(student_data) > 48 else '',
                    'father_qualification': student_data[49] if len(student_data) > 49 else '',
                    'mother_qualification': student_data[50] if len(student_data) > 50 else '',
                    'father_occupation': student_data[51] if len(student_data) > 51 else '',
                    'mother_occupation': student_data[52] if len(student_data) > 52 else '',
                    'father_income': student_data[53] if len(student_data) > 53 else '',
                    'mother_income': student_data[54] if len(student_data) > 54 else '',
                    'guardian_name': student_data[55] if len(student_data) > 55 else '',
                    'guardian_contact': student_data[56] if len(student_data) > 56 else '',
                    'guardian_email': student_data[57] if len(student_data) > 57 else '',
                    'child_living': student_data[58] if len(student_data) > 58 else '',
                    'rights_on_child': student_data[59] if len(student_data) > 59 else '',
                    'med_blood_group': student_data[60] if len(student_data) > 60 else '',
                    'diseases': student_data[61] if len(student_data) > 61 else '',
                    'allergies': student_data[62] if len(student_data) > 62 else '',
                    'medicines': student_data[63] if len(student_data) > 63 else '',
                    'hospital': student_data[64] if len(student_data) > 64 else '',
                    'doctor': student_data[65] if len(student_data) > 65 else '',
                    'gender_options': ['Male', 'Female', 'Other'],
                    'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                    'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                    'teacher_ward_options': ['yes', 'no'],
                    'rte_options': ['yes', 'no'],
                    'sports_quota_options': ['yes', 'no']
                }
                return render(request, 'users/add_update_student.html', context)

            user_id = student_data[1]  # user_id is at index 1 in the query result

            # Map yes/no to strings 'yes'/'no' for string-based boolean fields (to avoid truncation)
            teacher_ward_str = request.POST.get('teacher_ward', '').strip().lower()
            teacher_ward = 'yes' if teacher_ward_str == 'yes' else 'no'
            rte_str = request.POST.get('rte', '').strip().lower()
            rte = 'yes' if rte_str == 'yes' else 'no'
            sports_quota_str = request.POST.get('sports_quota', '').strip().lower()
            sports_quota = 'yes' if sports_quota_str == 'yes' else 'no'

            with connection.cursor() as cursor:
                # Check if new admission number already exists (if changed)
                if new_admission_number != admission_number:
                    cursor.execute("SELECT admission_number FROM student_page1 WHERE admission_number = %s", [new_admission_number])
                    if cursor.fetchone():
                        messages.error(request, f'Admission number {new_admission_number} already exists.')
                        # Inline context prep for error case (same as above)
                        post_data = dict(request.POST)
                        teacher_ward_db = student_data[36] if len(student_data) > 36 else None
                        teacher_ward_disp = 'yes' if teacher_ward_db in (1, 'yes', 'Yes', True) else 'no'
                        rte_db = student_data[37] if len(student_data) > 37 else None
                        rte_disp = 'yes' if rte_db in (1, 'yes', 'Yes', True) else 'no'
                        sports_quota_db = student_data[38] if len(student_data) > 38 else None
                        sports_quota_disp = 'yes' if sports_quota_db in (1, 'yes', 'Yes', True) else 'no'
                        context = {
                            'title': 'Update Student',
                            'admission_number': admission_number,
                            'post_data': post_data,
                            'id': student_data[0] if len(student_data) > 0 else '',
                            'user_id': student_data[1] if len(student_data) > 1 else '',
                            'name': student_data[2] if len(student_data) > 2 else '',
                            'class_section': f"{student_data[4]}-{student_data[5]}" if len(student_data) > 5 else '',
                            'roll_number': student_data[6] if len(student_data) > 6 else '',
                            'emis': student_data[7] if len(student_data) > 7 else '',
                            'email': student_data[8] if len(student_data) > 8 else '',
                            'gender': student_data[9] if len(student_data) > 9 else '',
                            'community': student_data[10] if len(student_data) > 10 else '',
                            'tamil_name': student_data[11] if len(student_data) > 11 else '',
                            'dob': student_data[12] if len(student_data) > 12 else '',
                            'nationality': student_data[13] if len(student_data) > 13 else '',
                            'blood_group': student_data[14] if len(student_data) > 14 else '',
                            'mother_tongue': student_data[15] if len(student_data) > 15 else '',
                            'caste': student_data[16] if len(student_data) > 16 else '',
                            'religion': student_data[17] if len(student_data) > 17 else '',
                            'place_of_birth': student_data[18] if len(student_data) > 18 else '',
                            'aadhaar': student_data[19] if len(student_data) > 19 else '',
                            'disability': student_data[20] if len(student_data) > 20 else '',
                            'id_mark1': student_data[21] if len(student_data) > 21 else '',
                            'id_mark2': student_data[22] if len(student_data) > 22 else '',
                            'current_class': student_data[23] if len(student_data) > 23 else '',
                            'admission_class': student_data[24] if len(student_data) > 24 else '',
                            'admission_year': student_data[25] if len(student_data) > 25 else '',
                            'admission_date': student_data[26] if len(student_data) > 26 else '',
                            'address': student_data[27] if len(student_data) > 27 else '',
                            'contact': student_data[28] if len(student_data) > 28 else '',
                            'alt_contact': student_data[29] if len(student_data) > 29 else '',
                            'country': student_data[30] if len(student_data) > 30 else '',
                            'state': student_data[31] if len(student_data) > 31 else '',
                            'city': student_data[32] if len(student_data) > 32 else '',
                            'pincode': student_data[33] if len(student_data) > 33 else '',
                            'status': student_data[34] if len(student_data) > 34 else '',
                            'house': student_data[35] if len(student_data) > 35 else '',
                            'teacher_ward': teacher_ward_disp,
                            'rte': rte_disp,
                            'sports_quota': sports_quota_disp,
                            'prev_school': student_data[39] if len(student_data) > 39 else '',
                            'prev_board': student_data[40] if len(student_data) > 40 else '',
                            'father_name': student_data[41] if len(student_data) > 41 else '',
                            'father_name_tamil': student_data[42] if len(student_data) > 42 else '',
                            'mother_name': student_data[43] if len(student_data) > 43 else '',
                            'mother_name_tamil': student_data[44] if len(student_data) > 44 else '',
                            'father_contact': student_data[45] if len(student_data) > 45 else '',
                            'mother_contact': student_data[46] if len(student_data) > 46 else '',
                            'father_email': student_data[47] if len(student_data) > 47 else '',
                            'mother_email': student_data[48] if len(student_data) > 48 else '',
                            'father_qualification': student_data[49] if len(student_data) > 49 else '',
                            'mother_qualification': student_data[50] if len(student_data) > 50 else '',
                            'father_occupation': student_data[51] if len(student_data) > 51 else '',
                            'mother_occupation': student_data[52] if len(student_data) > 52 else '',
                            'father_income': student_data[53] if len(student_data) > 53 else '',
                            'mother_income': student_data[54] if len(student_data) > 54 else '',
                            'guardian_name': student_data[55] if len(student_data) > 55 else '',
                            'guardian_contact': student_data[56] if len(student_data) > 56 else '',
                            'guardian_email': student_data[57] if len(student_data) > 57 else '',
                            'child_living': student_data[58] if len(student_data) > 58 else '',
                            'rights_on_child': student_data[59] if len(student_data) > 59 else '',
                            'med_blood_group': student_data[60] if len(student_data) > 60 else '',
                            'diseases': student_data[61] if len(student_data) > 61 else '',
                            'allergies': student_data[62] if len(student_data) > 62 else '',
                            'medicines': student_data[63] if len(student_data) > 63 else '',
                            'hospital': student_data[64] if len(student_data) > 64 else '',
                            'doctor': student_data[65] if len(student_data) > 65 else '',
                            'gender_options': ['Male', 'Female', 'Other'],
                            'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                            'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                            'teacher_ward_options': ['yes', 'no'],
                            'rte_options': ['yes', 'no'],
                            'sports_quota_options': ['yes', 'no']
                        }
                        return render(request, 'users/add_update_student.html', context)

                # 1. Update student_page1 (basic info)
                cursor.execute("""
                    UPDATE student_page1
                    SET name = %s, admission_number = %s, class = %s, section = %s, 
                        roll_number = %s, emis = %s
                    WHERE admission_number = %s
                """, [name, new_admission_number, class_part, section, roll_number, emis, admission_number])

                # 2. Update student_page2 (personal info)
                cursor.execute("""
                    UPDATE student_page2
                    SET 
                        gender = %s, community = %s, tamil_name = %s, dob = %s,
                        nationality = %s, blood_group = %s, mother_tongue = %s,
                        caste = %s, religion = %s, place_of_birth = %s, aadhaar = %s,
                        disability = %s, id_mark1 = %s, id_mark2 = %s,
                        current_class = %s, admission_class = %s, admission_year = %s,
                        admission_date = %s
                    WHERE user_id = %s
                """, [
                    request.POST.get('gender', ''),
                    request.POST.get('community', ''),
                    request.POST.get('tamil_name', ''),
                    request.POST.get('dob', ''),
                    request.POST.get('nationality', ''),
                    request.POST.get('blood_group', ''),
                    request.POST.get('mother_tongue', ''),
                    request.POST.get('caste', ''),
                    request.POST.get('religion', ''),
                    request.POST.get('place_of_birth', ''),
                    request.POST.get('aadhaar', ''),
                    request.POST.get('disability', ''),
                    request.POST.get('id_mark1', ''),
                    request.POST.get('id_mark2', ''),
                    class_part,  # current_class
                    request.POST.get('admission_class', class_part),
                    request.POST.get('admission_year', ''),
                    request.POST.get('admission_date', ''),
                    user_id
                ])

                # 3. Update student_page3 (contact info)
                cursor.execute("""
                    UPDATE student_page3
                    SET 
                        email = %s, address = %s, contact = %s, alt_contact = %s,
                        country = %s, state = %s, city = %s, pincode = %s,
                        status = %s, house = %s, teacher_ward = %s, rte = %s,
                        sports_quota = %s, prev_school = %s, prev_board = %s
                    WHERE user_id = %s
                """, [
                    email,
                    request.POST.get('address', ''),
                    request.POST.get('contact', ''),
                    request.POST.get('alt_contact', ''),
                    request.POST.get('country', ''),
                    request.POST.get('state', ''),
                    request.POST.get('city', ''),
                    request.POST.get('pincode', ''),
                    request.POST.get('status', ''),
                    request.POST.get('house', ''),
                    teacher_ward,
                    rte,
                    sports_quota,
                    request.POST.get('prev_school', ''),
                    request.POST.get('prev_board', ''),
                    user_id
                ])

                # 4. Update student_page4 (family info)
                cursor.execute("""
                    UPDATE student_page4
                    SET 
                        father_name = %s, father_name_tamil = %s, mother_name = %s,
                        mother_name_tamil = %s, father_contact = %s, mother_contact = %s,
                        father_email = %s, mother_email = %s, father_qualification = %s,
                        mother_qualification = %s, father_occupation = %s,
                        mother_occupation = %s, father_income = %s, mother_income = %s,
                        guardian_name = %s, guardian_contact = %s, guardian_email = %s,
                        child_living = %s, rights_on_child = %s, med_blood_group = %s,
                        diseases = %s, allergies = %s, medicines = %s, hospital = %s,
                        doctor = %s
                    WHERE user_id = %s
                """, [
                    request.POST.get('father_name', ''),
                    request.POST.get('father_name_tamil', ''),
                    request.POST.get('mother_name', ''),
                    request.POST.get('mother_name_tamil', ''),
                    request.POST.get('father_contact', ''),
                    request.POST.get('mother_contact', ''),
                    request.POST.get('father_email', ''),
                    request.POST.get('mother_email', ''),
                    request.POST.get('father_qualification', ''),
                    request.POST.get('mother_qualification', ''),
                    request.POST.get('father_occupation', ''),
                    request.POST.get('mother_occupation', ''),
                    request.POST.get('father_income', ''),
                    request.POST.get('mother_income', ''),
                    request.POST.get('guardian_name', ''),
                    request.POST.get('guardian_contact', ''),
                    request.POST.get('guardian_email', ''),
                    request.POST.get('child_living', ''),
                    request.POST.get('rights_on_child', ''),
                    request.POST.get('med_blood_group', ''),
                    request.POST.get('diseases', ''),
                    request.POST.get('allergies', ''),
                    request.POST.get('medicines', ''),
                    request.POST.get('hospital', ''),
                    request.POST.get('doctor', ''),
                    user_id
                ])

                # Update username if admission number changed
                if new_admission_number != admission_number:
                    new_username = f"student_{new_admission_number}"
                    cursor.execute("""
                        UPDATE users
                        SET username = %s
                        WHERE id = %s
                    """, [new_username, user_id])

            messages.success(request, f'Student {name} updated successfully.')
            return redirect('student_info')

        except Exception as e:
            messages.error(request, f'Error updating student: {str(e)}')
            # Inline context prep for error case (same as above)
            post_data = dict(request.POST)
            teacher_ward_db = student_data[36] if len(student_data) > 36 else None
            teacher_ward = 'yes' if teacher_ward_db in (1, 'yes', 'Yes', True) else 'no'
            rte_db = student_data[37] if len(student_data) > 37 else None
            rte = 'yes' if rte_db in (1, 'yes', 'Yes', True) else 'no'
            sports_quota_db = student_data[38] if len(student_data) > 38 else None
            sports_quota = 'yes' if sports_quota_db in (1, 'yes', 'Yes', True) else 'no'
            context = {
                'title': 'Update Student',
                'admission_number': admission_number,
                'post_data': post_data,
                'id': student_data[0] if len(student_data) > 0 else '',
                'user_id': student_data[1] if len(student_data) > 1 else '',
                'name': student_data[2] if len(student_data) > 2 else '',
                'class_section': f"{student_data[4]}-{student_data[5]}" if len(student_data) > 5 else '',
                'roll_number': student_data[6] if len(student_data) > 6 else '',
                'emis': student_data[7] if len(student_data) > 7 else '',
                'email': student_data[8] if len(student_data) > 8 else '',
                'gender': student_data[9] if len(student_data) > 9 else '',
                'community': student_data[10] if len(student_data) > 10 else '',
                'tamil_name': student_data[11] if len(student_data) > 11 else '',
                'dob': student_data[12] if len(student_data) > 12 else '',
                'nationality': student_data[13] if len(student_data) > 13 else '',
                'blood_group': student_data[14] if len(student_data) > 14 else '',
                'mother_tongue': student_data[15] if len(student_data) > 15 else '',
                'caste': student_data[16] if len(student_data) > 16 else '',
                'religion': student_data[17] if len(student_data) > 17 else '',
                'place_of_birth': student_data[18] if len(student_data) > 18 else '',
                'aadhaar': student_data[19] if len(student_data) > 19 else '',
                'disability': student_data[20] if len(student_data) > 20 else '',
                'id_mark1': student_data[21] if len(student_data) > 21 else '',
                'id_mark2': student_data[22] if len(student_data) > 22 else '',
                'current_class': student_data[23] if len(student_data) > 23 else '',
                'admission_class': student_data[24] if len(student_data) > 24 else '',
                'admission_year': student_data[25] if len(student_data) > 25 else '',
                'admission_date': student_data[26] if len(student_data) > 26 else '',
                'address': student_data[27] if len(student_data) > 27 else '',
                'contact': student_data[28] if len(student_data) > 28 else '',
                'alt_contact': student_data[29] if len(student_data) > 29 else '',
                'country': student_data[30] if len(student_data) > 30 else '',
                'state': student_data[31] if len(student_data) > 31 else '',
                'city': student_data[32] if len(student_data) > 32 else '',
                'pincode': student_data[33] if len(student_data) > 33 else '',
                'status': student_data[34] if len(student_data) > 34 else '',
                'house': student_data[35] if len(student_data) > 35 else '',
                'teacher_ward': teacher_ward,
                'rte': rte,
                'sports_quota': sports_quota,
                'prev_school': student_data[39] if len(student_data) > 39 else '',
                'prev_board': student_data[40] if len(student_data) > 40 else '',
                'father_name': student_data[41] if len(student_data) > 41 else '',
                'father_name_tamil': student_data[42] if len(student_data) > 42 else '',
                'mother_name': student_data[43] if len(student_data) > 43 else '',
                'mother_name_tamil': student_data[44] if len(student_data) > 44 else '',
                'father_contact': student_data[45] if len(student_data) > 45 else '',
                'mother_contact': student_data[46] if len(student_data) > 46 else '',
                'father_email': student_data[47] if len(student_data) > 47 else '',
                'mother_email': student_data[48] if len(student_data) > 48 else '',
                'father_qualification': student_data[49] if len(student_data) > 49 else '',
                'mother_qualification': student_data[50] if len(student_data) > 50 else '',
                'father_occupation': student_data[51] if len(student_data) > 51 else '',
                'mother_occupation': student_data[52] if len(student_data) > 52 else '',
                'father_income': student_data[53] if len(student_data) > 53 else '',
                'mother_income': student_data[54] if len(student_data) > 54 else '',
                'guardian_name': student_data[55] if len(student_data) > 55 else '',
                'guardian_contact': student_data[56] if len(student_data) > 56 else '',
                'guardian_email': student_data[57] if len(student_data) > 57 else '',
                'child_living': student_data[58] if len(student_data) > 58 else '',
                'rights_on_child': student_data[59] if len(student_data) > 59 else '',
                'med_blood_group': student_data[60] if len(student_data) > 60 else '',
                'diseases': student_data[61] if len(student_data) > 61 else '',
                'allergies': student_data[62] if len(student_data) > 62 else '',
                'medicines': student_data[63] if len(student_data) > 63 else '',
                'hospital': student_data[64] if len(student_data) > 64 else '',
                'doctor': student_data[65] if len(student_data) > 65 else '',
                'gender_options': ['Male', 'Female', 'Other'],
                'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
                'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
                'teacher_ward_options': ['yes', 'no'],
                'rte_options': ['yes', 'no'],
                'sports_quota_options': ['yes', 'no']
            }
            return render(request, 'users/add_update_student.html', context)

    # For GET request or after successful POST, prepare context (inline for safety)
    post_data = None  # No post_data for GET
    # Map boolean fields from DB robustly (handles int 0/1 or str 'yes'/'no')
    teacher_ward_db = student_data[36] if len(student_data) > 36 else None
    teacher_ward = 'yes' if teacher_ward_db in (1, 'yes', 'Yes', True) else 'no'
    rte_db = student_data[37] if len(student_data) > 37 else None
    rte = 'yes' if rte_db in (1, 'yes', 'Yes', True) else 'no'
    sports_quota_db = student_data[38] if len(student_data) > 38 else None
    sports_quota = 'yes' if sports_quota_db in (1, 'yes', 'Yes', True) else 'no'
    context = {
        'title': 'Update Student',
        'admission_number': admission_number,
        'post_data': post_data,  # For repopulating form on errors
        'id': student_data[0] if len(student_data) > 0 else '',
        'user_id': student_data[1] if len(student_data) > 1 else '',
        'name': student_data[2] if len(student_data) > 2 else '',
        'class_section': f"{student_data[4]}-{student_data[5]}" if len(student_data) > 5 else '',
        'roll_number': student_data[6] if len(student_data) > 6 else '',
        'emis': student_data[7] if len(student_data) > 7 else '',
        'email': student_data[8] if len(student_data) > 8 else '',
        'gender': student_data[9] if len(student_data) > 9 else '',
        'community': student_data[10] if len(student_data) > 10 else '',
        'tamil_name': student_data[11] if len(student_data) > 11 else '',
        'dob': student_data[12] if len(student_data) > 12 else '',
        'nationality': student_data[13] if len(student_data) > 13 else '',
        'blood_group': student_data[14] if len(student_data) > 14 else '',
        'mother_tongue': student_data[15] if len(student_data) > 15 else '',
        'caste': student_data[16] if len(student_data) > 16 else '',
        'religion': student_data[17] if len(student_data) > 17 else '',
        'place_of_birth': student_data[18] if len(student_data) > 18 else '',
        'aadhaar': student_data[19] if len(student_data) > 19 else '',
        'disability': student_data[20] if len(student_data) > 20 else '',
        'id_mark1': student_data[21] if len(student_data) > 21 else '',
        'id_mark2': student_data[22] if len(student_data) > 22 else '',
        'current_class': student_data[23] if len(student_data) > 23 else '',
        'admission_class': student_data[24] if len(student_data) > 24 else '',
        'admission_year': student_data[25] if len(student_data) > 25 else '',
        'admission_date': student_data[26] if len(student_data) > 26 else '',
        'address': student_data[27] if len(student_data) > 27 else '',
        'contact': student_data[28] if len(student_data) > 28 else '',
        'alt_contact': student_data[29] if len(student_data) > 29 else '',
        'country': student_data[30] if len(student_data) > 30 else '',
        'state': student_data[31] if len(student_data) > 31 else '',
        'city': student_data[32] if len(student_data) > 32 else '',
        'pincode': student_data[33] if len(student_data) > 33 else '',
        'status': student_data[34] if len(student_data) > 34 else '',
        'house': student_data[35] if len(student_data) > 35 else '',
        'teacher_ward': teacher_ward,
        'rte': rte,
        'sports_quota': sports_quota,
        'prev_school': student_data[39] if len(student_data) > 39 else '',
        'prev_board': student_data[40] if len(student_data) > 40 else '',
        'father_name': student_data[41] if len(student_data) > 41 else '',
        'father_name_tamil': student_data[42] if len(student_data) > 42 else '',
        'mother_name': student_data[43] if len(student_data) > 43 else '',
        'mother_name_tamil': student_data[44] if len(student_data) > 44 else '',
        'father_contact': student_data[45] if len(student_data) > 45 else '',
        'mother_contact': student_data[46] if len(student_data) > 46 else '',
        'father_email': student_data[47] if len(student_data) > 47 else '',
        'mother_email': student_data[48] if len(student_data) > 48 else '',
        'father_qualification': student_data[49] if len(student_data) > 49 else '',
        'mother_qualification': student_data[50] if len(student_data) > 50 else '',
        'father_occupation': student_data[51] if len(student_data) > 51 else '',
        'mother_occupation': student_data[52] if len(student_data) > 52 else '',
        'father_income': student_data[53] if len(student_data) > 53 else '',
        'mother_income': student_data[54] if len(student_data) > 54 else '',
        'guardian_name': student_data[55] if len(student_data) > 55 else '',
        'guardian_contact': student_data[56] if len(student_data) > 56 else '',
        'guardian_email': student_data[57] if len(student_data) > 57 else '',
        'child_living': student_data[58] if len(student_data) > 58 else '',
        'rights_on_child': student_data[59] if len(student_data) > 59 else '',
        'med_blood_group': student_data[60] if len(student_data) > 60 else '',
        'diseases': student_data[61] if len(student_data) > 61 else '',
        'allergies': student_data[62] if len(student_data) > 62 else '',
        'medicines': student_data[63] if len(student_data) > 63 else '',
        'hospital': student_data[64] if len(student_data) > 64 else '',
        'doctor': student_data[65] if len(student_data) > 65 else '',
        'gender_options': ['Male', 'Female', 'Other'],
        'community_options': ['General', 'OBC', 'SC', 'ST', 'Other'],
        'blood_group_options': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'Unknown'],
        'teacher_ward_options': ['yes', 'no'],
        'rte_options': ['yes', 'no'],
        'sports_quota_options': ['yes', 'no']
    }
    return render(request, 'users/add_update_student.html', context)

import os
import re
from django.conf import settings
from django.db import connection, transaction
from django.contrib import messages
from django.shortcuts import redirect
from django.db import IntegrityError

def delete_student(request, admission_number):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Get student info
                cursor.execute("""
                    SELECT user_id, name 
                    FROM student_page1 
                    WHERE admission_number = %s
                """, [admission_number])
                student_info = cursor.fetchone()
                
                if not student_info:
                    messages.error(request, 'Student not found.')
                    return redirect('student_info')

                user_id, student_name = student_info

                # Get profile picture path for file deletion
                cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
                profile_pic = cursor.fetchone()
                if profile_pic:
                    image_path = profile_pic[0]
                    # Assuming image_path is relative to MEDIA_ROOT; adjust if absolute
                    full_image_path = os.path.join(settings.MEDIA_ROOT, image_path)
                    if os.path.exists(full_image_path):
                        os.remove(full_image_path)

                # Delete from all child tables referencing student_page1.user_id
                cursor.execute("DELETE FROM admin_attendance WHERE student_id = %s", [user_id])
                cursor.execute("DELETE FROM attendance WHERE student_id = %s", [user_id])
                cursor.execute("DELETE FROM school_marks WHERE student_id = %s", [user_id])
                cursor.execute("DELETE FROM student_page2 WHERE user_id = %s", [user_id])
                cursor.execute("DELETE FROM student_page4 WHERE user_id = %s", [user_id])
                cursor.execute("DELETE FROM profile_pics WHERE user_id = %s", [user_id])
                cursor.execute("DELETE FROM homework WHERE user_id = %s", [user_id])
                cursor.execute("DELETE FROM leave_requests WHERE requested_by = %s", [user_id])
                cursor.execute("DELETE FROM student_leave_requests WHERE user_id = %s", [user_id])
                
                # Delete from student_page1
                cursor.execute("DELETE FROM student_page1 WHERE admission_number = %s", [admission_number])
                
                # Delete from users table
                cursor.execute("DELETE FROM users WHERE id = %s", [user_id])

        messages.success(request, f'Student {student_name} deleted successfully.')
    except IntegrityError as e:
        error_code, error_message = e.args
        if error_code == 1451:
            # Parse the error message to extract the table name
            table_match = re.search(r"constraint fails \(`[^`]*\.`([^`]+)`", error_message)
            if table_match:
                table_name = table_match.group(1)
                messages.error(request, f'Cannot delete student {student_name} due to related records in table `{table_name}`. Please delete or update records in this table first.')
            else:
                messages.error(request, f'Cannot delete student {student_name} due to related records in an unknown table. Contact the administrator. Full error: {error_message}')
        else:
            messages.error(request, f'Error deleting student: {str(e)}')
    except Exception as e:
        messages.error(request, f'Unexpected error deleting student: {str(e)}')
    
    return redirect('student_info')



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def add_batch(request):
    # Check admin authentication
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        academic_year = request.POST.get('academic_year')
        
        if not academic_year:
            messages.error(request, 'Please select an academic year')
            return redirect('view_batches')
        
        try:
            with connection.cursor() as cursor:
                # Check if batch already exists
                cursor.execute(
                    "SELECT id FROM admin_student_batch WHERE academic_year = %s",
                    [academic_year]
                )
                if cursor.fetchone():
                    messages.error(request, 'This academic year already exists')
                    return redirect('view_batches')
                
                # Insert new batch
                cursor.execute(
                    "INSERT INTO admin_student_batch (academic_year, created_at) VALUES (%s, NOW())",
                    [academic_year]
                )
                
                messages.success(request, 'Academic year added successfully')
                return redirect('view_batches')
                
        except Exception as e:
            messages.error(request, f'Error adding academic year: {str(e)}')
            return redirect('view_batches')
    
    return redirect('view_batches')

def view_batches(request):
    # Check admin authentication
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    try:
        with connection.cursor() as cursor:
            # Get all batches ordered by academic_year in descending order
            cursor.execute(
                "SELECT id, academic_year FROM admin_student_batch ORDER BY academic_year DESC"
            )
            batches = cursor.fetchall()
            
            # Convert to list of dictionaries for easier template handling
            batch_list = [{'id': row[0], 'academic_year': row[1]} for row in batches]
            
        return render(request, 'users/view_batches.html', {
            'batches': batch_list
        })
        
    except Exception as e:
        messages.error(request, f'Error fetching batches: {str(e)}')
        return render(request, 'users/view_batches.html', {
            'batches': []
        })

def update_batch(request, batch_id):
    # Check admin authentication
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        new_academic_year = request.POST.get('academic_year')
        
        if not new_academic_year:
            messages.error(request, 'Please select an academic year')
            return redirect('view_batches')
        
        try:
            with connection.cursor() as cursor:
                # Check if the new academic year already exists (excluding the current batch)
                cursor.execute(
                    "SELECT id FROM admin_student_batch WHERE academic_year = %s AND id != %s",
                    [new_academic_year, batch_id]
                )
                if cursor.fetchone():
                    messages.error(request, 'This academic year already exists')
                    return redirect('view_batches')
                
                # Check if batch exists
                cursor.execute(
                    "SELECT id FROM admin_student_batch WHERE id = %s",
                    [batch_id]
                )
                if not cursor.fetchone():
                    messages.error(request, 'Batch not found')
                    return redirect('view_batches')
                
                # Update the batch
                cursor.execute(
                    "UPDATE admin_student_batch SET academic_year = %s WHERE id = %s",
                    [new_academic_year, batch_id]
                )
                
                messages.success(request, 'Batch updated successfully')
                return redirect('view_batches')
                
        except Exception as e:
            messages.error(request, f'Error updating batch: {str(e)}')
            return redirect('view_batches')
    
    return redirect('view_batches')

def delete_batch(request, batch_id):
    # Check admin authentication
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    try:
        with connection.cursor() as cursor:
            # Check if batch exists
            cursor.execute(
                "SELECT id FROM admin_student_batch WHERE id = %s",
                [batch_id]
            )
            if not cursor.fetchone():
                messages.error(request, 'Batch not found')
                return redirect('view_batches')
            
            # Delete the batch
            cursor.execute(
                "DELETE FROM admin_student_batch WHERE id = %s",
                [batch_id]
            )
            
            messages.success(request, 'Batch deleted successfully')
            return redirect('view_batches')
            
    except Exception as e:
        messages.error(request, f'Error deleting batch: {str(e)}')
        return redirect('view_batches')






from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, transaction
from django.conf import settings
import os
import uuid

def manage_users(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, name, email, username, password, role, created_at, updated_at
                FROM admin_manage_users
                ORDER BY name
            """)
            users = cursor.fetchall()
            
            user_list = []
            for row in users:
                user_id, name, email, username, password, role, created_at, updated_at = row
                # Fetch profile picture based on role
                profile_pic_url = None
                if role == 'teacher':
                    cursor.execute("SELECT profile_pic_url FROM profile_pics_teachers WHERE teacher_id = %s", [user_id])
                    pic_result = cursor.fetchone()
                    profile_pic_url = f"{settings.MEDIA_URL}{pic_result[0]}" if pic_result else f"{settings.MEDIA_URL}pfpicsteacher/default.jpg"
                else:
                    cursor.execute("SELECT profile_pic_url FROM otherusers_profile_pic WHERE user_id = %s", [user_id])
                    pic_result = cursor.fetchone()
                    profile_pic_url = f"{settings.MEDIA_URL}{pic_result[0]}" if pic_result else f"{settings.MEDIA_URL}pfpicsusers/default.jpg"
                
                user_list.append({
                    'id': user_id,
                    'name': name,
                    'email': email,
                    'username': username,
                    'password': password,
                    'role': role,
                    'created_at': created_at,
                    'updated_at': updated_at,
                    'profile_pic_url': profile_pic_url
                })
            
        return render(request, 'users/manage_users.html', {
            'users': user_list
        })
        
    except Exception as e:
        messages.error(request, f'Error fetching users: {str(e)}')
        return render(request, 'users/manage_users.html', {
            'users': []
        })

def add_user(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        
        if not all([name, email, username, password, role]):
            messages.error(request, 'All fields are required')
            return redirect('manage_users')
        
        try:
            with transaction.atomic():
                with connection.cursor() as cursor:
                    # Check for duplicate username or email
                    cursor.execute(
                        "SELECT id FROM admin_manage_users WHERE username = %s OR email = %s",
                        [username, email]
                    )
                    if cursor.fetchone():
                        messages.error(request, 'Username or email already exists')
                        return redirect('manage_users')
                    
                    # Insert into admin_manage_users
                    cursor.execute(
                        """INSERT INTO admin_manage_users 
                        (name, email, username, password, role, created_at) 
                        VALUES (%s, %s, %s, %s, %s, NOW())""",
                        [name, email, username, password, role]
                    )
                    user_id = cursor.lastrowid

                    # Handle teacher-specific initialization
                    if role == 'teacher':
                        cursor.execute(
                            """INSERT INTO teachers (id, name, email, password, subject, created_at)
                            VALUES (%s, %s, %s, %s, %s, NOW())""",
                            [user_id, name, email, password, '']
                        )

                    # Handle profile picture
                    if 'profile_pic' in request.FILES:
                        profile_pic = request.FILES['profile_pic']
                        allowed_extensions = ['.png', '.jpg', '.jpeg']
                        file_ext = os.path.splitext(profile_pic.name)[1].lower()
                        if file_ext not in allowed_extensions:
                            messages.error(request, "Only PNG, JPG, or JPEG files are allowed.")
                            return redirect('manage_users')
                        if profile_pic.size > 5 * 1024 * 1024:
                            messages.error(request, "File size must be less than 5MB.")
                            return redirect('manage_users')

                        filename = f"{uuid.uuid4().hex}_{user_id}{file_ext}"
                        pfpics_dir = os.path.join(settings.MEDIA_ROOT, 'pfpicsteacher' if role == 'teacher' else 'pfpicsusers')
                        os.makedirs(pfpics_dir, exist_ok=True)
                        file_path = os.path.join(pfpics_dir, filename)

                        with open(file_path, 'wb+') as destination:
                            for chunk in profile_pic.chunks():
                                destination.write(chunk)

                        table = 'profile_pics_teachers' if role == 'teacher' else 'otherusers_profile_pic'
                        id_column = 'teacher_id' if role == 'teacher' else 'user_id'
                        pic_path = f"{'pfpicsteacher' if role == 'teacher' else 'pfpicsusers'}/{filename}"
                        cursor.execute(
                            f"INSERT INTO {table} ({id_column}, profile_pic_url, created_at) VALUES (%s, %s, NOW())",
                            [user_id, pic_path]
                        )

                messages.success(request, 'User created successfully')
                return redirect(f'/manage_users/?t={uuid.uuid4().hex}')
                
        except Exception as e:
            messages.error(request, f'Error creating user: {str(e)}')
            return redirect('manage_users')
    
    return redirect('manage_users')

def update_user(request, user_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        
        if not all([name, email, username, role]):
            messages.error(request, 'Name, email, username, and role are required')
            return redirect('manage_users')
        
        try:
            with transaction.atomic():
                with connection.cursor() as cursor:
                    # Check if user exists
                    cursor.execute("SELECT id, role, password FROM admin_manage_users WHERE id = %s", [user_id])
                    user_data = cursor.fetchone()
                    if not user_data:
                        messages.error(request, 'User not found')
                        return redirect('manage_users')
                    
                    current_role = user_data[1]
                    current_password = user_data[2]

                    # Check for duplicate username or email in admin_manage_users
                    cursor.execute(
                        "SELECT id FROM admin_manage_users WHERE (username = %s OR email = %s) AND id != %s",
                        [username, email, user_id]
                    )
                    if cursor.fetchone():
                        messages.error(request, 'Username or email already exists')
                        return redirect('manage_users')

                    # NEW: Check dependencies before changing from teacher role
                    if current_role == 'teacher' and role != 'teacher':
                        # Check if teacher is assigned to any timetable
                        cursor.execute("SELECT COUNT(*) FROM timetable WHERE teacher_id = %s", [user_id])
                        timetable_count = cursor.fetchone()[0]
                        
                        # Check if teacher is assigned as invigilator in exams
                        cursor.execute("SELECT COUNT(*) FROM exams WHERE invigilator_id = %s", [user_id])
                        exam_count = cursor.fetchone()[0]
                        
                        if timetable_count > 0 or exam_count > 0:
                            error_msg = f'Cannot change role from Teacher. This teacher is assigned to '
                            if timetable_count > 0:
                                error_msg += f'{timetable_count} timetable slot(s)'
                            if timetable_count > 0 and exam_count > 0:
                                error_msg += ' and '
                            if exam_count > 0:
                                error_msg += f'{exam_count} exam(s) as invigilator'
                            error_msg += '. Please remove these assignments first.'
                            messages.error(request, error_msg)
                            return redirect('manage_users')

                    # Update admin_manage_users
                    new_password = password if password else current_password
                    cursor.execute(
                        """UPDATE admin_manage_users 
                        SET name = %s, email = %s, username = %s, password = %s, role = %s, updated_at = NOW()
                        WHERE id = %s""",
                        [name, email, username, new_password, role, user_id]
                    )

                    # Handle role changes
                    if current_role == 'teacher' and role != 'teacher':
                        # Safe to delete now - we've already checked for dependencies above
                        # Delete teacher record and profile picture
                        cursor.execute("SELECT profile_pic_url FROM profile_pics_teachers WHERE teacher_id = %s", [user_id])
                        old_pic = cursor.fetchone()
                        if old_pic:
                            old_file_path = os.path.join(settings.MEDIA_ROOT, old_pic[0])
                            if os.path.exists(old_file_path):
                                os.remove(old_file_path)
                            cursor.execute("DELETE FROM profile_pics_teachers WHERE teacher_id = %s", [user_id])
                        cursor.execute("DELETE FROM teachers WHERE id = %s", [user_id])

                    elif role == 'teacher':
                        # Update or insert teacher record
                        cursor.execute("SELECT id FROM teachers WHERE id = %s", [user_id])
                        teacher_exists = cursor.fetchone()
                        
                        try:
                            if not teacher_exists:
                                # Inserting new teacher record
                                cursor.execute(
                                    """INSERT INTO teachers (id, name, email, password, subject, created_at)
                                    VALUES (%s, %s, %s, %s, %s, NOW())""",
                                    [user_id, name, email, new_password, '']
                                )
                            else:
                                # Updating existing teacher record
                                cursor.execute(
                                    """UPDATE teachers 
                                    SET name = %s, email = %s, password = %s
                                    WHERE id = %s""",
                                    [name, email, new_password, user_id]
                                )
                        except Exception as teacher_error:
                            # Handle duplicate email error in teachers table
                            if '1062' in str(teacher_error) or 'Duplicate entry' in str(teacher_error):
                                messages.error(request, f'Email {email} is already used by another teacher. Please use a different email.')
                                return redirect('manage_users')
                            else:
                                raise teacher_error

                    # Handle profile picture
                    if 'profile_pic' in request.FILES:
                        profile_pic = request.FILES['profile_pic']
                        allowed_extensions = ['.png', '.jpg', '.jpeg']
                        file_ext = os.path.splitext(profile_pic.name)[1].lower()
                        if file_ext not in allowed_extensions:
                            messages.error(request, "Only PNG, JPG, or JPEG files are allowed.")
                            return redirect('manage_users')
                        if profile_pic.size > 5 * 1024 * 1024:
                            messages.error(request, "File size must be less than 5MB.")
                            return redirect('manage_users')

                        filename = f"{uuid.uuid4().hex}_{user_id}{file_ext}"
                        pfpics_dir = os.path.join(settings.MEDIA_ROOT, 'pfpicsteacher' if role == 'teacher' else 'pfpicsusers')
                        os.makedirs(pfpics_dir, exist_ok=True)
                        file_path = os.path.join(pfpics_dir, filename)

                        # Delete old profile picture
                        table = 'profile_pics_teachers' if role == 'teacher' else 'otherusers_profile_pic'
                        id_column = 'teacher_id' if role == 'teacher' else 'user_id'
                        cursor.execute(f"SELECT profile_pic_url FROM {table} WHERE {id_column} = %s", [user_id])
                        old_pic = cursor.fetchone()
                        if old_pic:
                            old_file_path = os.path.join(settings.MEDIA_ROOT, old_pic[0])
                            if os.path.exists(old_file_path):
                                os.remove(old_file_path)
                            cursor.execute(f"DELETE FROM {table} WHERE {id_column} = %s", [user_id])

                        with open(file_path, 'wb+') as destination:
                            for chunk in profile_pic.chunks():
                                destination.write(chunk)

                        pic_path = f"{'pfpicsteacher' if role == 'teacher' else 'pfpicsusers'}/{filename}"
                        cursor.execute(
                            f"INSERT INTO {table} ({id_column}, profile_pic_url, created_at) VALUES (%s, %s, NOW())",
                            [user_id, pic_path]
                        )

                messages.success(request, 'User updated successfully')
                return redirect(f'/manage_users/?t={uuid.uuid4().hex}')

        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')
            return redirect('manage_users')
    
    return redirect('manage_users')

def delete_user(request, user_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Check user and role
                cursor.execute("SELECT role, name FROM admin_manage_users WHERE id = %s", [user_id])
                user_data = cursor.fetchone()
                if not user_data:
                    messages.error(request, 'User not found.')
                    return redirect('manage_users')
                role, name = user_data

                # NEW: Check dependencies before deleting teacher
                if role == 'teacher':
                    # Check if teacher is assigned to any timetable
                    cursor.execute("SELECT COUNT(*) FROM timetable WHERE teacher_id = %s", [user_id])
                    timetable_count = cursor.fetchone()[0]
                    
                    # Check if teacher is assigned as invigilator in exams
                    cursor.execute("SELECT COUNT(*) FROM exams WHERE invigilator_id = %s", [user_id])
                    exam_count = cursor.fetchone()[0]
                    
                    if timetable_count > 0 or exam_count > 0:
                        error_msg = f'Cannot delete teacher "{name}". This teacher is assigned to '
                        if timetable_count > 0:
                            error_msg += f'{timetable_count} timetable slot(s)'
                        if timetable_count > 0 and exam_count > 0:
                            error_msg += ' and '
                        if exam_count > 0:
                            error_msg += f'{exam_count} exam(s) as invigilator'
                        error_msg += '. Please remove these assignments first.'
                        messages.error(request, error_msg)
                        return redirect('manage_users')

                # Delete profile picture
                table = 'profile_pics_teachers' if role == 'teacher' else 'otherusers_profile_pic'
                id_column = 'teacher_id' if role == 'teacher' else 'user_id'
                cursor.execute(f"SELECT profile_pic_url FROM {table} WHERE {id_column} = %s", [user_id])
                old_pic = cursor.fetchone()
                if old_pic:
                    old_file_path = os.path.join(settings.MEDIA_ROOT, old_pic[0])
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                    cursor.execute(f"DELETE FROM {table} WHERE {id_column} = %s", [user_id])

                # Delete from teachers if applicable
                if role == 'teacher':
                    # Safe to delete now - we've already checked for dependencies
                    cursor.execute("DELETE FROM teachers WHERE id = %s", [user_id])

                # Delete from admin_manage_users
                cursor.execute("DELETE FROM admin_manage_users WHERE id = %s", [user_id])
                if cursor.rowcount == 0:
                    messages.error(request, 'User not found.')
                    return redirect('manage_users')

        messages.success(request, f'User "{name}" deleted successfully!')
        return redirect(f'/manage_users/?t={uuid.uuid4().hex}')

    except Exception as e:
        messages.error(request, f'Error deleting user: {str(e)}')
        return redirect('manage_users')




from django.shortcuts import render, redirect

def admin_page(request):
    admin_name = request.session.get('admin_name')
    if not admin_name:
        return redirect('admin_login')  # If not logged in, redirect to login

    return render(request, 'users/admin_page.html', {'admin_name': admin_name})



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def view_edit_class(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    # Fetch all class-section pairs from admin_student_classes for current admin
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, class, section 
            FROM admin_student_classes 
            WHERE admin_id = %s
            ORDER BY class DESC, section DESC
        """, [admin_id])
        classes = cursor.fetchall()

    # Format as "class-section" for display
    class_list = [{'id': row[0], 'class_name': f"{row[1]}-{row[2]}"} for row in classes]

    return render(request, 'users/view_edit_class.html', {
        'classes': class_list,
        'total_classes': len(class_list)
    })

def add_class(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    if request.method == 'POST':
        class_name = request.POST.get('class_name')
        class_number = request.POST.get('class_number', '').strip()
        section_name = request.POST.get('section_name', '').strip()
        
        # Fallback: Combine if class_name is empty but components are provided
        if not class_name and class_number and section_name:
            class_name = f"{class_number}-{section_name.upper()}"
        
        if class_name:
            try:
                class_part, section = class_name.split('-')
                class_part = class_part.strip()
                section = section.strip().upper()
                
                with connection.cursor() as cursor:
                    # Check if this exact class-section combo already exists for this admin
                    cursor.execute("""
                        SELECT COUNT(*) FROM admin_student_classes
                        WHERE admin_id = %s AND class = %s AND section = %s
                    """, [admin_id, class_part, section])
                    exists = cursor.fetchone()[0]

                    if exists:
                        messages.error(request, f'You already created class {class_name}.')
                    else:
                        # Insert new class-section for this admin
                        cursor.execute("""
                            INSERT INTO admin_student_classes (admin_id, class, section)
                            VALUES (%s, %s, %s)
                        """, [admin_id, class_part, section])
                        
                        messages.success(request, f'Class {class_name} added successfully.')
                        return redirect('view_edit_class')
                        
            except ValueError:
                messages.error(request, 'Class name must be in format "Class-Section" (e.g., 2-A).')
            except Exception as e:
                messages.error(request, f'Error adding class: {str(e)}')
        else:
            messages.error(request, 'Class name cannot be empty.')
    
    return render(request, 'users/add_update_class.html', {'title': 'Add New Class'})

def update_class(request, class_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    # Fetch the class-section pair from admin_student_classes
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, class, section 
            FROM admin_student_classes 
            WHERE id = %s AND admin_id = %s
        """, [class_id, admin_id])
        class_data = cursor.fetchone()
        if not class_data:
            messages.error(request, 'Class not found or you don\'t have permission.')
            return redirect('view_edit_class')

    if request.method == 'POST':
        new_class_name = request.POST.get('class_name')
        if new_class_name:
            try:
                new_class, new_section = new_class_name.split('-')
                with connection.cursor() as cursor:
                    # Check if the new class-section combo already exists for this admin
                    cursor.execute("""
                        SELECT COUNT(*) FROM admin_student_classes
                        WHERE admin_id = %s AND class = %s AND section = %s AND id != %s
                    """, [admin_id, new_class, new_section, class_id])
                    exists = cursor.fetchone()[0]

                    if exists:
                        messages.error(request, f'You already have class {new_class_name}.')
                    else:
                        # Update the record in admin_student_classes
                        cursor.execute("""
                            UPDATE admin_student_classes
                            SET class = %s, section = %s
                            WHERE id = %s AND admin_id = %s
                        """, [new_class, new_section, class_id, admin_id])
                        messages.success(request, f'Class updated to {new_class_name} successfully.')
                        return redirect('view_edit_class')
            except ValueError:
                messages.error(request, 'Class name must be in format "Class-Section" (e.g., 2-A).')
            except Exception as e:
                messages.error(request, f'Error updating class: {str(e)}')
        else:
            messages.error(request, 'Class name cannot be empty.')
    return render(request, 'users/add_update_class.html', {
        'title': 'Update Class',
        'class_name': f"{class_data[1]}-{class_data[2]}"
    })

def delete_class(request, class_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    admin_id = request.session['admin_id']

    try:
        with connection.cursor() as cursor:
            # Delete the record from admin_student_classes
            cursor.execute("""
                DELETE FROM admin_student_classes 
                WHERE id = %s AND admin_id = %s
            """, [class_id, admin_id])
        messages.success(request, 'Class deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error deleting class: {str(e)}')
    return redirect('view_edit_class')




from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import transaction, connection, IntegrityError
import pandas as pd
import numpy as np
import os
import uuid
from django.conf import settings

def bulk_upload(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        if 'preview' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
                return redirect('bulk_upload')

            if excel_file.size == 0:
                messages.error(request, 'The uploaded file is empty.')
                return redirect('bulk_upload')

            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'temp'))
            filename = f"temp_{uuid.uuid4()}_{excel_file.name}"
            fs.save(filename, excel_file)

            try:
                df = pd.read_excel(fs.path(filename))
                
                if df.empty:
                    messages.error(request, 'The Excel file contains no data.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                expected_columns = [
                    'user_id', 'name', 'admission_number', 'class', 'section', 'roll_number', 'emis',
                    'gender', 'community', 'tamil_name', 'dob', 'nationality', 'blood_group',
                    'mother_tongue', 'caste', 'religion', 'place_of_birth', 'aadhaar',
                    'disability', 'id_mark1', 'id_mark2', 'current_class', 'admission_class',
                    'admission_year', 'admission_date', 'email', 'address', 'contact',
                    'alt_contact', 'country', 'state', 'city', 'pincode', 'status', 'house',
                    'teacher_ward', 'rte', 'sports_quota', 'prev_school', 'prev_board',
                    'father_name', 'father_name_tamil', 'mother_name', 'mother_name_tamil',
                    'father_contact', 'mother_contact', 'father_email', 'mother_email',
                    'father_qualification', 'mother_qualification', 'father_occupation',
                    'mother_occupation', 'father_income', 'mother_income', 'guardian_name',
                    'guardian_contact', 'guardian_email', 'child_living', 'rights_on_child',
                    'med_blood_group', 'diseases', 'allergies', 'medicines', 'hospital', 'doctor'
                ]

                missing_columns = [col for col in expected_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f'Missing columns in Excel file: {", ".join(missing_columns)}')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Convert NaN, pd.NA, and None to None (MySQL NULL)
                df = df.replace([pd.NA, np.nan, None], None)

                # Validate user_id
                if df['user_id'].isna().any():
                    messages.error(request, 'The user_id column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                try:
                    df['user_id'] = df['user_id'].astype(int)
                except (ValueError, TypeError):
                    messages.error(request, 'Invalid data in user_id column. All values must be integers.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                if df['user_id'].duplicated().any():
                    messages.error(request, 'The user_id column contains duplicate values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Validate admission_number
                if df['admission_number'].isna().any():
                    messages.error(request, 'The admission_number column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                try:
                    df['admission_number'] = df['admission_number'].astype(str).str.strip()
                except Exception as e:
                    messages.error(request, f'Invalid data in admission_number column: {e}')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                if df['admission_number'].duplicated().any():
                    messages.error(request, 'The admission_number column contains duplicate values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Validate roll_number
                if df['roll_number'].isna().any():
                    messages.error(request, 'The roll_number column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                try:
                    df['roll_number'] = df['roll_number'].astype(int)
                except (ValueError, TypeError):
                    messages.error(request, 'Invalid data in roll_number column. All values must be integers.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Validate name (for username)
                if df['name'].isna().any():
                    messages.error(request, 'The name column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Convert date fields
                if 'dob' in df.columns:
                    df['dob'] = pd.to_datetime(df['dob'], errors='coerce').dt.strftime('%Y-%m-%d')
                if 'admission_date' in df.columns:
                    df['admission_date'] = pd.to_datetime(df['admission_date'], errors='coerce').dt.strftime('%Y-%m-%d')

                preview_data = df.head(10).to_dict('records')
                preview_columns = df.columns.tolist()

                request.session['temp_excel_file'] = filename

                return render(request, 'users/bulk_upload.html', {
                    'preview_data': preview_data,
                    'preview_columns': preview_columns
                })

            except Exception as e:
                messages.error(request, f'Error processing Excel file: {str(e)}')
                fs.delete(filename)
                return redirect('bulk_upload')

        elif 'upload' in request.POST:
            filename = request.session.get('temp_excel_file')
            if not filename:
                messages.error(request, 'No file selected for upload. Please upload a file first.')
                return redirect('bulk_upload')

            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'temp'))
            try:
                df = pd.read_excel(fs.path(filename))

                # Convert NaN, pd.NA, and None to None (MySQL NULL)
                df = df.replace([pd.NA, np.nan, None], None)

                # Validate user_id
                if df['user_id'].isna().any():
                    messages.error(request, 'The user_id column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                try:
                    df['user_id'] = df['user_id'].astype(int)
                except (ValueError, TypeError):
                    messages.error(request, 'Invalid data in user_id column. All values must be integers.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                if df['user_id'].duplicated().any():
                    messages.error(request, 'The user_id column contains duplicate values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Validate admission_number
                if df['admission_number'].isna().any():
                    messages.error(request, 'The admission_number column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                try:
                    # Convert all admission numbers to strings and strip extra spaces
                    df['admission_number'] = df['admission_number'].astype(str).str.strip()
                except Exception as e:
                    messages.error(request, f'Invalid data in admission_number column: {e}')
                    fs.delete(filename)
                    return redirect('bulk_upload')


                if df['admission_number'].duplicated().any():
                    messages.error(request, 'The admission_number column contains duplicate values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Validate roll_number
                if df['roll_number'].isna().any():
                    messages.error(request, 'The roll_number column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                try:
                    df['roll_number'] = df['roll_number'].astype(int)
                except (ValueError, TypeError):
                    messages.error(request, 'Invalid data in roll_number column. All values must be integers.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Validate name (for username)
                if df['name'].isna().any():
                    messages.error(request, 'The name column contains null or missing values.')
                    fs.delete(filename)
                    return redirect('bulk_upload')

                # Convert date fields
                if 'dob' in df.columns:
                    df['dob'] = pd.to_datetime(df['dob'], errors='coerce').dt.strftime('%Y-%m-%d')
                if 'admission_date' in df.columns:
                    df['admission_date'] = pd.to_datetime(df['admission_date'], errors='coerce').dt.strftime('%Y-%m-%d')

                skipped_rows = []
                with connection.cursor() as cursor:
                    for index, row in df.iterrows():
                        try:
                            with transaction.atomic():
                                user_id = row['user_id']  # Use Excel-provided user_id (integer)
                                base_username = str(row.get('name', ''))[:100]  # Reserve space for suffix
                                contact = str(row.get('contact', '')) if row.get('contact') else None
                                alt_contact = str(row.get('alt_contact', '')) if row.get('alt_contact') else None
                                unique_id = str(uuid.uuid4())[:8]  # Short UUID for uniqueness
                                email = f"{base_username.lower().replace(' ', '.')}_{user_id}_{unique_id}@example.com"[:255]  # Unique email
                                password = str(row.get('roll_number', ''))[:255]  # Use roll_number as password

                                # Check if user_id exists in users
                                cursor.execute("SELECT id FROM users WHERE id = %s", [user_id])
                                user_exists = cursor.fetchone()

                                # Handle users table insertion
                                if not user_exists:
                                    # Check if username exists
                                    username = base_username
                                    cursor.execute("SELECT username, email FROM users WHERE username = %s", [username])
                                    existing_user = cursor.fetchone()

                                    if existing_user:
                                        # If username exists, check phone numbers
                                        cursor.execute("""
                                            SELECT u.id 
                                            FROM users u
                                            JOIN student_page3 sp3 ON u.id = sp3.user_id
                                            WHERE u.username = %s AND (sp3.contact = %s OR sp3.alt_contact = %s)
                                        """, [username, contact, alt_contact])
                                        matching_phone = cursor.fetchone()

                                        if matching_phone:
                                            # Same username and phone number; skip this user
                                            print(f"Skipped user_id: {user_id} (duplicate username '{username}' and phone match)")
                                            skipped_rows.append(f"Row {index + 2}: user_id {user_id} (duplicate username and phone)")
                                            continue

                                        # Different phone or no phone match; append user_id to username
                                        username = f"{base_username}_{user_id}"[:150]

                                    # Insert into users
                                    cursor.execute("""
                                        INSERT INTO users (id, username, email, password)
                                        VALUES (%s, %s, %s, %s)
                                    """, [user_id, username, email, password])
                                    print(f"Inserted user_id: {user_id} into users with username: {username}")

                                else:
                                    print(f"Skipped user_id: {user_id} (already exists in users)")

                                # Insert into student_page1 (update name if exists)
                                cursor.execute("""
                                    INSERT INTO student_page1 (
                                        user_id, name, admission_number, class, section, roll_number, emis
                                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                                    ON DUPLICATE KEY UPDATE name = name
                                """, [
                                    user_id, row.get('name'), row.get('admission_number'), row.get('class'),
                                    row.get('section'), row.get('roll_number'), row.get('emis')
                                ])

                                # Insert into student_page2 (update gender if exists)
                                cursor.execute("""
                                    INSERT INTO student_page2 (
                                        user_id, gender, community, tamil_name, dob, nationality, blood_group,
                                        mother_tongue, caste, religion, place_of_birth, aadhaar, disability,
                                        id_mark1, id_mark2, current_class, admission_class, admission_year,
                                        admission_date
                                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    ON DUPLICATE KEY UPDATE gender = gender
                                """, [
                                    user_id, row.get('gender'), row.get('community'), row.get('tamil_name'), row.get('dob'),
                                    row.get('nationality'), row.get('blood_group'), row.get('mother_tongue'),
                                    row.get('caste'), row.get('religion'), row.get('place_of_birth'), row.get('aadhaar'),
                                    row.get('disability'), row.get('id_mark1'), row.get('id_mark2'), row.get('current_class'),
                                    row.get('admission_class'), row.get('admission_year'), row.get('admission_date')
                                ])

                                # Insert into student_page3 (update email if exists)
                                cursor.execute("""
                                    INSERT INTO student_page3 (
                                        user_id, email, address, contact, alt_contact, country, state, city, pincode,
                                        status, house, teacher_ward, rte, sports_quota, prev_school, prev_board
                                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    ON DUPLICATE KEY UPDATE email = email
                                """, [
                                    user_id, row.get('email'), row.get('address'), row.get('contact'), row.get('alt_contact'),
                                    row.get('country'), row.get('state'), row.get('city'), row.get('pincode'),
                                    row.get('status'), row.get('house'), row.get('teacher_ward'), row.get('rte'),
                                    row.get('sports_quota'), row.get('prev_school'), row.get('prev_board')
                                ])

                                # Insert into student_page4 (update father_name if exists)
                                cursor.execute("""
                                    INSERT INTO student_page4 (
                                        user_id, father_name, father_name_tamil, mother_name, mother_name_tamil,
                                        father_contact, mother_contact, father_email, mother_email, father_qualification,
                                        mother_qualification, father_occupation, mother_occupation, father_income,
                                        mother_income, guardian_name, guardian_contact, guardian_email, child_living,
                                        rights_on_child, med_blood_group, diseases, allergies, medicines, hospital, doctor
                                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    ON DUPLICATE KEY UPDATE father_name = father_name
                                """, [
                                    user_id, row.get('father_name'), row.get('father_name_tamil'), row.get('mother_name'),
                                    row.get('mother_name_tamil'), row.get('father_contact'), row.get('mother_contact'),
                                    row.get('father_email'), row.get('mother_email'), row.get('father_qualification'),
                                    row.get('mother_qualification'), row.get('father_occupation'), row.get('mother_occupation'),
                                    row.get('father_income'), row.get('mother_income'), row.get('guardian_name'),
                                    row.get('guardian_contact'), row.get('guardian_email'), row.get('child_living'),
                                    row.get('rights_on_child'), row.get('med_blood_group'), row.get('diseases'),
                                    row.get('allergies'), row.get('medicines'), row.get('hospital'), row.get('doctor')
                                ])

                        except IntegrityError as e:
                            error_msg = f"Row {index + 2}: user_id {user_id} failed due to {str(e)}"
                            print(error_msg)
                            skipped_rows.append(error_msg)
                            continue  # Continue with the next row

                if skipped_rows:
                    messages.warning(request, f"Some rows were skipped: {'; '.join(skipped_rows)}")
                messages.success(request, 'Data upload completed!')
                fs.delete(filename)
                if 'temp_excel_file' in request.session:
                    del request.session['temp_excel_file']
                return redirect('bulk_upload')

            except Exception as e:
                messages.error(request, f'Error uploading data: {str(e)}')
                fs.delete(filename)
                if 'temp_excel_file' in request.session:
                    del request.session['temp_excel_file']
                return redirect('bulk_upload')

    return render(request, 'users/bulk_upload.html')






# users/views.py



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, transaction, Error
from django.conf import settings
from datetime import datetime


def manage_teachers(request):
    """
    Displays the teacher management page with comprehensive teacher profiles.
    """
    if not request.session.get('admin_id'):
        messages.error(request, 'Please log in to access this page.')
        return redirect('admin_login')

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    t.id, t.name, t.email, t.subject, t.class_teacher_of, t.created_at, t.password,
                    p.profile_pic_url,
                    tp.full_name, tp.gender, tp.date_of_birth, tp.blood_group, tp.nationality,
                    tp.mobile_number, tp.alternate_contact, tp.official_email,
                    tp.residential_address, tp.city_state_pin,
                    tp.emergency_contact_name, tp.emergency_contact_number,
                    tp.designation, tp.department, tp.subjects_taught, tp.classes_assigned,
                    tp.sections, tp.employee_type, tp.employment_status, tp.joining_date,
                    tp.qualification, tp.specialization, tp.board_university, tp.year_of_passing,
                    tp.bed_ctet_tet, tp.special_training, tp.workshops_attended,
                    tp.is_class_teacher, tp.house_club_incharge, 
                    tp.cocurricular_responsibilities, tp.exam_duties
                FROM teachers t
                LEFT JOIN profile_pics_teachers p ON t.id = p.teacher_id
                LEFT JOIN teacher_profiles tp ON t.id = tp.teacher_id
                ORDER BY t.created_at DESC
            """)
            
            teachers = []
            for row in cursor.fetchall():
                teachers.append({
                    # Basic teacher info
                    'id': row[0],
                    'name': row[1],
                    'email': row[2],
                    'subject': row[3] or 'N/A',
                    'class_teacher_of': row[4] if row[4] else 'Not Assigned',
                    'created_at': row[5],
                    'password': row[6],
                    'profile_pic_url': f"{settings.MEDIA_URL}{row[7]}" if row[7] else f"{settings.MEDIA_URL}pfpicsteacher/default.jpg",
                    
                    # Extended profile info
                    'full_name': row[8] or row[1],
                    'gender': row[9],
                    'date_of_birth': row[10],
                    'blood_group': row[11],
                    'nationality': row[12] or 'Indian',
                    'mobile_number': row[13],
                    'alternate_contact': row[14],
                    'official_email': row[15] or row[2],
                    'residential_address': row[16],
                    'city_state_pin': row[17],
                    'emergency_contact_name': row[18],
                    'emergency_contact_number': row[19],
                    'designation': row[20],
                    'department': row[21],
                    'subjects_taught': row[22],
                    'classes_assigned': row[23],
                    'sections': row[24],
                    'employee_type': row[25],
                    'employment_status': row[26] or 'Active',
                    'joining_date': row[27],
                    'qualification': row[28],
                    'specialization': row[29],
                    'board_university': row[30],
                    'year_of_passing': row[31],
                    'bed_ctet_tet': row[32],
                    'special_training': row[33],
                    'workshops_attended': row[34],
                    'is_class_teacher': row[35] or 'No',
                    'house_club_incharge': row[36],
                    'cocurricular_responsibilities': row[37],
                    'exam_duties': row[38],
                })
                
        return render(request, 'users/manage_teachers_enhanced.html', {'teachers': teachers})
    except Error as e:
        messages.error(request, f'Error fetching teachers: {str(e)}')
        return redirect('manage_teachers')


def add_teacher(request):
    """
    Handles adding a new teacher with comprehensive profile information.
    """
    if request.method != 'POST':
        return redirect('manage_teachers')

    # Basic required fields
    name = request.POST.get('name')
    email = request.POST.get('email')
    subject = request.POST.get('subject')
    password = request.POST.get('password')
    class_teacher_of = request.POST.get('class_teacher_of') or None
    
    # Extended profile fields - Basic Details
    full_name = request.POST.get('full_name') or name
    gender = request.POST.get('gender')
    date_of_birth = request.POST.get('date_of_birth') or None
    blood_group = request.POST.get('blood_group')
    nationality = request.POST.get('nationality') or 'Indian'
    
    # Contact Information
    mobile_number = request.POST.get('mobile_number')
    alternate_contact = request.POST.get('alternate_contact')
    official_email = request.POST.get('official_email') or email
    residential_address = request.POST.get('residential_address')
    city_state_pin = request.POST.get('city_state_pin')
    emergency_contact_name = request.POST.get('emergency_contact_name')
    emergency_contact_number = request.POST.get('emergency_contact_number')
    
    # Employment Details
    designation = request.POST.get('designation')
    department = request.POST.get('department')
    subjects_taught = request.POST.get('subjects_taught')
    classes_assigned = request.POST.get('classes_assigned')
    sections = request.POST.get('sections')
    employee_type = request.POST.get('employee_type')
    employment_status = request.POST.get('employment_status') or 'Active'
    joining_date = request.POST.get('joining_date') or None
    
    # Qualifications
    qualification = request.POST.get('qualification')
    specialization = request.POST.get('specialization')
    board_university = request.POST.get('board_university')
    year_of_passing = request.POST.get('year_of_passing') or None
    
    # Professional Certifications
    bed_ctet_tet = request.POST.get('bed_ctet_tet')
    special_training = request.POST.get('special_training')
    workshops_attended = request.POST.get('workshops_attended')
    
    # Academic Responsibilities
    is_class_teacher = request.POST.get('is_class_teacher') or 'No'
    house_club_incharge = request.POST.get('house_club_incharge')
    cocurricular_responsibilities = request.POST.get('cocurricular_responsibilities')
    exam_duties = request.POST.get('exam_duties')

    # Validation
    if not all([name, email, subject, password]):
        messages.error(request, 'Name, email, subject, and password are required.')
        return redirect('manage_teachers')

    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Check if email already exists
                cursor.execute("SELECT id FROM teachers WHERE email = %s", [email])
                if cursor.fetchone():
                    messages.error(request, 'Email already exists.')
                    return redirect('manage_teachers')

                # Insert into teachers table (basic login info)
                cursor.execute(
                    "INSERT INTO teachers (name, email, subject, class_teacher_of, password) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    [name, email, subject, class_teacher_of, password]
                )
                teacher_id = cursor.lastrowid

                # Insert into admin_manage_users
                username = f"{name.lower().replace(' ', '')}_{teacher_id}"
                cursor.execute(
                    """INSERT INTO admin_manage_users 
                    (name, email, username, password, role, created_at, updated_at) 
                    VALUES (%s, %s, %s, %s, %s, NOW(), NOW())""",
                    [name, email, username, password, 'teacher']
                )
                
                # Insert comprehensive profile into teacher_profiles
                cursor.execute("""
                    INSERT INTO teacher_profiles (
                        teacher_id, full_name, gender, date_of_birth, blood_group, nationality,
                        mobile_number, alternate_contact, official_email, residential_address, city_state_pin,
                        emergency_contact_name, emergency_contact_number,
                        designation, department, subjects_taught, classes_assigned, sections,
                        employee_type, employment_status, joining_date,
                        qualification, specialization, board_university, year_of_passing,
                        bed_ctet_tet, special_training, workshops_attended,
                        is_class_teacher, house_club_incharge, cocurricular_responsibilities, exam_duties
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                """, [
                    teacher_id, full_name, gender, date_of_birth, blood_group, nationality,
                    mobile_number, alternate_contact, official_email, residential_address, city_state_pin,
                    emergency_contact_name, emergency_contact_number,
                    designation, department, subjects_taught, classes_assigned, sections,
                    employee_type, employment_status, joining_date,
                    qualification, specialization, board_university, year_of_passing,
                    bed_ctet_tet, special_training, workshops_attended,
                    is_class_teacher, house_club_incharge, cocurricular_responsibilities, exam_duties
                ])

        messages.success(request, f'Teacher "{full_name}" added successfully with complete profile!')
    except Error as e:
        messages.error(request, f'Error adding teacher: {str(e)}')
    return redirect('manage_teachers')



from django.db.utils import IntegrityError  # For better error handling

def update_teacher(request):
    """
    Handles updating an existing teacher's comprehensive profile.
    """
    if 'admin_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('admin_login')

    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('manage_teachers')

    teacher_id = request.POST.get('teacher_id')
    if not teacher_id:
        messages.error(request, 'Teacher ID is required.')
        return redirect('manage_teachers')

    # Basic required fields
    name = request.POST.get('name', '').strip()
    email = request.POST.get('email', '').strip()
    subject = request.POST.get('subject', '').strip()
    password = request.POST.get('password', '').strip()

    if not all([name, email, subject, password]):
        messages.error(request, 'Name, Email, Subject, and Password are required.')
        return redirect('manage_teachers')

    # Optional fields with defaults
    class_teacher_of = request.POST.get('class_teacher_of') or None  # Can be empty → None

    full_name = request.POST.get('full_name') or name
    gender = request.POST.get('gender')
    date_of_birth = request.POST.get('date_of_birth') or None
    blood_group = request.POST.get('blood_group')
    nationality = request.POST.get('nationality') or 'Indian'

    mobile_number = request.POST.get('mobile_number')
    alternate_contact = request.POST.get('alternate_contact')
    official_email = request.POST.get('official_email') or email
    residential_address = request.POST.get('residential_address')
    city_state_pin = request.POST.get('city_state_pin')
    emergency_contact_name = request.POST.get('emergency_contact_name')
    emergency_contact_number = request.POST.get('emergency_contact_number')

    designation = request.POST.get('designation')
    department = request.POST.get('department')
    subjects_taught = request.POST.get('subjects_taught')
    classes_assigned = request.POST.get('classes_assigned')
    sections = request.POST.get('sections')
    employee_type = request.POST.get('employee_type')
    employment_status = request.POST.get('employment_status') or 'Active'
    joining_date = request.POST.get('joining_date') or None

    qualification = request.POST.get('qualification')
    specialization = request.POST.get('specialization')
    board_university = request.POST.get('board_university')
    year_of_passing = request.POST.get('year_of_passing') or None
    bed_ctet_tet = request.POST.get('bed_ctet_tet')
    special_training = request.POST.get('special_training')
    workshops_attended = request.POST.get('workshops_attended')

    is_class_teacher = request.POST.get('is_class_teacher') or 'No'
    house_club_incharge = request.POST.get('house_club_incharge')
    cocurricular_responsibilities = request.POST.get('cocurricular_responsibilities')
    exam_duties = request.POST.get('exam_duties')

    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Get current email for admin_manage_users update
                cursor.execute("SELECT email FROM teachers WHERE id = %s", [teacher_id])
                current = cursor.fetchone()
                if not current:
                    messages.error(request, 'Teacher not found.')
                    return redirect('manage_teachers')
                old_email = current[0]

                # Check email uniqueness (excluding current teacher)
                cursor.execute(
                    "SELECT id FROM teachers WHERE email = %s AND id != %s",
                    [email, teacher_id]
                )
                if cursor.fetchone():
                    messages.error(request, 'This email is already used by another teacher.')
                    return redirect('manage_teachers')

                # Update main teachers table
                cursor.execute("""
                    UPDATE teachers 
                    SET name = %s, email = %s, subject = %s, 
                        class_teacher_of = %s, password = %s
                    WHERE id = %s
                """, [name, email, subject, class_teacher_of, password, teacher_id])

                # Update admin_manage_users (for login consistency)
                cursor.execute("""
                    UPDATE admin_manage_users 
                    SET name = %s, email = %s, password = %s, updated_at = NOW()
                    WHERE email = %s AND role = 'teacher'
                """, [name, email, password, old_email])

                # Update or Insert into teacher_profiles
                cursor.execute(
                    "SELECT 1 FROM teacher_profiles WHERE teacher_id = %s",
                    [teacher_id]
                )
                profile_exists = cursor.fetchone()

                profile_data = [
                    teacher_id, full_name, gender, date_of_birth, blood_group, nationality,
                    mobile_number, alternate_contact, official_email, residential_address, city_state_pin,
                    emergency_contact_name, emergency_contact_number,
                    designation, department, subjects_taught, classes_assigned, sections,
                    employee_type, employment_status, joining_date,
                    qualification, specialization, board_university, year_of_passing,
                    bed_ctet_tet, special_training, workshops_attended,
                    is_class_teacher, house_club_incharge, cocurricular_responsibilities, exam_duties
                ]

                if profile_exists:
                    # Update
                    cursor.execute("""
                        UPDATE teacher_profiles SET
                            full_name = %s, gender = %s, date_of_birth = %s, blood_group = %s, nationality = %s,
                            mobile_number = %s, alternate_contact = %s, official_email = %s, 
                            residential_address = %s, city_state_pin = %s,
                            emergency_contact_name = %s, emergency_contact_number = %s,
                            designation = %s, department = %s, subjects_taught = %s, 
                            classes_assigned = %s, sections = %s,
                            employee_type = %s, employment_status = %s, joining_date = %s,
                            qualification = %s, specialization = %s, board_university = %s, year_of_passing = %s,
                            bed_ctet_tet = %s, special_training = %s, workshops_attended = %s,
                            is_class_teacher = %s, house_club_incharge = %s, 
                            cocurricular_responsibilities = %s, exam_duties = %s,
                            updated_at = NOW()
                        WHERE teacher_id = %s
                    """, profile_data[1:] + [teacher_id])  # Skip teacher_id in VALUES, add at end for WHERE
                else:
                    # Insert
                    cursor.execute("""
                        INSERT INTO teacher_profiles (
                            teacher_id, full_name, gender, date_of_birth, blood_group, nationality,
                            mobile_number, alternate_contact, official_email, residential_address, city_state_pin,
                            emergency_contact_name, emergency_contact_number,
                            designation, department, subjects_taught, classes_assigned, sections,
                            employee_type, employment_status, joining_date,
                            qualification, specialization, board_university, year_of_passing,
                            bed_ctet_tet, special_training, workshops_attended,
                            is_class_teacher, house_club_incharge, cocurricular_responsibilities, exam_duties
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, profile_data)

        messages.success(request, f'Teacher "{full_name}" has been updated successfully!')
        
    except IntegrityError as e:
        messages.error(request, 'Database error: Possible duplicate entry or constraint violation.')
    except Exception as e:
        messages.error(request, f'Error updating teacher: {str(e)}')

    return redirect('manage_teachers')


def delete_teacher(request, teacher_id):
    """
    Deletes a teacher and cleans up all related records.
    """
    if request.method != 'POST' and 'delete_teacher' not in request.POST:
        return redirect('manage_teachers')

    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Get teacher email for admin_manage_users cleanup
                cursor.execute("SELECT email FROM teachers WHERE id = %s", [teacher_id])
                teacher = cursor.fetchone()
                if not teacher:
                    messages.error(request, 'Teacher not found.')
                    return redirect('manage_teachers')

                email = teacher[0]

                # Delete dependent records (CASCADE will handle teacher_profiles)
                cursor.execute("DELETE FROM timetable WHERE teacher_id = %s", [teacher_id])
                cursor.execute("DELETE FROM teacher_profiles WHERE teacher_id = %s", [teacher_id])

                # Delete from teachers table
                cursor.execute("DELETE FROM teachers WHERE id = %s", [teacher_id])

                # Delete from admin_manage_users
                cursor.execute(
                    "DELETE FROM admin_manage_users WHERE email = %s AND role = %s",
                    [email, 'teacher']
                )

        messages.success(request, 'Teacher and all associated records deleted successfully.')
    except Error as e:
        messages.error(request, f'Error deleting teacher: {str(e)}')
    return redirect('manage_teachers')


def view_teacher_profile(request, teacher_id):
    """
    Displays a comprehensive view of a single teacher's profile.
    """
    if not request.session.get('admin_id'):
        messages.error(request, 'Please log in to access this page.')
        return redirect('admin_login')

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    t.id, t.name, t.email, t.subject, t.class_teacher_of, t.created_at, t.password,
                    p.profile_pic_url,
                    tp.*
                FROM teachers t
                LEFT JOIN profile_pics_teachers p ON t.id = p.teacher_id
                LEFT JOIN teacher_profiles tp ON t.id = tp.teacher_id
                WHERE t.id = %s
            """, [teacher_id])
            
            row = cursor.fetchone()
            if not row:
                messages.error(request, 'Teacher not found.')
                return redirect('manage_teachers')
            
            # Build teacher profile dictionary
            teacher = {
                'id': row[0],
                'name': row[1],
                'email': row[2],
                'subject': row[3],
                'class_teacher_of': row[4],
                'created_at': row[5],
                'password': row[6],
                'profile_pic_url': f"{settings.MEDIA_URL}{row[7]}" if row[7] else f"{settings.MEDIA_URL}pfpicsteacher/default.jpg",
            }
            
            # Add profile data if exists
            if len(row) > 8 and row[8]:  # If teacher_profiles data exists
                profile_fields = [
                    'profile_id', 'teacher_id', 'full_name', 'gender', 'date_of_birth', 
                    'blood_group', 'nationality', 'profile_photo_link', 'mobile_number',
                    'alternate_contact', 'official_email', 'residential_address', 
                    'city_state_pin', 'emergency_contact_name', 'emergency_contact_number',
                    'designation', 'department', 'subjects_taught', 'classes_assigned',
                    'sections', 'employee_type', 'employment_status', 'joining_date',
                    'qualification', 'specialization', 'board_university', 'year_of_passing',
                    'bed_ctet_tet', 'special_training', 'workshops_attended',
                    'is_class_teacher', 'house_club_incharge', 'cocurricular_responsibilities',
                    'exam_duties', 'profile_created_at', 'profile_updated_at'
                ]
                
                for i, field in enumerate(profile_fields, start=8):
                    if i < len(row):
                        teacher[field] = row[i]
            
            return render(request, 'users/teacher_profile_view.html', {'teacher': teacher})
            
    except Error as e:
        messages.error(request, f'Error fetching teacher profile: {str(e)}')
        return redirect('manage_teachers')




from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import connection, Error


@require_http_methods(["GET"])
def get_teacher_data(request, teacher_id):
    """
    API endpoint to fetch comprehensive teacher data for editing.
    Returns JSON with all teacher profile information.
    """
    if not request.session.get('admin_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    t.id, t.name, t.email, t.subject, t.class_teacher_of, t.password,
                    tp.full_name, tp.gender, tp.date_of_birth, tp.blood_group, tp.nationality,
                    tp.mobile_number, tp.alternate_contact, tp.official_email,
                    tp.residential_address, tp.city_state_pin,
                    tp.emergency_contact_name, tp.emergency_contact_number,
                    tp.designation, tp.department, tp.subjects_taught, tp.classes_assigned,
                    tp.sections, tp.employee_type, tp.employment_status, tp.joining_date,
                    tp.qualification, tp.specialization, tp.board_university, tp.year_of_passing,
                    tp.bed_ctet_tet, tp.special_training, tp.workshops_attended,
                    tp.is_class_teacher, tp.house_club_incharge, 
                    tp.cocurricular_responsibilities, tp.exam_duties
                FROM teachers t
                LEFT JOIN teacher_profiles tp ON t.id = tp.teacher_id
                WHERE t.id = %s
            """, [teacher_id])
            
            row = cursor.fetchone()
            if not row:
                return JsonResponse({'error': 'Teacher not found'}, status=404)
            
            # Build teacher data dictionary
            teacher_data = {
                'id': row[0],
                'name': row[1],
                'email': row[2],
                'subject': row[3],
                'class_teacher_of': row[4] if row[4] else '',
                'password': row[5],
                'full_name': row[6] if row[6] else row[1],
                'gender': row[7] if row[7] else '',
                'date_of_birth': str(row[8]) if row[8] else '',
                'blood_group': row[9] if row[9] else '',
                'nationality': row[10] if row[10] else 'Indian',
                'mobile_number': row[11] if row[11] else '',
                'alternate_contact': row[12] if row[12] else '',
                'official_email': row[13] if row[13] else row[2],
                'residential_address': row[14] if row[14] else '',
                'city_state_pin': row[15] if row[15] else '',
                'emergency_contact_name': row[16] if row[16] else '',
                'emergency_contact_number': row[17] if row[17] else '',
                'designation': row[18] if row[18] else '',
                'department': row[19] if row[19] else '',
                'subjects_taught': row[20] if row[20] else '',
                'classes_assigned': row[21] if row[21] else '',
                'sections': row[22] if row[22] else '',
                'employee_type': row[23] if row[23] else '',
                'employment_status': row[24] if row[24] else 'Active',
                'joining_date': str(row[25]) if row[25] else '',
                'qualification': row[26] if row[26] else '',
                'specialization': row[27] if row[27] else '',
                'board_university': row[28] if row[28] else '',
                'year_of_passing': row[29] if row[29] else '',
                'bed_ctet_tet': row[30] if row[30] else '',
                'special_training': row[31] if row[31] else '',
                'workshops_attended': row[32] if row[32] else '',
                'is_class_teacher': row[33] if row[33] else 'No',
                'house_club_incharge': row[34] if row[34] else '',
                'cocurricular_responsibilities': row[35] if row[35] else '',
                'exam_duties': row[36] if row[36] else '',
            }
            
            return JsonResponse(teacher_data)
            
    except Error as e:
        return JsonResponse({'error': str(e)}, status=500)



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, transaction
from django.urls import reverse
from django.db import IntegrityError

def teacher_signup(request):
    """Final robust teacher signup with raw SQL, allowing custom Teacher ID"""
    if request.method != 'POST':
        return render(request, 'users/teacher_signup.html', {'form_cleared': request.GET.get('clear')})

    # Sanitize and validate inputs with trimming
    teacher_id = request.POST.get('id', '').strip()
    name = request.POST.get('name', '').strip()
    email = request.POST.get('email', '').lower().strip()
    subject = request.POST.get('subject', '').strip()
    password = request.POST.get('password', '').strip()
    class_teacher_of = request.POST.get('class_teacher_of', '').strip() or None

    # Optional: Temp debug logs (remove after fixing)
    print(f"DEBUG: Input - Teacher ID: '{repr(teacher_id)}' (len: {len(teacher_id)})")
    print(f"DEBUG: Input - Name: '{repr(name)}' (len: {len(name)})")
    print(f"DEBUG: Input - Email: '{repr(email)}' (len: {len(email)})")
    print(f"DEBUG: Input - Subject: '{repr(subject)}' (len: {len(subject)})")
    print(f"DEBUG: Input - Password: '{repr(password)}' (len: {len(password)})")

    # Clear existing messages
    storage = messages.get_messages(request)
    storage.used = True

    # Validate required fields after trimming
    if not all([teacher_id, name, email, subject, password]):
        messages.error(request, 'Teacher ID, Name, Email, Subject, and Password are required')
        return redirect(reverse('teacher_signup') + '?clear=1')

    # Validate input lengths based on table schema (after trimming)
    if len(name) > 100:
        messages.error(request, 'Name exceeds maximum length of 100 characters')
        return redirect(reverse('teacher_signup') + '?clear=1')
    if len(email) > 100:
        messages.error(request, 'Email exceeds maximum length of 100 characters')
        return redirect(reverse('teacher_signup') + '?clear=1')
    if len(subject) > 50:
        messages.error(request, 'Subject exceeds maximum length of 50 characters')
        return redirect(reverse('teacher_signup') + '?clear=1')
    if class_teacher_of and len(class_teacher_of) > 20:
        messages.error(request, 'Class Teacher Of exceeds maximum length of 20 characters')
        return redirect(reverse('teacher_signup') + '?clear=1')
    if len(password) > 255:
        messages.error(request, 'Password exceeds maximum length of 255 characters')
        return redirect(reverse('teacher_signup') + '?clear=1')

    # Validate teacher_id is numeric and positive
    if not teacher_id.isdigit() or int(teacher_id) <= 0:
        messages.error(request, 'Teacher ID must be a positive number')
        return redirect(reverse('teacher_signup') + '?clear=1')

    # Validate password is numeric and max 10 digits (after trimming)
    if not password.isdigit() or len(password) > 10:
        messages.error(request, 'Password must be numeric and up to 10 digits')
        return redirect(reverse('teacher_signup') + '?clear=1')

    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Check for existing teacher_id in teachers table
                cursor.execute("SELECT id FROM teachers WHERE id = %s", [teacher_id])
                if cursor.fetchone():
                    print(f"DEBUG: Duplicate teacher ID detected: {teacher_id}")  # Debug
                    messages.error(request, f'Teacher ID "{teacher_id}" already exists')
                    return redirect(reverse('teacher_signup') + '?clear=1')

                # Check for existing email in both tables with TRIM
                cursor.execute("""
                    SELECT 'teachers' as source FROM teachers WHERE TRIM(email) = %s
                    UNION ALL
                    SELECT 'admin' as source FROM admin_manage_users WHERE TRIM(email) = %s
                    LIMIT 1
                """, [email, email])
                
                existing = cursor.fetchone()
                if existing:
                    print(f"DEBUG: Duplicate email detected: {email} in {existing[0]}")  # Debug
                    messages.error(request, f'Email "{email}" already exists in our system')
                    return redirect(reverse('teacher_signup') + '?clear=1')

                # Insert into teachers with explicit id
                cursor.execute(
                    """INSERT INTO teachers (id, name, email, subject, class_teacher_of, password)
                    VALUES (%s, %s, %s, %s, %s, %s)""",
                    [teacher_id, name, email, subject, class_teacher_of, password]
                )

                # Generate unique username (fit varchar(50)) using trimmed name
                base_username = name.lower().replace(' ', '_')[:40]  # Reserve space for _id
                username = f"{base_username}_{teacher_id}"[:50]

                # Insert into admin_manage_users
                cursor.execute(
                    """INSERT INTO admin_manage_users
                    (name, email, username, password, role, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, NOW(), NOW())""",
                    [name, email, username, password, 'Teacher']
                )

                # Optional: Temp success log (remove after)
                print(f"DEBUG: SUCCESS for Teacher ID '{teacher_id}' ({name}, {email})")

        messages.success(request, 'Registration successful! Please login.')
        return redirect('teacher_login')

    except IntegrityError as e:
        print(f"DEBUG: IntegrityError: {str(e)}")  # Debug
        messages.error(request, f'Teacher ID "{teacher_id}" or Email "{email}" already exists in our system')
        return redirect(reverse('teacher_signup') + '?clear=1')
    except Exception as e:
        print(f"DEBUG: Exception: {str(e)}")  # Debug
        messages.error(request, 'System error during registration')
        return redirect(reverse('teacher_signup') + '?clear=1')


from django.db import connection
from django.http import HttpResponse
from django.shortcuts import render

def teacher_login(request):
    if request.method == "POST":
        teacher_id = request.POST.get("id", "").strip()  # Trim user input
        password = request.POST.get("password", "").strip()  # Trim user input

        # Optional: Temp debug logs (remove after fixing)
        print(f"DEBUG: Input - Teacher ID: '{repr(teacher_id)}' (len: {len(teacher_id)})")
        print(f"DEBUG: Input - Password: '{repr(password)}' (len: {len(password)})")

        # Check user credentials in MySQL with TRIM for exact match
        with connection.cursor() as cursor:
            # Optional: Temp debug - find similar teachers (remove after)
            cursor.execute(
                "SELECT id, name, LENGTH(id), LENGTH(password) FROM teachers WHERE id LIKE %s OR TRIM(password) LIKE %s", 
                (f"%{teacher_id}%", f"%{password}%")
            )
            similar_teachers = cursor.fetchall()
            if similar_teachers:
                print("DEBUG: Similar teachers found:")
                for t in similar_teachers:
                    print(f"  ID: {t[0]}, Name: '{t[1]}', ID len: {t[2]}, Pass len: {t[3]}")

            # Exact match query with TRIM (id likely numeric, but safe)
            cursor.execute(
                "SELECT id, name, password FROM teachers WHERE id = %s AND TRIM(password) = %s", 
                (teacher_id, password)
            )
            user = cursor.fetchone()

        if user:
            # Store trimmed values in session
            clean_name = user[1].strip()
            request.session["teacher_id"] = user[0]
            request.session["username"] = clean_name
            
            # Optional: Temp success log (remove after)
            print(f"DEBUG: SUCCESS for Teacher '{clean_name}' (ID: {user[0]})")
            
            return HttpResponse("Success")

        # Optional: Temp failure log (remove after)
        print("DEBUG: No exact match found")
        
        # If credentials are invalid, send error message
        return HttpResponse("Invalid credentials!")  

    return render(request, "users/teacher_login.html")


def teacher_change_credentials(request):
    if request.method == "POST":
        current_id = request.POST.get("current_id")
        new_id = request.POST.get("new_id")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        if new_password != confirm_password:
            return HttpResponse("Passwords do not match!")

        # Check if current teacher exists
        with connection.cursor() as cursor:
            cursor.execute("SELECT id FROM teachers WHERE id = %s", (current_id,))
            teacher = cursor.fetchone()

            if not teacher:
                return HttpResponse("Teacher not found!")

            # Check if new ID already exists (if different from current)
            if new_id != current_id:
                cursor.execute("SELECT id FROM teachers WHERE id = %s", (new_id,))
                existing_teacher = cursor.fetchone()
                if existing_teacher:
                    return HttpResponse("Teacher ID already exists!")

            # Update ID and/or password
            if new_id != current_id:
                # Note: Updating primary key directly; assume no foreign keys or handle accordingly
                cursor.execute("UPDATE teachers SET id = %s, password = %s WHERE id = %s", (new_id, new_password, current_id))
            else:
                cursor.execute("UPDATE teachers SET password = %s WHERE id = %s", (new_password, current_id))
            connection.commit()

        return HttpResponse("Success")

    # For GET requests, redirect to teacher login
    return redirect('teacher_login')

def teacher_dashboard(request):
    username = request.session.get('username')
    if not username:
        messages.error(request, 'Please log in to access the mark entry system.')
        return redirect('teacher_login')
    
    return render(request, 'users/teacher_dashboard.html', {'username': username})






from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, IntegrityError
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

import json

def teacher_portal(request):
    """
    Main teacher portal page for marking attendance
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'Please log in to access this page.')
        return redirect('teacher_login')

    today_date = datetime.now().date().strftime('%Y-%m-%d')  # Changed this line
    selected_date = request.GET.get('date', today_date)
    
    # Get all unique classes
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT DISTINCT class 
            FROM student_page1 
            WHERE class IS NOT NULL AND class != '' 
            ORDER BY class
        """)
        classes = [row[0] for row in cursor.fetchall()]

    return render(request, 'users/teacher_portal.html', {
        'classes': classes,
        'selected_date': selected_date,
    })


def mark_single_attendance(request):
    """
    Mark attendance for a single student via AJAX
    This is called every time a teacher clicks Present/Absent/Leave
    """
    if not request.session.get('teacher_id'):
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    try:
        selected_class = request.POST.get('class')
        selected_section = request.POST.get('section')
        selected_date = request.POST.get('date')
        student_name = request.POST.get('student_name')
        status = request.POST.get('status')

        # Validate all required fields
        if not all([selected_class, selected_section, selected_date, student_name, status]):
            return JsonResponse({
                'success': False, 
                'message': 'Missing required fields'
            }, status=400)

        with connection.cursor() as cursor:
            # Get student info
            cursor.execute("""
                SELECT user_id, name, admission_number, section
                FROM student_page1 
                WHERE name = %s AND class = %s AND section = %s
            """, [student_name, selected_class, selected_section])
            
            student_info = cursor.fetchone()
            
            if not student_info:
                return JsonResponse({
                    'success': False, 
                    'message': f'Student {student_name} not found'
                }, status=404)
            
            student_id, name, admission_number, section = student_info
            
            # Check if attendance already exists for this student on this date
            cursor.execute("""
                SELECT id, status 
                FROM attendance 
                WHERE student_id = %s AND class = %s AND section = %s AND date = %s
            """, [student_id, selected_class, selected_section, selected_date])
            
            existing = cursor.fetchone()
            
            if existing:
                # Attendance already exists - don't allow changes
                return JsonResponse({
                    'success': False, 
                    'message': f'Attendance for {name} on {selected_date} is already marked as {existing[1].upper()}. Cannot modify existing records.',
                    'locked': True
                }, status=409)
            
            # Insert new attendance record
            try:
                cursor.execute("""
                    INSERT INTO attendance (student_id, class, section, date, status)
                    VALUES (%s, %s, %s, %s, %s)
                """, [student_id, selected_class, selected_section, selected_date, status])
                
                # Sync with admin_attendance table
                cursor.execute("""
                    INSERT INTO admin_attendance (student_id, name, admission_number, class, section, date, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, [student_id, name, admission_number, selected_class, section, selected_date, status])
                
                # Check if all students in this class/section now have attendance
                cursor.execute("""
                    SELECT COUNT(*) FROM student_page1 
                    WHERE class = %s AND section = %s
                """, [selected_class, selected_section])
                total_students = cursor.fetchone()[0]
                
                cursor.execute("""
                    SELECT COUNT(*) FROM attendance 
                    WHERE class = %s AND section = %s AND date = %s
                """, [selected_class, selected_section, selected_date])
                marked_students = cursor.fetchone()[0]
                
                all_marked = (marked_students == total_students)
                
                return JsonResponse({
                    'success': True,
                    'message': f'✓ Attendance saved for {name}: {status.upper()}',
                    'status': status,
                    'all_marked': all_marked,
                    'marked_count': marked_students,
                    'total_count': total_students
                })
                
            except IntegrityError:
                return JsonResponse({
                    'success': False,
                    'message': 'Database error: Record already exists',
                    'locked': True
                }, status=409)
                
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Server error: {str(e)}'
        }, status=500)


def get_attendance(request):
    """
    Fetch existing attendance records for a specific class, section, and date
    Called when loading students to show previously marked attendance
    """
    if not request.session.get('teacher_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    class_name = request.GET.get('class', '')
    section = request.GET.get('section', '')
    date = request.GET.get('date', '')
    
    if not class_name or not section or not date:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    attendance_dict = {}
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.name, a.status
            FROM attendance a
            JOIN student_page1 s ON a.student_id = s.user_id
            WHERE a.class = %s AND a.section = %s AND a.date = %s
        """, [class_name, section, date])
        
        for row in cursor.fetchall():
            attendance_dict[row[0]] = row[1]  # {student_name: status}
    
    return JsonResponse({'attendance': attendance_dict})


def get_students_by_class_section(request):
    """
    Get list of student names for a specific class and section
    Used for populating the student dropdown
    """
    if not request.session.get('teacher_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    class_name = request.GET.get('class', '')
    section = request.GET.get('section', '')
    
    if not class_name or not section:
        return JsonResponse({'students': []})
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT name 
            FROM student_page1 
            WHERE class = %s AND section = %s 
            ORDER BY name
        """, [class_name, section])
        
        students = [row[0] for row in cursor.fetchall()]
    
    return JsonResponse({'students': students})




def admin_attendance_portal(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'Please log in to access this page.')
        return redirect('admin_login')

    today_date = datetime.now().date().strftime('%Y-%m-%d')
    selected_date = request.GET.get('date', today_date)
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class, section FROM student_page1 WHERE section IS NOT NULL AND section != '' ORDER BY class, section")
        class_sections = [(row[0], row[1]) for row in cursor.fetchall()]
        classes = sorted(set(row[0] for row in class_sections)) or []

    selected_class = request.GET.get('class', '')
    selected_section = request.GET.get('section', '')
    students = []
    
    if selected_class and selected_section:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT user_id, name, admission_number, class, section
                FROM student_page1 
                WHERE class = %s AND section = %s 
                ORDER BY name, admission_number
                """, 
                [selected_class, selected_section]
            )
            students = [
                {
                    'user_id': row[0],
                    'name': row[1],
                    'admission_number': row[2],
                    'class': row[3],
                    'section': row[4] if row[4] else 'N/A'
                } for row in cursor.fetchall()
            ]
            
            cursor.execute(
                """
                SELECT student_id, status 
                FROM admin_attendance 
                WHERE class = %s AND section = %s AND date = %s
                """, 
                [selected_class, selected_section, selected_date]
            )
            attendance_records = {row[0]: row[1] for row in cursor.fetchall()}
            
            for student in students:
                student['status'] = attendance_records.get(student['user_id'], '')

    return render(request, 'users/admin_attendance.html', {
        'classes': classes,
        'class_sections': json.dumps(class_sections),
        'selected_class': selected_class,
        'selected_section': selected_section,
        'selected_date': selected_date,
        'students': students,
    })


def attendance_get_students(request):
    """
    AJAX endpoint specifically for Admin Attendance Portal.
    Returns list of student names for selected class and section.
    """
    if not request.session.get('admin_id'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    class_num = request.GET.get('class')
    section = request.GET.get('section')

    if not class_num or not section:
        return JsonResponse({'students': []})

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT name 
                FROM student_page1 
                WHERE class = %s 
                  AND section = %s 
                  AND name IS NOT NULL
                ORDER BY name ASC
            """, [class_num, section])
            
            students = [row[0] for row in cursor.fetchall()]
        
        return JsonResponse({'students': students})
    
    except Exception as e:
        print("Error in attendance_get_students:", str(e))
        return JsonResponse({'error': 'Database error'}, status=500)
    
    

def admin_mark_attendance(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'Please log in to access this page.')
        return redirect('admin_login')

    if request.method == 'POST':
        selected_class = request.POST.get('class')
        selected_section = request.POST.get('section')
        selected_date = request.POST.get('date')
        with connection.cursor() as cursor:
            for key, value in request.POST.items():
                if key.startswith('student_'):
                    student_id = int(key.split('_')[1])
                    status = value
                    cursor.execute(
                        """
                        SELECT name, admission_number, section
                        FROM student_page1 
                        WHERE user_id = %s
                        """,
                        [student_id]
                    )
                    student_info = cursor.fetchone()
                    if not student_info:
                        messages.error(request, f"Student ID {student_id} not found.")
                        continue
                    name, admission_number, section = student_info
                    
                    try:
                        cursor.execute(
                            """
                            SELECT id FROM admin_attendance 
                            WHERE student_id = %s AND class = %s AND section = %s AND date = %s
                            """,
                            [student_id, selected_class, selected_section, selected_date]
                        )
                        admin_exists = cursor.fetchone()
                        if admin_exists:
                            cursor.execute(
                                """
                                UPDATE admin_attendance 
                                SET status = %s, name = %s, admission_number = %s, section = %s
                                WHERE student_id = %s AND class = %s AND section = %s AND date = %s
                                """,
                                [status, name, admission_number, section, student_id, selected_class, selected_section, selected_date]
                            )
                            cursor.execute(
                                """
                                SELECT id FROM attendance 
                                WHERE student_id = %s AND class = %s AND section = %s AND date = %s
                                """,
                                [student_id, selected_class, selected_section, selected_date]
                            )
                            att_exists = cursor.fetchone()
                            if att_exists:
                                cursor.execute(
                                    """
                                    UPDATE attendance 
                                    SET status = %s
                                    WHERE student_id = %s AND class = %s AND section = %s AND date = %s
                                    """,
                                    [status, student_id, selected_class, selected_section, selected_date]
                                )
                            else:
                                cursor.execute(
                                    """
                                    INSERT INTO attendance (student_id, class, section, date, status)
                                    VALUES (%s, %s, %s, %s, %s)
                                    """,
                                    [student_id, selected_class, selected_section, selected_date, status]
                                )
                        else:
                            cursor.execute(
                                """
                                INSERT INTO admin_attendance (student_id, name, admission_number, class, section, date, status)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                """,
                                [student_id, name, admission_number, selected_class, section, selected_date, status]
                            )
                            cursor.execute(
                                """
                                INSERT INTO attendance (student_id, class, section, date, status)
                                VALUES (%s, %s, %s, %s, %s)
                                """,
                                [student_id, selected_class, selected_section, selected_date, status]
                            )
                    except IntegrityError:
                        messages.error(request, f"Duplicate attendance record for student ID {student_id} on {selected_date}.")
                        continue
        messages.success(request, f'Attendance updated for {selected_class} - {selected_section} on {selected_date}')
        return redirect(f"/admin_attendance/?class={selected_class}&section={selected_section}&date={selected_date}")
    return redirect('admin_attendance_portal')

def admin_generate_attendance_pdf(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'Please log in to access this page.')
        return redirect('admin_login')

    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')
    selected_date = request.GET.get('date')

    # Input validation
    if not all([selected_class, selected_section, selected_date]):
        messages.error(request, 'Please select class, section, and date to generate the PDF.')
        return redirect('admin_attendance_portal')

    # Debugging logs
    print(f"Generating PDF - Class: {selected_class}, Section: {selected_section}, Date: {selected_date}")

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT s.user_id, s.name, s.admission_number, s.section, a.status 
            FROM student_page1 s
            LEFT JOIN admin_attendance a ON s.user_id = a.student_id 
                AND a.class = %s AND a.section = %s AND a.date = %s
            WHERE s.class = %s AND COALESCE(s.section, '') = COALESCE(%s, '')
            ORDER BY s.name, s.admission_number
            """,
            [selected_class, selected_section, selected_date, selected_class, selected_section]
        )
        data = cursor.fetchall()
        print(f"Fetched {len(data)} records: {data}")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    header_style = ParagraphStyle(
        name='Header',
        parent=styles['Heading2'],
        fontSize=16,
        alignment=1,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=10
    )
    subheader_style = ParagraphStyle(
        name='Subheader',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=10
    )
    no_data_style = ParagraphStyle(
        name='NoData',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        textColor=colors.HexColor('#dc2626'),
        spaceAfter=10
    )
    
    logo_url = "https://via.placeholder.com/80"
    logo_img = None
    try:
        response = requests.get(logo_url, stream=True)
        if response.status_code == 200:
            logo_data = BytesIO(response.content)
            logo_img = Image(logo_data, width=0.8*inch, height=0.8*inch)
            logo_img.hAlign = 'LEFT'
            elements.append(logo_img)
    except Exception as e:
        print(f"Failed to load logo: {e}")
    
    elements.append(Paragraph("Manavargal School Management System", header_style))
    elements.append(Paragraph("Admin Attendance Report", header_style))
    elements.append(Paragraph(f"Class: {selected_class} | Section: {selected_section} | Date: {selected_date}", subheader_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subheader_style))
    elements.append(Spacer(1, 0.25*inch))
    
    if data:
        table_data = [['Student ID', 'Name', 'Admission Number', 'Class', 'Section', 'Status']]
        for row in data:
            status = row[4] if row[4] else 'Not Marked'
            table_data.append([
                str(row[0]),
                row[1] if row[1] else 'N/A',
                row[2] if row[2] else 'N/A',
                selected_class,
                row[3] if row[3] else 'N/A',
                status.capitalize()
            ])
        table = Table(table_data, colWidths=[0.75*inch, 1.5*inch, 1.75*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafd')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bfdbfe')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No attendance records found for the selected class, section, and date.", no_data_style))
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Admin_Attendance_{selected_class}_{selected_section}_{selected_date}.pdf"'
    response.write(pdf)
    return response



def student_portal(request):
    if "user_id" not in request.session:
        return redirect("/")
    
    user_id = request.session['user_id']
    selected_date = request.GET.get('date', '')
    
    with connection.cursor() as cursor:
        if selected_date:
            cursor.execute(
                """
                SELECT a.date, s.admission_number, s.name, a.class, a.section, a.status 
                FROM attendance a
                JOIN student_page1 s ON a.student_id = s.user_id
                WHERE s.user_id = %s AND a.date = %s
                ORDER BY a.date DESC
                """,
                [user_id, selected_date]
            )
        else:
            cursor.execute(
                """
                SELECT a.date, s.admission_number, s.name, a.class, a.section, a.status
                FROM attendance a
                JOIN student_page1 s ON a.student_id = s.user_id
                WHERE s.user_id = %s
                ORDER BY a.date DESC
                """,
                [user_id]
            )
        attendance_records = [
            {
                'date': row[0],
                'admission_number': row[1],
                'name': row[2],
                'class': row[3],
                'section': row[4] if row[4] else 'N/A',
                'status': row[5]
            } for row in cursor.fetchall()
        ]

    return render(request, 'users/student_attendance.html', {
        'attendance_records': attendance_records,
        'selected_date': selected_date
    })




from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
import re

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
import re

def parent_signup(request):
    if request.method == 'POST':
        admission_number = request.POST.get('admission_number')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        class_grade = request.POST.get('class')
        section = request.POST.get('section')
        roll_number = request.POST.get('roll_number')

        # Validate inputs
        if not all([admission_number, contact, email, class_grade, section, roll_number]):
            messages.error(request, 'All fields are required')
            return render(request, 'users/parent_signup.html')

        # Validate email format
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
            messages.error(request, 'Invalid email format')
            return render(request, 'users/parent_signup.html')

        # Validate contact number (10 digits)
        if not re.match(r'^\d{10}$', contact):
            messages.error(request, 'Contact number must be 10 digits')
            return render(request, 'users/parent_signup.html')

        # Validate roll_number (positive integer)
        try:
            roll_number = int(roll_number)
            if roll_number <= 0:
                raise ValueError
        except ValueError:
            messages.error(request, 'Roll number must be a positive integer')
            return render(request, 'users/parent_signup.html')

        # Check if admission_number or email already exists
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM student_page1 WHERE admission_number = %s", [admission_number])
            if cursor.fetchone()[0] > 0:
                messages.error(request, 'Admission number already exists')
                return render(request, 'users/parent_signup.html')

            cursor.execute("SELECT COUNT(*) FROM users WHERE email = %s", [email])
            if cursor.fetchone()[0] > 0:
                messages.error(request, 'Email already exists')
                return render(request, 'users/parent_signup.html')

        # Insert new user and related data
        try:
            with connection.cursor() as cursor:
                # Insert into users (using admission_number as username, contact as password)
                cursor.execute(
                    "INSERT INTO users (username, email, password) VALUES (%s, %s, %s)",
                    [admission_number, email, contact]
                )
                user_id = cursor.lastrowid

                # Insert into student_page1 (using placeholder for name)
                cursor.execute(
                    "INSERT INTO student_page1 (user_id, name, admission_number, class, section, roll_number) VALUES (%s, %s, %s, %s, %s, %s)",
                    [user_id, 'Placeholder Name', admission_number, class_grade, section, roll_number]
                )

                # Insert into student_page3
                cursor.execute(
                    "INSERT INTO student_page3 (user_id, contact) VALUES (%s, %s)",
                    [user_id, contact]
                )

                connection.commit()
                messages.success(request, 'Account created successfully! Please log in.')
                return redirect('parent_login')
        except Exception as e:
            connection.rollback()
            messages.error(request, 'An error occurred during signup: ' + str(e))
            return render(request, 'users/parent_signup.html')
    
    return render(request, 'users/parent_signup.html')

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def parent_login(request):
    if request.method == "POST":
        admission_number = request.POST.get("username")
        contact = request.POST.get("password")

        # Check user credentials in MySQL
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT u.id, sp1.admission_number
                FROM users u
                JOIN student_page1 sp1 ON u.id = sp1.user_id
                JOIN student_page3 sp3 ON u.id = sp3.user_id
                WHERE sp1.admission_number = %s AND sp3.contact = %s
            """, (admission_number, contact))
            user = cursor.fetchone()

        if user:
            request.session["user_id"] = user[0]  # Store user ID in session
            request.session["username"] = user[1]  # Store admission number in session
            
            return HttpResponse("Success") 

        # If credentials are invalid, send error message
        return HttpResponse("Invalid credentials!")  

    return render(request, "users/parent_login.html")


def parent_change_credentials(request):
    if request.method == "POST":
        current_admission = request.POST.get("current_username")
        new_admission = request.POST.get("new_username")
        new_contact = request.POST.get("new_password")
        confirm_contact = request.POST.get("confirm_password")

        if new_contact != confirm_contact:
            return HttpResponse("Phone numbers do not match!")

        # Check if current user exists
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT u.id 
                FROM users u
                JOIN student_page1 sp1 ON u.id = sp1.user_id
                WHERE sp1.admission_number = %s
            """, (current_admission,))
            user = cursor.fetchone()

            if not user:
                return HttpResponse("User not found!")

            # Check if new admission number already exists (if different from current)
            if new_admission != current_admission:
                cursor.execute("""
                    SELECT id 
                    FROM student_page1 
                    WHERE admission_number = %s
                """, (new_admission,))
                existing = cursor.fetchone()
                if existing:
                    return HttpResponse("Admission number already exists!")

            # Update admission number and/or contact
            if new_admission != current_admission:
                cursor.execute("""
                    UPDATE student_page1 
                    SET admission_number = %s 
                    WHERE user_id = %s
                """, (new_admission, user[0]))
                cursor.execute("""
                    UPDATE student_page3 
                    SET contact = %s 
                    WHERE user_id = %s
                """, (new_contact, user[0]))
                # Optionally update users.username if it's used elsewhere
                cursor.execute("""
                    UPDATE users 
                    SET username = %s 
                    WHERE id = %s
                """, (new_admission, user[0]))
            else:
                cursor.execute("""
                    UPDATE student_page3 
                    SET contact = %s 
                    WHERE user_id = %s
                """, (new_contact, user[0]))
            connection.commit()

        return HttpResponse("Success")

    # For GET requests, redirect to parent login
    return redirect('parent_login')

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def parent_dashboard(request):
    if 'user_id' not in request.session:
        messages.error(request, 'Please log in to access the dashboard.')
        return redirect('parent_login')

    # Fetch username from users table
    admin_name = "Guest"
    with connection.cursor() as cursor:
        cursor.execute("SELECT username FROM users WHERE id = %s", [request.session['user_id']])
        result = cursor.fetchone()
        if result:
            admin_name = result[0]

    return render(request, 'users/parent_dashboard.html', {'admin_name': admin_name})




from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.db import connection
import json

def mark_entry(request):
    if 'admin_id' not in request.session:
        messages.error(request, 'Please log in to access the mark entry system.')
        return redirect('admin_login')

    with connection.cursor() as cursor:
        # Verify admin exists (only id)
        cursor.execute("SELECT id FROM admins WHERE id = %s", [request.session['admin_id']])
        if not cursor.fetchone():
            messages.error(request, 'Admin not found.')
            return redirect('admin_login')

        # Safe defaults for admins (full access)
        role = 'classTeacher'
        teacher_subject = None
        teacher_name = request.session.get('admin_display_name', 'Admin')  # Set in login view if possible

        # Classes/sections
        cursor.execute("SELECT DISTINCT class FROM student_page1 ORDER BY class")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL ORDER BY section")
        sections = [row[0] for row in cursor.fetchall()]

        selected_class = request.GET.get('class', classes[0] if classes else '')
        selected_section = request.GET.get('section', sections[0] if sections else '')

        # Validate
        if selected_class not in classes:
            selected_class = classes[0] if classes else ''
        if selected_section not in sections:
            selected_section = sections[0] if sections else ''
        if not classes or not sections:
            messages.warning(request, 'No classes or sections found.')

        # Subjects
        subjects = []
        if selected_class:
            cursor.execute("SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name", [selected_class])
            subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]

        # Students with admission_number fallback (now includes class/section for display/differentiation)
        students = []
        if selected_class and selected_section:
            try:
                cursor.execute(
                    """
                    SELECT user_id AS id, name, roll_number, class, section, 
                           COALESCE(admission_number, '') AS admission_number 
                    FROM student_page1 
                    WHERE class = %s AND section = %s 
                    ORDER BY name
                    """,
                    [selected_class, selected_section]
                )
                students = [
                    {
                        'id': row[0], 
                        'name': row[1], 
                        'roll_number': row[2], 
                        'class': row[3], 
                        'section': row[4], 
                        'admission_number': row[5]
                    } 
                    for row in cursor.fetchall()
                ]
            except Exception as col_err:
                if "Unknown column 'admission_number'" in str(col_err):
                    cursor.execute(
                        """
                        SELECT user_id AS id, name, roll_number, class, section
                        FROM student_page1 
                        WHERE class = %s AND section = %s 
                        ORDER BY name
                        """,
                        [selected_class, selected_section]
                    )
                    students = [
                        {
                            'id': row[0], 
                            'name': row[1], 
                            'roll_number': row[2], 
                            'class': row[3], 
                            'section': row[4], 
                            'admission_number': ''
                        } 
                        for row in cursor.fetchall()
                    ]
                else:
                    raise col_err

    return render(request, 'users/mark.html', {
        'subjects': subjects,
        'students': students,
        'teacher_name': teacher_name,
        'role': role,
        'teacher_subject': teacher_subject,
        'classes': classes,
        'sections': sections,
        'selected_class': selected_class,
        'selected_section': selected_section
    })

def save_marks(request):
    # Support both admin and teacher sessions
    if 'admin_id' in request.session:
        user_id = request.session['admin_id']
        user_type = 'admin'
    elif 'teacher_id' in request.session:
        user_id = request.session['teacher_id']
        user_type = 'teacher'
    else:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            role = data.get('role')  # 'classTeacher' expected
            student_id = data.get('studentId')
            marks_data = data.get('marks')
            class_name = data.get('class')
            section = data.get('section')

            if not all([student_id, role, marks_data, class_name, section]):
                return JsonResponse({'success': False, 'message': 'Missing required fields.'}, status=400)

            with connection.cursor() as cursor:
                # Verify user exists (admin or teacher)
                if user_type == 'admin':
                    cursor.execute("SELECT id FROM admins WHERE id = %s", [user_id])
                else:
                    cursor.execute("SELECT id FROM teachers WHERE id = %s", [user_id])
                if not cursor.fetchone():
                    return JsonResponse({'success': False, 'message': 'User not found.'}, status=403)

                # Only allow classTeacher role
                if role != 'classTeacher':
                    return JsonResponse({'success': False, 'message': 'Invalid role.'}, status=403)

                # Verify student belongs to the class/section
                cursor.execute(
                    "SELECT user_id FROM student_page1 WHERE user_id = %s AND class = %s AND section = %s",
                    [student_id, class_name, section]
                )
                if not cursor.fetchone():
                    return JsonResponse({'success': False, 'message': 'Invalid student or class/section.'}, status=400)

                # Normalize marks_data to list
                if isinstance(marks_data, dict):
                    marks_data = [marks_data]
                if not marks_data:
                    return JsonResponse({'success': False, 'message': 'No marks data provided.'}, status=400)

                # Extract signatures
                teacher_signature_data = marks_data[0].get('signature')
                principal_signature_data = data.get('principalSignature')

                if not teacher_signature_data:
                    return JsonResponse({'success': False, 'message': 'Teacher signature is required.'}, status=400)
                if not principal_signature_data:
                    return JsonResponse({'success': False, 'message': 'Principal signature is required.'}, status=400)

                # === KEY FIX: Find the actual class teacher for this class-section ===
                class_section = f"{class_name}-{section}"
                cursor.execute(
                    "SELECT id FROM teachers WHERE class_teacher_of = %s",
                    [class_section]
                )
                teacher_row = cursor.fetchone()
                if not teacher_row:
                    return JsonResponse({
                        'success': False,
                        'message': f'No class teacher assigned for {class_section}. Please assign one first.'
                    }, status=400)

                class_teacher_id = teacher_row[0]

                # Save Teacher Signature under the CLASS TEACHER's ID
                cursor.execute(
                    """INSERT INTO teacher_signature (teacher_id, signature) 
                       VALUES (%s, %s) 
                       ON DUPLICATE KEY UPDATE signature = %s""",
                    [class_teacher_id, teacher_signature_data, teacher_signature_data]
                )

                # Save Principal Signature (global - use fixed principal_id = 1)
                cursor.execute(
                    """INSERT INTO principal_signature (principal_id, signature) 
                       VALUES (1, %s) 
                       ON DUPLICATE KEY UPDATE signature = %s""",
                    [principal_signature_data, principal_signature_data]
                )

                # Grade calculation
                def calculate_grade(marks, max_marks):
                    if max_marks == 0:
                        return 'E'
                    percentage = (marks / max_marks) * 100
                    if percentage >= 80:
                        return 'A'
                    elif percentage >= 60:
                        return 'B'
                    elif percentage >= 40:
                        return 'C'
                    elif percentage >= 33:
                        return 'D'
                    else:
                        return 'E'

                marks_updated = False
                marks_inserted = False

                # Process each subject
                for subject in marks_data:
                    subject_id = subject.get('subjectId')
                    marks_val = subject.get('marks')
                    max_marks_val = subject.get('maxMarks')

                    if not all([subject_id, marks_val is not None, max_marks_val]):
                        return JsonResponse({'success': False, 'message': 'Invalid subject data.'}, status=400)

                    # Verify subject exists for this class
                    cursor.execute(
                        "SELECT name, max_marks FROM school_subjects WHERE id = %s AND class = %s",
                        [subject_id, class_name]
                    )
                    subject_row = cursor.fetchone()
                    if not subject_row:
                        return JsonResponse({'success': False, 'message': 'Subject not found or not for this class.'}, status=404)

                    marks = int(marks_val)
                    max_marks = int(max_marks_val)

                    if marks < 0 or marks > max_marks or max_marks < 1:
                        return JsonResponse({'success': False, 'message': 'Invalid marks value.'}, status=400)

                    grade = calculate_grade(marks, max_marks)

                    # Insert or update marks
                    cursor.execute(
                        """INSERT INTO school_marks 
                           (student_id, subject_id, marks, max_marks, grade) 
                           VALUES (%s, %s, %s, %s, %s) 
                           ON DUPLICATE KEY UPDATE 
                           marks = %s, max_marks = %s, grade = %s""",
                        [student_id, subject_id, marks, max_marks, grade,
                         marks, max_marks, grade]
                    )

                    # Track insert vs update
                    cursor.execute(
                        "SELECT id FROM school_marks WHERE student_id = %s AND subject_id = %s AND marks = %s",
                        [student_id, subject_id, marks]
                    )
                    if cursor.rowcount > 1 or cursor.fetchone():  # rough check
                        marks_updated = True
                    else:
                        marks_inserted = True

                connection.commit()

                # Success message
                if marks_updated and marks_inserted:
                    message = 'Marks saved successfully! Some updated, some newly added.'
                elif marks_updated:
                    message = 'Marks updated successfully!'
                else:
                    message = 'Marks saved successfully!'

                return JsonResponse({
                    'success': True,
                    'message': message
                })

        except Exception as e:
            connection.rollback()
            return JsonResponse({
                'success': False,
                'message': f'Error saving marks: {str(e)}'
            }, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

def add_subject(request):
    if 'admin_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            subject_name = data.get('subjectName')
            max_marks = data.get('maxMarks')
            class_name = data.get('class')

            if not all([subject_name, max_marks, class_name]):
                return JsonResponse({'success': False, 'message': 'Missing fields.'}, status=400)

            max_marks = int(max_marks)
            if max_marks < 1:
                return JsonResponse({'success': False, 'message': 'Max marks must be positive.'}, status=400)

            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM student_page1 WHERE class = %s", [class_name])
                if cursor.fetchone()[0] == 0:
                    return JsonResponse({'success': False, 'message': 'Invalid class.'}, status=400)

                cursor.execute("SELECT COUNT(*) FROM school_subjects WHERE name = %s AND class = %s", [subject_name, class_name])
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'success': False, 'message': 'Subject exists.'}, status=400)

                cursor.execute("INSERT INTO school_subjects (name, max_marks, class) VALUES (%s, %s, %s)", [subject_name, max_marks, class_name])
                connection.commit()

                cursor.execute("SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name", [class_name])
                subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]

                return JsonResponse({'success': True, 'message': f'Subject {subject_name} added.', 'subjects': subjects})

        except Exception as e:
            connection.rollback()
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid method.'}, status=405)

def delete_subject(request):
    if 'admin_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            subject_id = data.get('subjectId')
            class_name = data.get('class')

            if not all([subject_id, class_name]):
                return JsonResponse({'success': False, 'message': 'Missing fields.'}, status=400)

            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM student_page1 WHERE class = %s", [class_name])
                if cursor.fetchone()[0] == 0:
                    return JsonResponse({'success': False, 'message': 'Invalid class.'}, status=400)

                cursor.execute("SELECT name FROM school_subjects WHERE id = %s AND class = %s", [subject_id, class_name])
                subject = cursor.fetchone()
                if not subject:
                    return JsonResponse({'success': False, 'message': 'Subject not found.'}, status=404)

                cursor.execute("DELETE FROM school_marks WHERE subject_id = %s", [subject_id])
                cursor.execute("DELETE FROM school_subjects WHERE id = %s", [subject_id])
                connection.commit()

                cursor.execute("SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name", [class_name])
                subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]

                return JsonResponse({'success': True, 'message': f'Subject {subject[0]} deleted.', 'subjects': subjects})

        except Exception as e:
            connection.rollback()
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid method.'}, status=405)

def progress_card(request):
    if 'admin_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('studentId')

            if not student_id:
                return JsonResponse({'success': False, 'message': 'Student ID required.'}, status=400)

            with connection.cursor() as cursor:
                # Verify admin
                cursor.execute("SELECT id FROM admins WHERE id = %s", [request.session['admin_id']])
                if not cursor.fetchone():
                    return JsonResponse({'success': False, 'message': 'Admin not found.'}, status=403)

                # Student with admission_number fallback
                try:
                    cursor.execute(
                        """
                        SELECT s.name, s.roll_number, s.class, s.section, COALESCE(s.admission_number, '') AS admission_number, p.image_path
                        FROM student_page1 s LEFT JOIN profile_pics p ON s.user_id = p.user_id
                        WHERE s.user_id = %s
                        """,
                        [student_id]
                    )
                    student_row = cursor.fetchone()
                    if student_row:
                        name, roll_number, class_name, section, admission_number, image_path = student_row
                    else:
                        name = roll_number = class_name = section = admission_number = image_path = ''
                except Exception as col_err:
                    if "Unknown column 'admission_number'" in str(col_err):
                        cursor.execute(
                            """
                            SELECT s.name, s.roll_number, s.class, s.section, p.image_path
                            FROM student_page1 s LEFT JOIN profile_pics p ON s.user_id = p.user_id
                            WHERE s.user_id = %s
                            """,
                            [student_id]
                        )
                        student_row = cursor.fetchone()
                        if student_row:
                            name, roll_number, class_name, section, image_path = student_row
                            admission_number = ''
                        else:
                            name = roll_number = class_name = section = admission_number = image_path = ''
                    else:
                        raise col_err

                if not name:
                    return JsonResponse({'success': False, 'message': 'Student not found.'}, status=404)

                # Marks
                cursor.execute(
                    """
                    SELECT ss.name, m.marks, m.max_marks, m.grade
                    FROM school_marks m JOIN school_subjects ss ON m.subject_id = ss.id
                    WHERE m.student_id = %s AND ss.class = %s
                    """,
                    [student_id, class_name]
                )
                marks = [{'subject': row[0], 'marks': row[1] or 0, 'max_marks': row[2], 'grade': row[3] or 'E'} for row in cursor.fetchall()]

                # Signature
                cursor.execute("SELECT signature FROM teacher_signature WHERE teacher_id = %s", [request.session['admin_id']])
                signature_row = cursor.fetchone()
                signature = signature_row[0] if signature_row else None

                # Principal Signature
                cursor.execute("SELECT signature FROM principal_signature WHERE principal_id = %s", [request.session['admin_id']])
                principal_signature_row = cursor.fetchone()
                principal_signature = principal_signature_row[0] if principal_signature_row else None

                teacher_name = request.session.get('admin_display_name', 'Admin')

                # Calculations
                total_marks = sum(mark['marks'] for mark in marks)
                total_max_marks = sum(mark['max_marks'] for mark in marks)
                percentage = (total_marks / total_max_marks * 100) if total_max_marks > 0 else 0
                grade = 'A' if percentage >= 80 else 'B' if percentage >= 60 else 'C' if percentage >= 40 else 'D' if percentage >= 33 else 'E'
                passed = all(mark['marks'] >= 0.33 * mark['max_marks'] for mark in marks) if marks else False

                response_data = {
                    'success': True,
                    'student': {
                        'name': name,
                        'roll_number': roll_number,
                        'class': class_name,
                        'section': section,
                        'admission_number': admission_number,  # Always included
                        'image_path': image_path or 'images/default_profile.jpg'
                    },
                    'marks': marks,
                    'total_marks': total_marks,
                    'total_max_marks': total_max_marks,
                    'percentage': round(percentage, 2),
                    'grade': grade,
                    'status': 'Pass' if passed else 'Fail',
                    'teacher_name': teacher_name
                }
                if signature:
                    response_data['signature'] = signature
                if principal_signature:
                    response_data['principal_signature'] = principal_signature
                return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid method.'}, status=405)

def get_students(request):
    if 'admin_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    try:
        class_name = request.GET.get('class')
        section = request.GET.get('section')

        if not class_name or not section:
            return JsonResponse({'success': False, 'message': 'Class and section required.'}, status=400)

        with connection.cursor() as cursor:
            # Verify admin
            cursor.execute("SELECT id FROM admins WHERE id = %s", [request.session['admin_id']])
            if not cursor.fetchone():
                return JsonResponse({'success': False, 'message': 'Admin not found.'}, status=403)

            # Students with admission_number fallback (and class/section for display)
            try:
                cursor.execute(
                    """
                    SELECT user_id AS id, name, roll_number, class, section, 
                           COALESCE(admission_number, '') AS admission_number 
                    FROM student_page1 
                    WHERE class = %s AND section = %s 
                    ORDER BY name
                    """,
                    [class_name, section]
                )
                students = [
                    {
                        'id': row[0], 
                        'name': row[1], 
                        'roll_number': row[2], 
                        'class': row[3], 
                        'section': row[4], 
                        'admission_number': row[5]
                    } 
                    for row in cursor.fetchall()
                ]
            except Exception as col_err:
                if "Unknown column 'admission_number'" in str(col_err):
                    cursor.execute(
                        """
                        SELECT user_id AS id, name, roll_number, class, section
                        FROM student_page1 
                        WHERE class = %s AND section = %s 
                        ORDER BY name
                        """,
                        [class_name, section]
                    )
                    students = [
                        {
                            'id': row[0], 
                            'name': row[1], 
                            'roll_number': row[2], 
                            'class': row[3], 
                            'section': row[4], 
                            'admission_number': ''
                        } 
                        for row in cursor.fetchall()
                    ]
                else:
                    raise col_err

            return JsonResponse({'success': True, 'students': students})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

def get_subjects(request):
    if 'admin_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    try:
        class_name = request.GET.get('class')
        if not class_name:
            return JsonResponse({'success': False, 'message': 'Class required.'}, status=400)

        with connection.cursor() as cursor:
            # Verify admin
            cursor.execute("SELECT id FROM admins WHERE id = %s", [request.session['admin_id']])
            if not cursor.fetchone():
                return JsonResponse({'success': False, 'message': 'Admin not found.'}, status=403)

            cursor.execute("SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name", [class_name])
            subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]
            return JsonResponse({'success': True, 'subjects': subjects})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)



def teacher_mark_entry(request):
    if 'teacher_id' not in request.session:
        messages.error(request, 'Please log in to access the mark entry system.')
        return redirect('teacher_login')

    with connection.cursor() as cursor:
        # Get teacher details
        cursor.execute(
            "SELECT subject, class_teacher_of, name FROM teachers WHERE id = %s",
            [request.session['teacher_id']]
        )
        teacher = cursor.fetchone()
        if not teacher:
            messages.error(request, 'Teacher not found.')
            return redirect('teacher_login')

        role = 'classTeacher' if teacher[1] else 'subjectTeacher'
        teacher_subject = teacher[0] if role == 'subjectTeacher' else None
        teacher_name = teacher[2]

        # Get distinct classes and sections
        cursor.execute("SELECT DISTINCT class FROM student_page1 ORDER BY class")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL ORDER BY section")
        sections = [row[0] for row in cursor.fetchall()]

        # Default class and section
        default_class = ''
        default_section = ''
        if teacher[1]:
            try:
                class_section = teacher[1].split('-')
                if len(class_section) == 2:
                    default_class, default_section = class_section
                else:
                    messages.warning(request, 'Invalid class_teacher_of format.')
            except Exception as e:
                messages.error(request, f'Error parsing class_teacher_of: {str(e)}')

        selected_class = request.GET.get('class', default_class or (classes[0] if classes else ''))
        selected_section = request.GET.get('section', default_section or (sections[0] if sections else ''))

        # Validate class and section
        if selected_class not in classes or selected_section not in sections:
            selected_class = classes[0] if classes else ''
            selected_section = sections[0] if sections else ''
            if not classes or not sections:
                messages.warning(request, 'No classes or sections found.')

        # Get subjects for selected class
        subjects = []
        if selected_class:
            cursor.execute(
                "SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name",
                [selected_class]
            )
            subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]

        # Get students for selected class and section
        students = []
        if selected_class and selected_section:
            try:
                cursor.execute(
                    """
                    SELECT user_id AS id, name, roll_number, class, section, 
                           COALESCE(admission_number, '') AS admission_number 
                    FROM student_page1 
                    WHERE class = %s AND section = %s 
                    ORDER BY name
                    """,
                    [selected_class, selected_section]
                )
                students = [
                    {
                        'id': row[0], 
                        'name': row[1], 
                        'roll_number': row[2], 
                        'class': row[3], 
                        'section': row[4], 
                        'admission_number': row[5]
                    } 
                    for row in cursor.fetchall()
                ]
            except Exception as col_err:
                if "Unknown column 'admission_number'" in str(col_err):
                    cursor.execute(
                        """
                        SELECT user_id AS id, name, roll_number, class, section
                        FROM student_page1 
                        WHERE class = %s AND section = %s 
                        ORDER BY name
                        """,
                        [selected_class, selected_section]
                    )
                    students = [
                        {
                            'id': row[0], 
                            'name': row[1], 
                            'roll_number': row[2], 
                            'class': row[3], 
                            'section': row[4], 
                            'admission_number': ''
                        } 
                        for row in cursor.fetchall()
                    ]
                else:
                    raise col_err

    return render(request, 'users/teacher_mark_entry.html', {
        'subjects': subjects,
        'students': students,
        'teacher_name': teacher_name,
        'role': role,
        'teacher_subject': teacher_subject,
        'classes': classes,
        'sections': sections,
        'selected_class': selected_class,
        'selected_section': selected_section
    })

from django.db import connection, transaction
from django.http import JsonResponse
import json

def teacher_save_marks(request):
    if 'teacher_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    try:
        data = json.loads(request.body)
        role = data.get('role')  # 'subjectTeacher' or 'classTeacher'
        student_id = data.get('studentId')
        marks_data = data.get('marks')
        class_name = data.get('class')
        section = data.get('section')
        principal_signature_data = data.get('principalSignature')

        # Basic validation
        if not all([role, student_id, marks_data, class_name, section, principal_signature_data]):
            return JsonResponse({'success': False, 'message': 'Missing required fields.'}, status=400)

        # Normalize marks_data
        if isinstance(marks_data, dict):
            marks_data = [marks_data]
        if not marks_data:
            return JsonResponse({'success': False, 'message': 'No marks data provided.'}, status=400)

        teacher_signature_data = marks_data[0].get('signature')
        if not teacher_signature_data:
            return JsonResponse({'success': False, 'message': 'Teacher signature is required.'}, status=400)

        with connection.cursor() as cursor:
            # Get logged-in teacher details
            cursor.execute(
                "SELECT subject, class_teacher_of FROM teachers WHERE id = %s",
                [request.session['teacher_id']]
            )
            teacher_row = cursor.fetchone()
            if not teacher_row:
                return JsonResponse({'success': False, 'message': 'Teacher not found.'}, status=403)

            assigned_subject, is_class_teacher_str = teacher_row
            expected_role = 'classTeacher' if is_class_teacher_str else 'subjectTeacher'

            # Validate role matches teacher's actual role
            if role != expected_role:
                return JsonResponse({'success': False, 'message': 'Invalid role for this teacher.'}, status=403)

            # Verify student is in the selected class/section
            cursor.execute(
                "SELECT user_id FROM student_page1 WHERE user_id = %s AND class = %s AND section = %s",
                [student_id, class_name, section]
            )
            if not cursor.fetchone():
                return JsonResponse({'success': False, 'message': 'Student not found in specified class/section.'}, status=400)

            # === CRITICAL FIX: Find the actual class teacher for this class-section ===
            class_section_key = f"{class_name}-{section}"
            cursor.execute(
                "SELECT id FROM teachers WHERE class_teacher_of = %s",
                [class_section_key]
            )
            class_teacher_row = cursor.fetchone()
            if not class_teacher_row:
                return JsonResponse({
                    'success': False,
                    'message': f'No class teacher assigned for {class_section_key}. Contact admin.'
                }, status=400)

            class_teacher_id = class_teacher_row[0]

            # Save Teacher Signature under the CLASS TEACHER's ID (not necessarily the logged-in one)
            cursor.execute(
                """INSERT INTO teacher_signature (teacher_id, signature) 
                   VALUES (%s, %s) 
                   ON DUPLICATE KEY UPDATE signature = %s""",
                [class_teacher_id, teacher_signature_data, teacher_signature_data]
            )

            # Save Principal Signature globally (fixed principal_id = 1)
            cursor.execute(
                """INSERT INTO principal_signature (principal_id, signature) 
                   VALUES (1, %s) 
                   ON DUPLICATE KEY UPDATE signature = %s""",
                [principal_signature_data, principal_signature_data]
            )

            # Grade calculation
            def calculate_grade(marks, max_marks):
                if max_marks == 0:
                    return 'E'
                percentage = (marks / max_marks) * 100
                if percentage >= 80:
                    return 'A'
                elif percentage >= 60:
                    return 'B'
                elif percentage >= 40:
                    return 'C'
                elif percentage >= 33:
                    return 'D'
                else:
                    return 'E'

            marks_updated = False
            marks_inserted = False

            # Subject Teacher: Only one subject allowed
            if role == 'subjectTeacher':
                if len(marks_data) > 1:
                    return JsonResponse({'success': False, 'message': 'Subject teachers can only submit one subject.'}, status=400)

                entry = marks_data[0]
                subject_id = entry.get('subjectId')
                marks_val = entry.get('marks')
                max_marks_val = entry.get('maxMarks')

                if not all([subject_id, marks_val is not None, max_marks_val]):
                    return JsonResponse({'success': False, 'message': 'Missing subject or marks data.'}, status=400)

                # Verify teacher is assigned to this subject
                cursor.execute(
                    "SELECT id FROM school_subjects WHERE id = %s AND class = %s AND name = %s",
                    [subject_id, class_name, assigned_subject]
                )
                if not cursor.fetchone():
                    return JsonResponse({'success': False, 'message': 'You are not authorized to enter marks for this subject.'}, status=403)

                marks = int(marks_val)
                max_marks = int(max_marks_val)
                if marks < 0 or marks > max_marks or max_marks < 1:
                    return JsonResponse({'success': False, 'message': 'Invalid marks value.'}, status=400)

                grade = calculate_grade(marks, max_marks)

                cursor.execute(
                    """INSERT INTO school_marks (student_id, subject_id, marks, max_marks, grade)
                       VALUES (%s, %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE marks=%s, max_marks=%s, grade=%s""",
                    [student_id, subject_id, marks, max_marks, grade, marks, max_marks, grade]
                )

                marks_inserted = True  # or check rowcount if needed

            # Class Teacher: Can enter all subjects
            else:
                for entry in marks_data:
                    subject_id = entry.get('subjectId')
                    marks_val = entry.get('marks')
                    max_marks_val = entry.get('maxMarks')

                    if not all([subject_id, marks_val is not None, max_marks_val]):
                        return JsonResponse({'success': False, 'message': 'Invalid data in marks entry.'}, status=400)

                    cursor.execute(
                        "SELECT id FROM school_subjects WHERE id = %s AND class = %s",
                        [subject_id, class_name]
                    )
                    if not cursor.fetchone():
                        return JsonResponse({'success': False, 'message': 'Subject not valid for this class.'}, status=400)

                    marks = int(marks_val)
                    max_marks = int(max_marks_val)
                    if marks < 0 or marks > max_marks or max_marks < 1:
                        return JsonResponse({'success': False, 'message': 'Invalid marks entered.'}, status=400)

                    grade = calculate_grade(marks, max_marks)

                    cursor.execute(
                        """INSERT INTO school_marks (student_id, subject_id, marks, max_marks, grade)
                           VALUES (%s, %s, %s, %s, %s)
                           ON DUPLICATE KEY UPDATE marks=%s, max_marks=%s, grade=%s""",
                        [student_id, subject_id, marks, max_marks, grade, marks, max_marks, grade]
                    )

                    marks_inserted = True

            connection.commit()

            # Success message
            message = "Marks saved successfully!"
            if marks_updated:
                message = "Marks updated successfully!"

            return JsonResponse({'success': True, 'message': message})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'}, status=400)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid number format in marks.'}, status=400)
    except Exception as e:
        connection.rollback()
        return JsonResponse({'success': False, 'message': f'Error saving marks: {str(e)}'}, status=500)

def teacher_add_subject(request):
    if 'teacher_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            subject_name = data.get('subjectName')
            max_marks = data.get('maxMarks')
            class_name = data.get('class')

            if not all([subject_name, max_marks, class_name]):
                return JsonResponse({'success': False, 'message': 'Subject name, max marks, and class required.'}, status=400)

            max_marks = int(max_marks)
            if max_marks < 1:
                return JsonResponse({'success': False, 'message': 'Max marks must be positive.'}, status=400)

            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM student_page1 WHERE class = %s", [class_name])
                if cursor.fetchone()[0] == 0:
                    return JsonResponse({'success': False, 'message': 'Invalid class.'}, status=400)

                cursor.execute(
                    "SELECT COUNT(*) FROM school_subjects WHERE name = %s AND class = %s",
                    [subject_name, class_name]
                )
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'success': False, 'message': 'Subject already exists for this class.'}, status=400)

                cursor.execute(
                    "INSERT INTO school_subjects (name, max_marks, class) VALUES (%s, %s, %s)",
                    [subject_name, max_marks, class_name]
                )
                connection.commit()

                cursor.execute(
                    "SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name",
                    [class_name]
                )
                subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]

                return JsonResponse({'success': True, 'message': f'Subject {subject_name} added for class {class_name}.', 'subjects': subjects})

        except Exception as e:
            connection.rollback()
            return JsonResponse({'success': False, 'message': f'Error adding subject: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

def teacher_delete_subject(request):
    if 'teacher_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            subject_id = data.get('subjectId')
            class_name = data.get('class')

            if not all([subject_id, class_name]):
                return JsonResponse({'success': False, 'message': 'Subject ID and class required.'}, status=400)

            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM student_page1 WHERE class = %s", [class_name])
                if cursor.fetchone()[0] == 0:
                    return JsonResponse({'success': False, 'message': 'Invalid class.'}, status=400)

                cursor.execute(
                    "SELECT name FROM school_subjects WHERE id = %s AND class = %s",
                    [subject_id, class_name]
                )
                subject = cursor.fetchone()
                if not subject:
                    return JsonResponse({'success': False, 'message': 'Subject not found for this class.'}, status=404)

                cursor.execute("DELETE FROM school_marks WHERE subject_id = %s", [subject_id])
                cursor.execute("DELETE FROM school_subjects WHERE id = %s", [subject_id])
                connection.commit()

                cursor.execute(
                    "SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name",
                    [class_name]
                )
                subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]

                return JsonResponse({'success': True, 'message': f'Subject {subject[0]} deleted for class {class_name}.', 'subjects': subjects})

        except Exception as e:
            connection.rollback()
            return JsonResponse({'success': False, 'message': f'Error deleting subject: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

def teacher_progress_card(request):
    if 'teacher_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('studentId')

            if not student_id:
                return JsonResponse({'success': False, 'message': 'Student ID required.'}, status=400)

            with connection.cursor() as cursor:
                # Fetch student details including profile picture and admission_number fallback
                try:
                    cursor.execute(
                        """
                        SELECT s.name, s.roll_number, s.class, s.section, COALESCE(s.admission_number, '') AS admission_number, p.image_path
                        FROM student_page1 s LEFT JOIN profile_pics p ON s.user_id = p.user_id
                        WHERE s.user_id = %s
                        """,
                        [student_id]
                    )
                    student_row = cursor.fetchone()
                    if student_row:
                        name, roll_number, class_name, section, admission_number, image_path = student_row
                    else:
                        name = roll_number = class_name = section = admission_number = image_path = ''
                except Exception as col_err:
                    if "Unknown column 'admission_number'" in str(col_err):
                        cursor.execute(
                            """
                            SELECT s.name, s.roll_number, s.class, s.section, p.image_path
                            FROM student_page1 s LEFT JOIN profile_pics p ON s.user_id = p.user_id
                            WHERE s.user_id = %s
                            """,
                            [student_id]
                        )
                        student_row = cursor.fetchone()
                        if student_row:
                            name, roll_number, class_name, section, image_path = student_row
                            admission_number = ''
                        else:
                            name = roll_number = class_name = section = admission_number = image_path = ''
                    else:
                        raise col_err

                if not name:
                    return JsonResponse({'success': False, 'message': 'Student not found.'}, status=404)

                # Fetch marks
                cursor.execute(
                    """
                    SELECT ss.name, m.marks, m.max_marks, m.grade
                    FROM school_marks m JOIN school_subjects ss ON m.subject_id = ss.id
                    WHERE m.student_id = %s AND ss.class = %s
                    """,
                    [student_id, class_name]
                )
                marks = [{'subject': row[0], 'marks': row[1] or 0, 'max_marks': row[2], 'grade': row[3] or 'E'} for row in cursor.fetchall()]

                # Fetch teacher signature
                cursor.execute(
                    "SELECT signature FROM teacher_signature WHERE teacher_id = %s",
                    [request.session['teacher_id']]
                )
                signature_row = cursor.fetchone()
                signature = signature_row[0] if signature_row else None

                # Fetch principal signature (using teacher_id as proxy, similar to admin setup)
                cursor.execute(
                    "SELECT signature FROM principal_signature WHERE principal_id = %s",
                    [request.session['teacher_id']]
                )
                principal_signature_row = cursor.fetchone()
                principal_signature = principal_signature_row[0] if principal_signature_row else None

                # Fetch teacher name
                cursor.execute("SELECT name FROM teachers WHERE id = %s", [request.session['teacher_id']])
                teacher_name = cursor.fetchone()[0]

                # Calculate totals and grade
                total_marks = sum(mark['marks'] for mark in marks)
                total_max_marks = sum(mark['max_marks'] for mark in marks)
                percentage = (total_marks / total_max_marks * 100) if total_max_marks > 0 else 0
                grade = 'A' if percentage >= 80 else 'B' if percentage >= 60 else 'C' if percentage >= 40 else 'D' if percentage >= 33 else 'E'
                passed = all(mark['marks'] >= 0.33 * mark['max_marks'] for mark in marks) if marks else False

                # Prepare response
                response_data = {
                    'success': True,
                    'student': {
                        'name': name,
                        'roll_number': roll_number,
                        'class': class_name,
                        'section': section,
                        'admission_number': admission_number,
                        'image_path': image_path or 'pfpics/default_profile.jpg'  # Fallback to default
                    },
                    'marks': marks,
                    'total_marks': total_marks,
                    'total_max_marks': total_max_marks,
                    'percentage': round(percentage, 2),
                    'grade': grade,
                    'status': 'Pass' if passed else 'Fail',
                    'teacher_name': teacher_name
                }
                if signature:
                    response_data['signature'] = signature
                if principal_signature:
                    response_data['principal_signature'] = principal_signature
                return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error fetching progress card: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

def teacher_get_students(request):
    if 'teacher_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    try:
        class_name = request.GET.get('class')
        section = request.GET.get('section')

        if not class_name or not section:
            return JsonResponse({'success': False, 'message': 'Class and section required.'}, status=400)

        with connection.cursor() as cursor:
            # Students with admission_number fallback (and class/section for display)
            try:
                cursor.execute(
                    """
                    SELECT user_id AS id, name, roll_number, class, section, 
                           COALESCE(admission_number, '') AS admission_number 
                    FROM student_page1 
                    WHERE class = %s AND section = %s 
                    ORDER BY name
                    """,
                    [class_name, section]
                )
                students = [
                    {
                        'id': row[0], 
                        'name': row[1], 
                        'roll_number': row[2], 
                        'class': row[3], 
                        'section': row[4], 
                        'admission_number': row[5]
                    } 
                    for row in cursor.fetchall()
                ]
            except Exception as col_err:
                if "Unknown column 'admission_number'" in str(col_err):
                    cursor.execute(
                        """
                        SELECT user_id AS id, name, roll_number, class, section
                        FROM student_page1 
                        WHERE class = %s AND section = %s 
                        ORDER BY name
                        """,
                        [class_name, section]
                    )
                    students = [
                        {
                            'id': row[0], 
                            'name': row[1], 
                            'roll_number': row[2], 
                            'class': row[3], 
                            'section': row[4], 
                            'admission_number': ''
                        } 
                        for row in cursor.fetchall()
                    ]
                else:
                    raise col_err

            return JsonResponse({'success': True, 'students': students})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error fetching students: {str(e)}'}, status=500)

def teacher_get_subjects(request):
    if 'teacher_id' not in request.session:
        return JsonResponse({'success': False, 'message': 'Please log in.'}, status=403)

    try:
        class_name = request.GET.get('class')
        if not class_name:
            return JsonResponse({'success': False, 'message': 'Class required.'}, status=400)

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, name, max_marks FROM school_subjects WHERE class = %s ORDER BY name",
                [class_name]
            )
            subjects = [{'id': row[0], 'name': row[1], 'max_marks': row[2]} for row in cursor.fetchall()]
            return JsonResponse({'success': True, 'subjects': subjects})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error fetching subjects: {str(e)}'}, status=500)




from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.db import connection

# Helper function to fetch timetable data
def fetch_timetable_data(query, params):
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()
        return [
            {
                'id': row[0], 'class_id': row[1], 'subject': row[2], 
                'teacher_name': row[3] if len(row) > 3 else None, 
                'day_of_week': row[4], 'start_time': row[5], 
                'end_time': row[6], 'room': row[7] if len(row) > 7 else None
            } for row in rows
        ]

# Admin Timetable Dashboard (with filters)
# First, update admin_timetable_view to fetch exams and subjects
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, IntegrityError
from datetime import datetime, timedelta

def admin_timetable_view(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT id, name FROM teachers")
        teachers = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT subject FROM exams ORDER BY subject")
        subjects = [row[0] for row in cursor.fetchall() if row[0]]
        
        cursor.execute("""
            SELECT e.id, e.class_id, e.subject, e.exam_date, e.start_time, 
                   e.end_time, e.room, COALESCE(tch.name, 'N/A') as invigilator_name
            FROM exams e
            LEFT JOIN teachers tch ON e.invigilator_id = tch.id
            ORDER BY e.exam_date, e.start_time
        """)
        exams = []
        for row in cursor.fetchall():
            exams.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'exam_date': row[3], 'start_time': row[4], 'end_time': row[5],
                'room': row[6], 'invigilator_name': row[7]
            })
        
        cursor.execute("""
            SELECT t.id, t.class_id, t.subject, COALESCE(tch.name, 'N/A') as teacher_name, 
                   t.day_of_week, t.start_time, t.end_time, COALESCE(t.room, 'N/A') as room,
                   t.week_start_date, t.week_end_date
            FROM timetable t
            LEFT JOIN teachers tch ON t.teacher_id = tch.id
            ORDER BY t.week_start_date DESC, t.day_of_week, t.start_time
        """)
        timetables = []
        for row in cursor.fetchall():
            timetables.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'teacher_name': row[3], 'day_of_week': row[4],
                'start_time': row[5], 'end_time': row[6], 'room': row[7],
                'week_start_date': row[8], 'week_end_date': row[9]
            })
    
    return render(request, 'users/admin_timetable.html', {
        'classes': classes, 'sections': sections, 'teachers': teachers,
        'subjects': subjects, 'exams': exams, 'timetables': timetables
    })

def admin_exam_filter(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    class_filter = request.GET.get('class', '')
    subject_filter = request.GET.get('subject', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    invigilator_id = request.GET.get('invigilator', '')
    
    query = """
        SELECT e.id, e.class_id, e.subject, e.exam_date, e.start_time, 
               e.end_time, e.room, COALESCE(tch.name, 'N/A') as invigilator_name
        FROM exams e
        LEFT JOIN teachers tch ON e.invigilator_id = tch.id
        WHERE 1=1
    """
    params = []
    
    if class_filter:
        query += " AND e.class_id LIKE %s"
        params.append(f"{class_filter}%")
    if subject_filter:
        query += " AND e.subject = %s"
        params.append(subject_filter)
    if start_date and end_date:
        query += " AND e.exam_date BETWEEN %s AND %s"
        params.append(start_date)
        params.append(end_date)
    elif start_date:
        query += " AND e.exam_date >= %s"
        params.append(start_date)
    elif end_date:
        query += " AND e.exam_date <= %s"
        params.append(end_date)
    if invigilator_id:
        query += " AND e.invigilator_id = %s"
        params.append(invigilator_id)
    
    query += " ORDER BY e.exam_date, e.start_time"
    
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        exams = []
        for row in cursor.fetchall():
            exams.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'exam_date': row[3], 'start_time': row[4], 'end_time': row[5],
                'room': row[6], 'invigilator_name': row[7]
            })
        
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT id, name FROM teachers")
        teachers = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT subject FROM exams ORDER BY subject")
        subjects = [row[0] for row in cursor.fetchall() if row[0]]
        
        cursor.execute("""
            SELECT t.id, t.class_id, t.subject, COALESCE(tch.name, 'N/A') as teacher_name, 
                   t.day_of_week, t.start_time, t.end_time, COALESCE(t.room, 'N/A') as room,
                   t.week_start_date, t.week_end_date
            FROM timetable t
            LEFT JOIN teachers tch ON t.teacher_id = tch.id
            ORDER BY t.week_start_date DESC, t.day_of_week, t.start_time
        """)
        timetables = []
        for row in cursor.fetchall():
            timetables.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'teacher_name': row[3], 'day_of_week': row[4],
                'start_time': row[5], 'end_time': row[6], 'room': row[7],
                'week_start_date': row[8], 'week_end_date': row[9]
            })
    
    return render(request, 'users/admin_timetable.html', {
        'exams': exams, 'classes': classes, 'sections': sections, 
        'teachers': teachers, 'subjects': subjects, 'timetables': timetables,
        'selected_exam_class': class_filter,
        'selected_exam_subject': subject_filter,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'selected_invigilator': invigilator_id
    })

def admin_timetable_filter(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    class_filter = request.GET.get('class', '')
    section_filter = request.GET.get('section', '')
    teacher_id = request.GET.get('teacher_id', '')
    day_filter = request.GET.get('day', '')
    week_filter = request.GET.get('week', '')
    
    query = """
        SELECT t.id, t.class_id, t.subject, tch.name, t.day_of_week, 
               t.start_time, t.end_time, t.room, t.week_start_date, t.week_end_date
        FROM timetable t
        JOIN teachers tch ON t.teacher_id = tch.id
        WHERE 1=1
    """
    params = []
    
    if class_filter and section_filter:
        class_id = f"{class_filter}{section_filter}"
        query += " AND t.class_id = %s"
        params.append(class_id)
    elif class_filter:
        query += " AND t.class_id LIKE %s"
        params.append(f"{class_filter}%")
    if teacher_id:
        query += " AND t.teacher_id = %s"
        params.append(teacher_id)
    if day_filter:
        query += " AND t.day_of_week = %s"
        params.append(day_filter)
    if week_filter:
        query += " AND t.week_start_date = %s"
        params.append(week_filter)
    
    query += " ORDER BY t.week_start_date DESC, t.day_of_week, t.start_time"
    
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        timetables = []
        for row in cursor.fetchall():
            timetables.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'teacher_name': row[3], 'day_of_week': row[4],
                'start_time': row[5], 'end_time': row[6], 'room': row[7],
                'week_start_date': row[8], 'week_end_date': row[9]
            })
        
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT id, name FROM teachers")
        teachers = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT week_start_date FROM timetable ORDER BY week_start_date DESC")
        weeks = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/admin_timetable.html', {
        'timetables': timetables, 'classes': classes, 'sections': sections, 
        'teachers': teachers, 'selected_class': class_filter, 
        'selected_section': section_filter, 'selected_teacher': teacher_id, 
        'selected_day': day_filter, 'weeks': weeks, 'selected_week': week_filter
    })

def admin_timetable_add(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    if request.method == 'POST':
        class_name = request.POST.get('class')
        section = request.POST.get('section')
        subject = request.POST.get('subject')
        teacher_id = request.POST.get('teacher_id')
        day_of_week = request.POST.get('day_of_week')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        room = request.POST.get('room')
        week_start_date = request.POST.get('week_start_date')

        class_id = f"{class_name}{section}" if section else class_name

        if not class_name or not class_id:
            messages.error(request, 'Please select a valid class.')
            return redirect('admin_timetable_add')
        if not subject or not teacher_id or not day_of_week or not start_time or not end_time or not week_start_date:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('admin_timetable_add')

        week_start = datetime.strptime(week_start_date, '%Y-%m-%d').date()
        week_end = week_start + timedelta(days=6)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM timetable 
                WHERE (class_id = %s OR teacher_id = %s)
                AND day_of_week = %s
                AND start_time <= %s AND end_time >= %s
                AND week_start_date = %s
            """, [class_id, teacher_id, day_of_week, end_time, start_time, week_start])
            conflict = cursor.fetchone()
            
            if conflict:
                messages.error(request, 'Scheduling conflict detected for the selected week.')
                return redirect('admin_timetable')
            
            cursor.execute("""
                INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                     start_time, end_time, room, week_start_date, week_end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [class_id, subject, teacher_id, day_of_week, start_time, end_time, room or None, week_start, week_end])
        
        messages.success(request, 'Timetable entry added successfully.')
        return redirect('admin_timetable')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, name, subject FROM teachers")
        teachers = [{'id': row[0], 'name': row[1], 'subject': row[2]} for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/admin_timetable_add.html', {
        'teachers': teachers, 'classes': classes, 'sections': sections
    })

def admin_timetable_edit(request, id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, class_id, subject, teacher_id, day_of_week, 
                   start_time, end_time, room, week_start_date, week_end_date
            FROM timetable WHERE id = %s
        """, [id])
        timetable = cursor.fetchone()
        if not timetable:
            messages.error(request, 'Timetable entry not found.')
            return redirect('admin_timetable')
        
        class_id = timetable[1]
        class_name = class_id[:-1] if class_id[-1].isalpha() else class_id
        section = class_id[-1] if class_id[-1].isalpha() else ''
        
        timetable_data = {
            'id': timetable[0], 'class_id': timetable[1], 'subject': timetable[2],
            'teacher_id': timetable[3], 'day_of_week': timetable[4],
            'start_time': timetable[5], 'end_time': timetable[6], 'room': timetable[7],
            'week_start_date': timetable[8], 'week_end_date': timetable[9],
            'class': class_name, 'section': section
        }
        
        if request.method == 'POST':
            class_name = request.POST.get('class')
            section = request.POST.get('section')
            class_id = f"{class_name}{section}" if section else class_name
            subject = request.POST.get('subject')
            teacher_id = request.POST.get('teacher_id')
            day_of_week = request.POST.get('day_of_week')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            room = request.POST.get('room')
            week_start_date = request.POST.get('week_start_date')
            
            if not class_name or not class_id:
                messages.error(request, 'Please select a valid class.')
                return redirect('admin_timetable_edit', id=id)
            if not subject or not teacher_id or not day_of_week or not start_time or not end_time or not week_start_date:
                messages.error(request, 'Please fill in all required fields.')
                return redirect('admin_timetable_edit', id=id)
            
            week_start = datetime.strptime(week_start_date, '%Y-%m-%d').date()
            week_end = week_start + timedelta(days=6)
            
            cursor.execute("""
                SELECT id FROM timetable 
                WHERE (class_id = %s OR teacher_id = %s)
                AND day_of_week = %s
                AND start_time <= %s AND end_time >= %s
                AND week_start_date = %s
                AND id != %s
            """, [class_id, teacher_id, day_of_week, end_time, start_time, week_start, id])
            conflict = cursor.fetchone()
            
            if conflict:
                messages.error(request, 'Scheduling conflict detected.')
                return redirect('admin_timetable_edit', id=id)
            
            try:
                cursor.execute("""
                    UPDATE timetable 
                    SET class_id = %s, subject = %s, teacher_id = %s, 
                        day_of_week = %s, start_time = %s, end_time = %s, room = %s,
                        week_start_date = %s, week_end_date = %s
                    WHERE id = %s
                """, [class_id, subject, teacher_id, day_of_week, start_time, end_time, room or None, week_start, week_end, id])
            except IntegrityError as e:
                messages.error(request, f'Error updating timetable entry: {str(e)}')
                return redirect('admin_timetable_edit', id=id)
            
            messages.success(request, 'Timetable entry updated successfully.')
            return redirect('admin_timetable')
        
        cursor.execute("SELECT id, name, subject FROM teachers")
        teachers = [{'id': row[0], 'name': row[1], 'subject': row[2]} for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/admin_timetable_edit.html', {
        'timetable': timetable_data, 'teachers': teachers, 
        'classes': classes, 'sections': sections
    })

def admin_timetable_delete(request, id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM timetable WHERE id = %s", [id])
    
    messages.success(request, 'Timetable entry deleted successfully.')
    return redirect('admin_timetable')

def admin_timetable_weekly(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    if request.method == 'POST':
        class_name = request.POST.get('class')
        section = request.POST.get('section')
        class_id = f"{class_name}{section}" if section else class_name
        week_start_date = request.POST.get('week_start_date')
        num_weeks = int(request.POST.get('num_weeks', 1))
        num_periods = int(request.POST.get('num_periods', 6))
        
        if not class_name or not class_id:
            messages.error(request, 'Please select a valid class.')
            return redirect('admin_timetable_weekly')
        
        if not week_start_date:
            messages.error(request, 'Please select a start date for the week.')
            return redirect('admin_timetable_weekly')
        
        week_start = datetime.strptime(week_start_date, '%Y-%m-%d').date()
        
        with connection.cursor() as cursor:
            for week_offset in range(num_weeks):
                current_week_start = week_start + timedelta(weeks=week_offset)
                current_week_end = current_week_start + timedelta(days=6)
                
                for day in days:
                    for period in range(1, num_periods + 1):
                        subject = request.POST.get(f'subject_{day}_{period}')
                        teacher_id = request.POST.get(f'teacher_{day}_{period}')
                        start_time = request.POST.get(f'start_time_{day}_{period}')
                        end_time = request.POST.get(f'end_time_{day}_{period}')
                        room = request.POST.get(f'room_{day}_{period}')
                        
                        if not (subject and teacher_id and start_time and end_time):
                            continue
                        
                        cursor.execute("""
                            SELECT id FROM timetable 
                            WHERE (class_id = %s OR teacher_id = %s)
                            AND day_of_week = %s
                            AND start_time <= %s AND end_time >= %s
                            AND week_start_date = %s
                        """, [class_id, teacher_id, day, end_time, start_time, current_week_start])
                        conflict = cursor.fetchone()
                        
                        if conflict:
                            messages.warning(request, f'Conflict detected for {day} period {period} in week starting {current_week_start}. Skipped.')
                            continue
                        
                        try:
                            cursor.execute("""
                                INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                                     start_time, end_time, room, week_start_date, week_end_date)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, [class_id, subject, teacher_id, day, start_time, end_time, room or None, current_week_start, current_week_end])
                        except IntegrityError as e:
                            messages.warning(request, f'Error adding timetable for {day} period {period} in week {current_week_start}: {str(e)}')
                            continue
        
        messages.success(request, f'Weekly timetable created successfully for {num_weeks} week(s).')
        return redirect('admin_timetable')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, name, subject FROM teachers")
        teachers = [{'id': row[0], 'name': row[1], 'subject': row[2]} for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/admin_timetable_weekly.html', {
        'days': days, 'teachers': teachers, 
        'classes': classes, 'sections': sections
    })

def admin_timetable_copy_week(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    if request.method == 'POST':
        source_week = request.POST.get('source_week')
        target_week = request.POST.get('target_week')
        num_weeks = int(request.POST.get('num_weeks', 1))
        class_filter = request.POST.get('class_filter', '')
        
        if not source_week or not target_week:
            messages.error(request, 'Please select both source and target weeks.')
            return redirect('admin_timetable')
        
        source_date = datetime.strptime(source_week, '%Y-%m-%d').date()
        target_date = datetime.strptime(target_week, '%Y-%m-%d').date()
        
        with connection.cursor() as cursor:
            query = "SELECT * FROM timetable WHERE week_start_date = %s"
            params = [source_date]
            
            if class_filter:
                query += " AND class_id LIKE %s"
                params.append(f"{class_filter}%")
            
            cursor.execute(query, params)
            source_timetables = cursor.fetchall()
            
            if not source_timetables:
                messages.error(request, 'No timetable entries found for the selected source week.')
                return redirect('admin_timetable')
            
            for week_offset in range(num_weeks):
                current_target = target_date + timedelta(weeks=week_offset)
                current_target_end = current_target + timedelta(days=6)
                
                for entry in source_timetables:
                    try:
                        cursor.execute("""
                            INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                                 start_time, end_time, room, week_start_date, week_end_date)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, [entry[1], entry[2], entry[3], entry[4], entry[5], entry[6], entry[7], current_target, current_target_end])
                    except IntegrityError:
                        continue
            
            messages.success(request, f'Timetable copied successfully to {num_weeks} week(s).')
    
    return redirect('admin_timetable')

def admin_timetable_bulk_delete(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.error(request, 'No entries selected.')
            return redirect('admin_timetable')
        
        with connection.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(ids))
            cursor.execute(f"DELETE FROM timetable WHERE id IN ({placeholders})", ids)
        
        messages.success(request, f'{len(ids)} timetable entries deleted successfully.')
    
    return redirect('admin_timetable')


def admin_timetable_bulk_copy(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    ids = request.GET.get('ids', '').split(',')
    if not ids or ids == ['']:
        messages.error(request, 'No entries selected.')
        return redirect('admin_timetable')
    
    with connection.cursor() as cursor:
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            SELECT class_id, subject, teacher_id, day_of_week, 
                   start_time, end_time, room, week_start_date
            FROM timetable 
            WHERE id IN ({placeholders})
        """, ids)
        
        entries = cursor.fetchall()
        
        for entry in entries:
            week_start = entry[7] + timedelta(weeks=1)  # Next week
            week_end = week_start + timedelta(days=6)
            
            try:
                cursor.execute("""
                    INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                         start_time, end_time, room, week_start_date, week_end_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [entry[0], entry[1], entry[2], entry[3], entry[4], entry[5], 
                      entry[6], week_start, week_end])
            except IntegrityError:
                continue
        
        messages.success(request, f'{len(entries)} entries copied to next week.')
    
    return redirect('admin_timetable')


import openpyxl
from django.http import HttpResponse

def admin_timetable_export_excel(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    ids = request.GET.get('ids', '')
    
    with connection.cursor() as cursor:
        if ids:
            id_list = ids.split(',')
            placeholders = ','.join(['%s'] * len(id_list))
            query = f"""
                SELECT t.class_id, t.subject, tch.name, t.day_of_week, 
                       t.start_time, t.end_time, t.room
                FROM timetable t
                LEFT JOIN teachers tch ON t.teacher_id = tch.id
                WHERE t.id IN ({placeholders})
                ORDER BY t.class_id, t.day_of_week, t.start_time
            """
            cursor.execute(query, id_list)
        else:
            cursor.execute("""
                SELECT t.class_id, t.subject, tch.name, t.day_of_week, 
                       t.start_time, t.end_time, t.room
                FROM timetable t
                LEFT JOIN teachers tch ON t.teacher_id = tch.id
                ORDER BY t.class_id, t.day_of_week, t.start_time
            """)
        
        entries = cursor.fetchall()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Headers
    headers = ['Class', 'Subject', 'Teacher', 'Day', 'Start Time', 'End Time', 'Room']
    ws.append(headers)
    
    # Data
    for entry in entries:
        ws.append(list(entry))
    
    # Style headers
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="00a676", end_color="00a676", fill_type="solid")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=timetable.xlsx'
    
    wb.save(response)
    return response




    # Teacher Timetable View (with integrated filtering)
from django.contrib import messages
from django.db import connection, IntegrityError
from django.shortcuts import render, redirect
from datetime import datetime, timedelta


def teacher_timetable_view(request):
    """
    Enhanced teacher timetable view with filtering similar to admin
    """
    if 'teacher_id' not in request.session:
        messages.error(request, 'Please log in to access the teacher portal.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    
    with connection.cursor() as cursor:
        # Fetch distinct values for filters
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        
        # Fetch all timetable entries for this teacher
        cursor.execute("""
            SELECT t.id, t.class_id, t.subject, t.day_of_week, 
                   t.start_time, t.end_time, COALESCE(t.room, 'N/A') as room,
                   t.week_start_date, t.week_end_date
            FROM timetable t
            WHERE t.teacher_id = %s
            ORDER BY t.week_start_date DESC, 
                     FIELD(t.day_of_week, 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'),
                     t.start_time
        """, [teacher_id])
        
        timetables = []
        for row in cursor.fetchall():
            timetables.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'day_of_week': row[3], 'start_time': row[4], 'end_time': row[5],
                'room': row[6], 'week_start_date': row[7], 'week_end_date': row[8]
            })
        
        # Get distinct weeks for week filter
        cursor.execute("""
            SELECT DISTINCT week_start_date 
            FROM timetable 
            WHERE teacher_id = %s 
            ORDER BY week_start_date DESC
        """, [teacher_id])
        weeks = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/teacher_timetable.html', {
        'timetables': timetables, 
        'classes': classes, 
        'sections': sections,
        'weeks': weeks,
        'teacher_id': teacher_id
    })


def teacher_timetable_filter(request):
    """
    Filter teacher's timetable entries
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    class_filter = request.GET.get('class', '')
    section_filter = request.GET.get('section', '')
    day_filter = request.GET.get('day', '')
    week_filter = request.GET.get('week', '')
    
    query = """
        SELECT t.id, t.class_id, t.subject, t.day_of_week, 
               t.start_time, t.end_time, COALESCE(t.room, 'N/A') as room,
               t.week_start_date, t.week_end_date
        FROM timetable t
        WHERE t.teacher_id = %s
    """
    params = [teacher_id]
    
    if class_filter and section_filter:
        class_id = f"{class_filter}{section_filter}"
        query += " AND t.class_id = %s"
        params.append(class_id)
    elif class_filter:
        query += " AND t.class_id LIKE %s"
        params.append(f"{class_filter}%")
    
    if day_filter:
        query += " AND t.day_of_week = %s"
        params.append(day_filter)
    
    if week_filter:
        query += " AND t.week_start_date = %s"
        params.append(week_filter)
    
    query += """ 
        ORDER BY t.week_start_date DESC, 
                 FIELD(t.day_of_week, 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'),
                 t.start_time
    """
    
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        timetables = []
        for row in cursor.fetchall():
            timetables.append({
                'id': row[0], 'class_id': row[1], 'subject': row[2],
                'day_of_week': row[3], 'start_time': row[4], 'end_time': row[5],
                'room': row[6], 'week_start_date': row[7], 'week_end_date': row[8]
            })
        
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("""
            SELECT DISTINCT week_start_date 
            FROM timetable 
            WHERE teacher_id = %s 
            ORDER BY week_start_date DESC
        """, [teacher_id])
        weeks = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/teacher_timetable.html', {
        'timetables': timetables, 
        'classes': classes, 
        'sections': sections,
        'weeks': weeks,
        'selected_class': class_filter, 
        'selected_section': section_filter, 
        'selected_day': day_filter, 
        'selected_week': week_filter
    })


def teacher_timetable_add(request):
    """
    Add single timetable entry for teacher
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    
    if request.method == 'POST':
        class_name = request.POST.get('class')
        section = request.POST.get('section')
        subject = request.POST.get('subject')
        day_of_week = request.POST.get('day_of_week')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        room = request.POST.get('room')
        week_start_date = request.POST.get('week_start_date')

        class_id = f"{class_name}{section}" if section else class_name

        if not class_name or not class_id:
            messages.error(request, 'Please select a valid class.')
            return redirect('teacher_timetable_add')
        if not subject or not day_of_week or not start_time or not end_time or not week_start_date:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('teacher_timetable_add')

        week_start = datetime.strptime(week_start_date, '%Y-%m-%d').date()
        week_end = week_start + timedelta(days=6)

        with connection.cursor() as cursor:
            # Check for conflicts
            cursor.execute("""
                SELECT id FROM timetable 
                WHERE (class_id = %s OR teacher_id = %s)
                AND day_of_week = %s
                AND start_time <= %s AND end_time >= %s
                AND week_start_date = %s
            """, [class_id, teacher_id, day_of_week, end_time, start_time, week_start])
            conflict = cursor.fetchone()
            
            if conflict:
                messages.error(request, 'Scheduling conflict detected for the selected week.')
                return redirect('teacher_timetable')
            
            cursor.execute("""
                INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                     start_time, end_time, room, week_start_date, week_end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [class_id, subject, teacher_id, day_of_week, start_time, end_time, room or None, week_start, week_end])
        
        messages.success(request, 'Timetable entry added successfully.')
        return redirect('teacher_timetable')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/teacher_timetable_add.html', {
        'classes': classes, 'sections': sections
    })


def teacher_timetable_edit(request, id):
    """
    Edit timetable entry - teacher can only edit their own entries
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, class_id, subject, teacher_id, day_of_week, 
                   start_time, end_time, room, week_start_date, week_end_date
            FROM timetable WHERE id = %s AND teacher_id = %s
        """, [id, teacher_id])
        timetable = cursor.fetchone()
        
        if not timetable:
            messages.error(request, 'Timetable entry not found or you do not have permission.')
            return redirect('teacher_timetable')
        
        class_id = timetable[1]
        class_name = class_id[:-1] if class_id[-1].isalpha() else class_id
        section = class_id[-1] if class_id[-1].isalpha() else ''
        
        timetable_data = {
            'id': timetable[0], 'class_id': timetable[1], 'subject': timetable[2],
            'teacher_id': timetable[3], 'day_of_week': timetable[4],
            'start_time': timetable[5], 'end_time': timetable[6], 'room': timetable[7],
            'week_start_date': timetable[8], 'week_end_date': timetable[9],
            'class': class_name, 'section': section
        }
        
        if request.method == 'POST':
            class_name = request.POST.get('class')
            section = request.POST.get('section')
            class_id = f"{class_name}{section}" if section else class_name
            subject = request.POST.get('subject')
            day_of_week = request.POST.get('day_of_week')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            room = request.POST.get('room')
            week_start_date = request.POST.get('week_start_date')
            
            if not class_name or not class_id:
                messages.error(request, 'Please select a valid class.')
                return redirect('teacher_timetable_edit', id=id)
            if not subject or not day_of_week or not start_time or not end_time or not week_start_date:
                messages.error(request, 'Please fill in all required fields.')
                return redirect('teacher_timetable_edit', id=id)
            
            week_start = datetime.strptime(week_start_date, '%Y-%m-%d').date()
            week_end = week_start + timedelta(days=6)
            
            cursor.execute("""
                SELECT id FROM timetable 
                WHERE (class_id = %s OR teacher_id = %s)
                AND day_of_week = %s
                AND start_time <= %s AND end_time >= %s
                AND week_start_date = %s
                AND id != %s
            """, [class_id, teacher_id, day_of_week, end_time, start_time, week_start, id])
            conflict = cursor.fetchone()
            
            if conflict:
                messages.error(request, 'Scheduling conflict detected.')
                return redirect('teacher_timetable_edit', id=id)
            
            try:
                cursor.execute("""
                    UPDATE timetable 
                    SET class_id = %s, subject = %s, teacher_id = %s, 
                        day_of_week = %s, start_time = %s, end_time = %s, room = %s,
                        week_start_date = %s, week_end_date = %s
                    WHERE id = %s
                """, [class_id, subject, teacher_id, day_of_week, start_time, end_time, room or None, week_start, week_end, id])
            except IntegrityError as e:
                messages.error(request, f'Error updating timetable entry: {str(e)}')
                return redirect('teacher_timetable_edit', id=id)
            
            messages.success(request, 'Timetable entry updated successfully.')
            return redirect('teacher_timetable')
        
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/teacher_timetable_edit.html', {
        'timetable': timetable_data, 
        'classes': classes, 'sections': sections
    })


def teacher_timetable_delete(request, id):
    """
    Delete timetable entry - teacher can only delete their own entries
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM timetable WHERE id = %s AND teacher_id = %s", [id, teacher_id])
        if cursor.rowcount == 0:
            messages.error(request, 'Timetable entry not found or you do not have permission.')
        else:
            messages.success(request, 'Timetable entry deleted successfully.')
    
    return redirect('teacher_timetable')


def teacher_timetable_weekly(request):
    """
    Create weekly timetable - similar to admin but auto-assigns teacher_id
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    if request.method == 'POST':
        class_name = request.POST.get('class')
        section = request.POST.get('section')
        class_id = f"{class_name}{section}" if section else class_name
        week_start_date = request.POST.get('week_start_date')
        num_weeks = int(request.POST.get('num_weeks', 1))
        num_periods = int(request.POST.get('num_periods', 6))
        
        if not class_name or not class_id:
            messages.error(request, 'Please select a valid class.')
            return redirect('teacher_timetable_weekly')
        
        if not week_start_date:
            messages.error(request, 'Please select a start date for the week.')
            return redirect('teacher_timetable_weekly')
        
        week_start = datetime.strptime(week_start_date, '%Y-%m-%d').date()
        
        with connection.cursor() as cursor:
            success_count = 0
            for week_offset in range(num_weeks):
                current_week_start = week_start + timedelta(weeks=week_offset)
                current_week_end = current_week_start + timedelta(days=6)
                
                for day in days:
                    for period in range(1, num_periods + 1):
                        subject = request.POST.get(f'subject_{day}_{period}')
                        start_time = request.POST.get(f'start_time_{day}_{period}')
                        end_time = request.POST.get(f'end_time_{day}_{period}')
                        room = request.POST.get(f'room_{day}_{period}')
                        
                        if not (subject and start_time and end_time):
                            continue
                        
                        cursor.execute("""
                            SELECT id FROM timetable 
                            WHERE (class_id = %s OR teacher_id = %s)
                            AND day_of_week = %s
                            AND start_time <= %s AND end_time >= %s
                            AND week_start_date = %s
                        """, [class_id, teacher_id, day, end_time, start_time, current_week_start])
                        conflict = cursor.fetchone()
                        
                        if conflict:
                            messages.warning(request, f'Conflict detected for {day} period {period} in week starting {current_week_start}. Skipped.')
                            continue
                        
                        try:
                            cursor.execute("""
                                INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                                     start_time, end_time, room, week_start_date, week_end_date)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, [class_id, subject, teacher_id, day, start_time, end_time, room or None, current_week_start, current_week_end])
                            success_count += 1
                        except IntegrityError as e:
                            messages.warning(request, f'Error adding timetable for {day} period {period} in week {current_week_start}: {str(e)}')
                            continue
        
        if success_count > 0:
            messages.success(request, f'Weekly timetable created successfully for {num_weeks} week(s) with {success_count} entries.')
        return redirect('teacher_timetable')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
    
    return render(request, 'users/teacher_timetable_weekly.html', {
        'days': days, 
        'classes': classes, 
        'sections': sections
    })


def teacher_timetable_copy_week(request):
    """
    Copy timetable from one week to another
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    
    if request.method == 'POST':
        source_week = request.POST.get('source_week')
        target_week = request.POST.get('target_week')
        num_weeks = int(request.POST.get('num_weeks', 1))
        
        if not source_week or not target_week:
            messages.error(request, 'Please select both source and target weeks.')
            return redirect('teacher_timetable')
        
        source_date = datetime.strptime(source_week, '%Y-%m-%d').date()
        target_date = datetime.strptime(target_week, '%Y-%m-%d').date()
        
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM timetable 
                WHERE week_start_date = %s AND teacher_id = %s
            """, [source_date, teacher_id])
            source_timetables = cursor.fetchall()
            
            if not source_timetables:
                messages.error(request, 'No timetable entries found for the selected source week.')
                return redirect('teacher_timetable')
            
            success_count = 0
            for week_offset in range(num_weeks):
                current_target = target_date + timedelta(weeks=week_offset)
                current_target_end = current_target + timedelta(days=6)
                
                for entry in source_timetables:
                    try:
                        cursor.execute("""
                            INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                                 start_time, end_time, room, week_start_date, week_end_date)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, [entry[1], entry[2], entry[3], entry[4], entry[5], entry[6], entry[7], current_target, current_target_end])
                        success_count += 1
                    except IntegrityError:
                        continue
            
            if success_count > 0:
                messages.success(request, f'Timetable copied successfully to {num_weeks} week(s) with {success_count} entries.')
            else:
                messages.warning(request, 'No entries were copied. Check for conflicts.')
    
    return redirect('teacher_timetable')


def teacher_timetable_bulk_delete(request):
    """
    Delete multiple timetable entries
    """
    if not request.session.get('teacher_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('teacher_login')
    
    teacher_id = request.session['teacher_id']
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.error(request, 'No entries selected.')
            return redirect('teacher_timetable')
        
        with connection.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(ids))
            cursor.execute(
                f"DELETE FROM timetable WHERE id IN ({placeholders}) AND teacher_id = %s", 
                ids + [teacher_id]
            )
        
        messages.success(request, f'{len(ids)} timetable entries deleted successfully.')
    
    return redirect('teacher_timetable')

# Student Timetable View
def student_timetable_view(request):
    if 'user_id' not in request.session:
        messages.error(request, 'Please log in to access the student portal.')
        return redirect('/login/')
    
    user_id = request.session['user_id']
    with connection.cursor() as cursor:
        # Fetch class and section for the student
        cursor.execute("SELECT class, section FROM student_page1 WHERE user_id = %s", [user_id])
        student = cursor.fetchone()
        if not student:
            messages.error(request, 'Student class information not found.')
            return redirect('student_timetable')
        
        class_name, section = student
        if not class_name:
            messages.error(request, 'Invalid class information for the student.')
            return redirect('student_timetable')
        
        # Construct class_id
        class_id = f"{class_name}{section}" if section else class_name
        
        # Check if timetable entries exist
        cursor.execute("SELECT COUNT(*) FROM timetable WHERE class_id = %s", [class_id])
        timetable_count = cursor.fetchone()[0]
        if timetable_count == 0:
            messages.warning(request, f'No timetable entries found for class {class_id}.')
            return render(request, 'users/student_timetable.html', {
                'timetable_data': [], 'class_id': class_id
            })
        
        # Fetch timetable with teacher details
        query = """
            SELECT t.id, t.class_id, t.subject, tch.name, t.day_of_week, 
                   t.start_time, t.end_time, t.room
            FROM timetable t
            JOIN teachers tch ON t.teacher_id = tch.id
            WHERE t.class_id = %s
            ORDER BY FIELD(t.day_of_week, 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday')
        """
        cursor.execute(query, [class_id])
        timetables = [
            {
                'id': row[0],
                'class_id': row[1],
                'subject': row[2],
                'name': row[3],
                'day_of_week': row[4],
                'start_time': row[5],
                'end_time': row[6],
                'room': row[7]
            } for row in cursor.fetchall()
        ]
    
    # Organize by day as a list of (day, entries)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    timetable_data = []
    for day in days:
        entries = [entry for entry in timetables if entry['day_of_week'] == day]
        timetable_data.append((day, entries))
    
    return render(request, 'users/student_timetable.html', {
        'timetable_data': timetable_data, 'class_id': class_id
    })




# users/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.conf import settings
from django.db import connection, transaction
import os
import uuid
from django.utils import timezone

def teacher_profile(request):
    if 'teacher_id' not in request.session:
        messages.error(request, "Please log in to view your profile.")
        return redirect('login')

    teacher_id = request.session['teacher_id']
    
    # Fetch profile picture
    profile_picture = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT profile_pic_url FROM profile_pics_teachers WHERE teacher_id = %s", [teacher_id])
            profile_picture_result = cursor.fetchone()
            if profile_picture_result:
                profile_picture = f"{settings.MEDIA_URL}{profile_picture_result[0]}"
                print(f"DEBUG: Found profile picture: {profile_picture}")
    except Exception as e:
        print("Error fetching profile picture:", e)

    if request.method == "POST":
        print(f"DEBUG: POST request received")
        print(f"DEBUG: FILES in request: {request.FILES}")
        print(f"DEBUG: POST data: {list(request.POST.keys())}")
        
        try:
            with transaction.atomic():
                # Handle ONLY profile picture upload
                if 'profile_pic' in request.FILES and request.FILES['profile_pic']:
                    print("DEBUG: Processing profile picture upload")
                    profile_pic_file = request.FILES['profile_pic']
                    print(f"DEBUG: File name: {profile_pic_file.name}, Size: {profile_pic_file.size}")
                    
                    # Validate file type
                    allowed_extensions = ['.png', '.jpg', '.jpeg']
                    file_ext = os.path.splitext(profile_pic_file.name)[1].lower()
                    if file_ext not in allowed_extensions:
                        messages.error(request, "Only PNG, JPG, or JPEG files are allowed.")
                        return redirect('teacher_profile')

                    # Validate file size (5MB limit)
                    if profile_pic_file.size > 5 * 1024 * 1024:
                        messages.error(request, "File size must be less than 5MB.")
                        return redirect('teacher_profile')

                    # Generate file path using UUID and teacher_id
                    filename = f"{uuid.uuid4().hex}_{teacher_id}{file_ext}"
                    
                    # Create pfpicsteacher directory in MEDIA_ROOT
                    pfpics_dir = os.path.join(settings.MEDIA_ROOT, 'pfpicsteacher')
                    os.makedirs(pfpics_dir, exist_ok=True)
                    
                    file_path = os.path.join(pfpics_dir, filename)
                    print(f"DEBUG: Saving file to: {file_path}")
                    
                    # Delete old profile picture if exists
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("SELECT profile_pic_url FROM profile_pics_teachers WHERE teacher_id = %s", [teacher_id])
                            old_pic = cursor.fetchone()
                            if old_pic:
                                old_file_path = os.path.join(settings.MEDIA_ROOT, old_pic[0])
                                if os.path.exists(old_file_path):
                                    os.remove(old_file_path)
                                    print(f"DEBUG: Deleted old file: {old_file_path}")
                    except Exception as e:
                        print(f"Error deleting old profile picture: {e}")

                    # Save new file
                    try:
                        with open(file_path, 'wb+') as destination:
                            for chunk in profile_pic_file.chunks():
                                destination.write(chunk)
                        print(f"DEBUG: File saved successfully to {file_path}")
                    except Exception as e:
                        print(f"ERROR: Failed to save file: {e}")
                        messages.error(request, "Failed to save file.")
                        return redirect('teacher_profile')

                    # Update database: delete existing and insert new record
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("DELETE FROM profile_pics_teachers WHERE teacher_id = %s", [teacher_id])
                            cursor.execute("""
                                INSERT INTO profile_pics_teachers (teacher_id, profile_pic_url, created_at)
                                VALUES (%s, %s, %s)
                            """, [teacher_id, f"pfpicsteacher/{filename}", timezone.now()])
                        print(f"DEBUG: Updated database with new path: pfpicsteacher/{filename}")
                    except Exception as e:
                        print(f"ERROR: Database operation failed: {e}")
                        messages.error(request, "Failed to update database.")
                        return redirect('teacher_profile')
                    
                    messages.success(request, "Profile picture uploaded successfully!")
                    return redirect('teacher_profile')
        except Exception as e:
            print(f"ERROR: Transaction failed: {e}")
            messages.error(request, "An error occurred during upload.")
            return redirect('teacher_profile')

    # Fetch comprehensive teacher details
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    t.id, t.name, t.email, t.subject, t.class_teacher_of, t.created_at,
                    tp.full_name, tp.gender, tp.date_of_birth, tp.blood_group, tp.nationality,
                    tp.mobile_number, tp.alternate_contact, tp.official_email,
                    tp.residential_address, tp.city_state_pin,
                    tp.emergency_contact_name, tp.emergency_contact_number,
                    tp.designation, tp.department, tp.subjects_taught, tp.classes_assigned,
                    tp.sections, tp.employee_type, tp.employment_status, tp.joining_date,
                    tp.qualification, tp.specialization, tp.board_university, tp.year_of_passing,
                    tp.bed_ctet_tet, tp.special_training, tp.workshops_attended,
                    tp.is_class_teacher, tp.house_club_incharge, 
                    tp.cocurricular_responsibilities, tp.exam_duties
                FROM teachers t
                LEFT JOIN teacher_profiles tp ON t.id = tp.teacher_id
                WHERE t.id = %s
            """, [teacher_id])
            
            teacher = cursor.fetchone()
            
            if not teacher:
                messages.error(request, "Teacher not found.")
                return redirect('login')

            # Convert tuple to comprehensive dictionary for template
            teacher_data = {
                'id': teacher[0],
                'name': teacher[1],
                'email': teacher[2],
                'subject': teacher[3],
                'class_teacher_of': teacher[4],
                'created_at': teacher[5],
                'full_name': teacher[6] if teacher[6] else teacher[1],
                'gender': teacher[7],
                'date_of_birth': teacher[8],
                'blood_group': teacher[9],
                'nationality': teacher[10] if teacher[10] else 'Indian',
                'mobile_number': teacher[11],
                'alternate_contact': teacher[12],
                'official_email': teacher[13] if teacher[13] else teacher[2],
                'residential_address': teacher[14],
                'city_state_pin': teacher[15],
                'emergency_contact_name': teacher[16],
                'emergency_contact_number': teacher[17],
                'designation': teacher[18],
                'department': teacher[19],
                'subjects_taught': teacher[20],
                'classes_assigned': teacher[21],
                'sections': teacher[22],
                'employee_type': teacher[23],
                'employment_status': teacher[24] if teacher[24] else 'Active',
                'joining_date': teacher[25],
                'qualification': teacher[26],
                'specialization': teacher[27],
                'board_university': teacher[28],
                'year_of_passing': teacher[29],
                'bed_ctet_tet': teacher[30],
                'special_training': teacher[31],
                'workshops_attended': teacher[32],
                'is_class_teacher': teacher[33] if teacher[33] else 'No',
                'house_club_incharge': teacher[34],
                'cocurricular_responsibilities': teacher[35],
                'exam_duties': teacher[36],
            }
            profile_pic_url = profile_picture if profile_picture else f"{settings.MEDIA_URL}pfpicsteacher/default.jpg"
    except Exception as e:
        print(f"Error fetching teacher details: {e}")
        messages.error(request, "Failed to load profile.")
        return redirect('login')

    context = {
        'teacher': teacher_data,
        'profile_pic_url': profile_pic_url,
    }
    return render(request, 'users/teacher_profile.html', context)





def parent_profile_view(request):
    if "user_id" not in request.session:
        return redirect("/parent_login/")  # Redirect to login if not authenticated

    user_id = request.session["user_id"]  # Get logged-in user's ID

    # Fetch profile picture
    profile_picture = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
            profile_picture_result = cursor.fetchone()
            if profile_picture_result:
                profile_picture = f"{settings.MEDIA_URL}{profile_picture_result[0]}"
                print(f"DEBUG: Found profile picture: {profile_picture}")
    except Exception as e:
        print("Error fetching profile picture:", e)

    if request.method == "POST":
        print(f"DEBUG: POST request received")
        print(f"DEBUG: FILES in request: {request.FILES}")
        
        # Handle ONLY profile picture upload
        if 'profile_picture' in request.FILES and request.FILES['profile_picture']:
            print("DEBUG: Processing profile picture upload")
            profile_picture_file = request.FILES['profile_picture']
            print(f"DEBUG: File name: {profile_picture_file.name}, Size: {profile_picture_file.size}")
            
            # Validate file type
            allowed_extensions = ['.png', '.jpg', '.jpeg']
            file_ext = os.path.splitext(profile_picture_file.name)[1].lower()
            if file_ext not in allowed_extensions:
                messages.error(request, "Only PNG, JPG, or JPEG files are allowed.")
                return redirect('parent_profile_view')

            # Validate file size (5MB limit)
            if profile_picture_file.size > 5 * 1024 * 1024:
                messages.error(request, "File size must be less than 5MB.")
                return redirect('parent_profile_view')

            # Generate file path using UUID and user_id
            filename = f"{uuid.uuid4().hex}_{user_id}{file_ext}"
            
            # Create pfpics directory in MEDIA_ROOT
            pfpics_dir = os.path.join(settings.MEDIA_ROOT, 'pfpics')
            os.makedirs(pfpics_dir, exist_ok=True)
            
            file_path = os.path.join(pfpics_dir, filename)
            print(f"DEBUG: Saving file to: {file_path}")
            
            try:
                with transaction.atomic():
                    # Delete old profile picture if exists
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
                            old_pic = cursor.fetchone()
                            if old_pic:
                                old_file_path = os.path.join(settings.MEDIA_ROOT, old_pic[0])
                                if os.path.exists(old_file_path):
                                    os.remove(old_file_path)
                                    print(f"DEBUG: Deleted old file: {old_file_path}")
                    except Exception as e:
                        print(f"Error deleting old profile picture: {e}")

                    # Save new file
                    try:
                        with open(file_path, 'wb+') as destination:
                            for chunk in profile_picture_file.chunks():
                                destination.write(chunk)
                        print(f"DEBUG: File saved successfully to {file_path}")
                    except Exception as e:
                        print(f"ERROR: Failed to save file: {e}")
                        messages.error(request, "Failed to save file.")
                        return redirect('parent_profile_view')

                    # Update or insert profile picture path in database
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("SELECT id FROM profile_pics WHERE user_id = %s", [user_id])
                            existing = cursor.fetchone()
                            
                            db_path = f"pfpics/{filename}"
                            if existing:
                                cursor.execute(
                                    "UPDATE profile_pics SET image_path = %s, uploaded_at = %s WHERE user_id = %s",
                                    [db_path, timezone.now(), user_id]
                                )
                                print(f"DEBUG: Updated existing record with path: {db_path}")
                            else:
                                cursor.execute(
                                    "INSERT INTO profile_pics (user_id, image_path, uploaded_at) VALUES (%s, %s, %s)",
                                    [user_id, db_path, timezone.now()]
                                )
                                print(f"DEBUG: Inserted new record with path: {db_path}")
                    except Exception as e:
                        print(f"ERROR: Database operation failed: {e}")
                        messages.error(request, "Failed to update database.")
                        return redirect('parent_profile_view')
                    
                    messages.success(request, "Profile picture uploaded successfully!")
                    return redirect('parent_profile_view')
            except Exception as e:
                print(f"ERROR: Exception in POST processing: {e}")
                messages.error(request, f"Failed to process request: {str(e)}")
                return redirect('parent_profile_view')
        else:
            # No file uploaded, redirect without error
            return redirect('parent_profile_view')

    # Fetch ALL student details from all tables
    student = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    s1.name, s1.admission_number, s1.class, s1.section, s1.roll_number, s1.emis,
                    s2.tamil_name, s2.gender, s2.community, s2.dob, s2.nationality, 
                    s2.blood_group, s2.mother_tongue, s2.caste, s2.religion, 
                    s2.place_of_birth, s2.aadhaar, s2.disability, s2.id_mark1, 
                    s2.id_mark2, s2.current_class, s2.admission_class, s2.admission_year, 
                    s2.admission_date,
                    s3.email, s3.address, s3.contact, s3.alt_contact, s3.country, 
                    s3.state, s3.city, s3.pincode, s3.status, s3.house, 
                    s3.teacher_ward, s3.rte, s3.sports_quota, s3.prev_school, 
                    s3.prev_board,
                    s4.father_name, s4.father_name_tamil, s4.mother_name, s4.mother_name_tamil,
                    s4.father_contact, s4.mother_contact, s4.father_email, s4.mother_email,
                    s4.father_qualification, s4.mother_qualification, s4.father_occupation,
                    s4.mother_occupation, s4.father_income, s4.mother_income, s4.guardian_name,
                    s4.guardian_contact, s4.guardian_email, s4.child_living, s4.rights_on_child,
                    s4.med_blood_group, s4.diseases, s4.allergies, s4.medicines, 
                    s4.hospital, s4.doctor
                FROM student_page1 s1
                LEFT JOIN student_page2 s2 ON s1.user_id = s2.user_id
                LEFT JOIN student_page3 s3 ON s1.user_id = s3.user_id
                LEFT JOIN student_page4 s4 ON s1.user_id = s4.user_id
                WHERE s1.user_id = %s
            """, [user_id])
            student_data = cursor.fetchone()
            
            if student_data:
                student = {
                    # From student_page1
                    'name': student_data[0],
                    'admission_number': student_data[1],
                    'class': student_data[2],
                    'section': student_data[3],
                    'roll_number': student_data[4],
                    'emis': student_data[5],
                    
                    # From student_page2
                    'tamil_name': student_data[6],
                    'gender': student_data[7],
                    'community': student_data[8],
                    'dob': student_data[9],
                    'nationality': student_data[10],
                    'blood_group': student_data[11],
                    'mother_tongue': student_data[12],
                    'caste': student_data[13],
                    'religion': student_data[14],
                    'place_of_birth': student_data[15],
                    'aadhaar': student_data[16],
                    'disability': student_data[17],
                    'id_mark1': student_data[18],
                    'id_mark2': student_data[19],
                    'current_class': student_data[20],
                    'admission_class': student_data[21],
                    'admission_year': student_data[22],
                    'admission_date': student_data[23],
                    
                    # From student_page3
                    'email': student_data[24],
                    'address': student_data[25],
                    'contact': student_data[26],
                    'alt_contact': student_data[27],
                    'country': student_data[28],
                    'state': student_data[29],
                    'city': student_data[30],
                    'pincode': student_data[31],
                    'status': student_data[32],
                    'house': student_data[33],
                    'teacher_ward': student_data[34],
                    'rte': student_data[35],
                    'sports_quota': student_data[36],
                    'prev_school': student_data[37],
                    'prev_board': student_data[38],
                    
                    # From student_page4
                    'father_name': student_data[39],
                    'father_name_tamil': student_data[40],
                    'mother_name': student_data[41],
                    'mother_name_tamil': student_data[42],
                    'father_contact': student_data[43],
                    'mother_contact': student_data[44],
                    'father_email': student_data[45],
                    'mother_email': student_data[46],
                    'father_qualification': student_data[47],
                    'mother_qualification': student_data[48],
                    'father_occupation': student_data[49],
                    'mother_occupation': student_data[50],
                    'father_income': student_data[51],
                    'mother_income': student_data[52],
                    'guardian_name': student_data[53],
                    'guardian_contact': student_data[54],
                    'guardian_email': student_data[55],
                    'child_living': student_data[56],
                    'rights_on_child': student_data[57],
                    'med_blood_group': student_data[58],
                    'diseases': student_data[59],
                    'allergies': student_data[60],
                    'medicines': student_data[61],
                    'hospital': student_data[62],
                    'doctor': student_data[63],
                }
                
                print(f"DEBUG: Fetched complete student data for user_id: {user_id}")
            else:
                print(f"DEBUG: No student data found for user_id: {user_id}")
                
    except Exception as e:
        print(f"DEBUG: Error fetching student data: {e}")
        messages.error(request, f"Error loading student information: {str(e)}")

    print(f"DEBUG: Rendering template with profile_picture: {profile_picture}")
    return render(request, "users/parent_profile_view.html", {
        "student": student,
        "profile_picture": profile_picture,
        "user_id": user_id
    })




def parent_student_portal(request):
    if "user_id" not in request.session:
        return redirect("/")
    
    user_id = request.session['user_id']
    selected_date = request.GET.get('date', '')
    
    with connection.cursor() as cursor:
        if selected_date:
            cursor.execute(
                """
                SELECT a.date, s.admission_number, s.name, a.class, a.section, a.status 
                FROM attendance a
                JOIN student_page1 s ON a.student_id = s.user_id
                WHERE s.user_id = %s AND a.date = %s
                ORDER BY a.date DESC
                """,
                [user_id, selected_date]
            )
        else:
            cursor.execute(
                """
                SELECT a.date, s.admission_number, s.name, a.class, a.section, a.status
                FROM attendance a
                JOIN student_page1 s ON a.student_id = s.user_id
                WHERE s.user_id = %s
                ORDER BY a.date DESC
                """,
                [user_id]
            )
        attendance_records = [
            {
                'date': row[0],
                'admission_number': row[1],
                'name': row[2],
                'class': row[3],
                'section': row[4] if row[4] else 'N/A',
                'status': row[5]
            } for row in cursor.fetchall()
        ]

    return render(request, 'users/parent_student_portal.html', {
        'attendance_records': attendance_records,
        'selected_date': selected_date
    })



def parent_student_leave(request):
    """Handle parent student leave request submission and viewing with auto-filled student data."""
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access the parent student portal.")
        return redirect("/parent_login/")

    user_id = request.session["user_id"]
    
    # Fetch student data from database to pre-fill the form
    student_data = {}
    with connection.cursor() as cursor:
        try:
            # Fetch student details from student_page1
            cursor.execute("""
                SELECT name, admission_number, class, section
                FROM student_page1
                WHERE user_id = %s
                LIMIT 1
            """, [user_id])
            result = cursor.fetchone()
            if result:
                student_data = {
                    'name': result[0],
                    'admission_number': result[1],
                    'class': result[2],
                    'section': result[3]
                }
        except Exception as e:
            messages.error(request, f"Error fetching student data: {str(e)}")
    
    if request.method == "POST":
        try:
            # Get data from readonly fields (already validated from database)
            form_data = {
                "student_name": student_data.get('name', ''),
                "reg_number": student_data.get('admission_number', ''),
                "class_number": student_data.get('class', ''),
                "section": student_data.get('section', ''),
                "leave_reason": request.POST.get("leave_reason", "").strip(),
                "leave_start_date": request.POST.get("leave_start_date", ""),
                "leave_end_date": request.POST.get("leave_end_date", ""),
                "leave_duration": request.POST.get("leave_duration", ""),
                "half_day_type": request.POST.get("half_day_type", "")
            }

            # Validate required fields
            required_fields = ["student_name", "reg_number", "class_number", "section", 
                             "leave_reason", "leave_start_date", "leave_end_date", "leave_duration"]
            missing_fields = [field for field in required_fields if not form_data[field]]
            if missing_fields:
                messages.error(request, f"Missing required fields: {', '.join(missing_fields)}")
                return redirect("parent_student_leave")
            
            # Validate leave duration
            if form_data["leave_duration"] not in ["full", "half"]:
                messages.error(request, "Invalid leave duration.")
                return redirect("parent_student_leave")
                
            # Validate half day type if half day is selected
            if form_data["leave_duration"] == "half" and not form_data["half_day_type"]:
                messages.error(request, "Please select half day type for half-day leave.")
                return redirect("parent_student_leave")

            # Validate dates
            try:
                start_date = datetime.strptime(form_data["leave_start_date"], "%Y-%m-%d")
                end_date = datetime.strptime(form_data["leave_end_date"], "%Y-%m-%d")
                if start_date > end_date:
                    messages.error(request, "End date must be on or after start date.")
                    return redirect("parent_student_leave")
            except ValueError:
                messages.error(request, "Invalid date format.")
                return redirect("parent_student_leave")

            # Insert leave request into database
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO student_leave_requests 
                    (user_id, student_name, reg_number, class_number, section, leave_reason,
                    leave_start_date, leave_end_date, leave_duration, half_day_type, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [user_id, form_data["student_name"], form_data["reg_number"], 
                      form_data["class_number"], form_data["section"],
                      form_data["leave_reason"], 
                      form_data["leave_start_date"], form_data["leave_end_date"],
                      form_data["leave_duration"], 
                      form_data["half_day_type"] if form_data["leave_duration"] == "half" else None,
                      "Pending"])
                connection.commit()
            messages.success(request, "Leave request submitted successfully.")
        except Exception as e:
            connection.rollback()
            messages.error(request, f"Error submitting leave request: {str(e)}")
        return redirect("parent_student_leave")

    # Fetch leave requests for this parent student
    leave_requests = []
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                SELECT id, student_name, reg_number, class_number, section, leave_reason, 
                leave_start_date, leave_end_date, leave_duration, half_day_type, status
                FROM student_leave_requests WHERE user_id = %s
                ORDER BY leave_start_date DESC
            """, [user_id])
            leave_requests = cursor.fetchall()
        except Exception as e:
            messages.error(request, f"Error fetching leave requests: {str(e)}")

    return render(request, "users/parent_student_leave.html", {
        "leave_requests": leave_requests,
        "student_data": student_data
    })


from django.conf import settings
import os

def parent_student_circular(request):
    CIRCULARS_DIR = os.path.join(settings.MEDIA_ROOT, 'circulars')
    # Get the logged-in user's user_id from session and fetch class/section from student_page1
    student_class = None
    student_section = None
    error_message = None

    if "user_id" not in request.session:
        error_message = "Please log in to view circulars."
        print("No user_id found in session")
    else:
        user_id = request.session['user_id']
        print(f"Session user_id: {user_id}")
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT class, section FROM student_page1 WHERE user_id = %s",
                    [user_id]
                )
                result = cursor.fetchone()
                if result:
                    student_class, student_section = [normalize_value(r) for r in result]
                    print(f"Parent student user_id={user_id}: class={student_class}, section={student_section}")
                else:
                    error_message = "No class or section found for your account. Please contact the admin."
                    print(f"No student record found for user_id: {user_id}")
        except Exception as e:
            error_message = "Error fetching your class/section. Please try again later."
            print(f"Error fetching parent student class/section for user_id={user_id}: {e}")

    # Get filter type from POST request (default to 'all')
    filter_type = request.POST.get('filter_type', 'all')
    print(f"Filter type: {filter_type}")

    circulars = []
    for file in os.listdir(CIRCULARS_DIR):
        if file.endswith(('.jpg', '.png', '.jpeg', '.webp', '.gif')):
            full_path = os.path.join(CIRCULARS_DIR, file)
            if not os.path.exists(full_path):
                print(f"Image file missing in parent student view: {full_path}")
                continue

            title_file = f"{file}.txt"
            title_path = os.path.join(CIRCULARS_DIR, title_file)
            title = "Untitled"
            target = "all"
            class_name = ""
            section = ""

            if os.path.exists(title_path):
                try:
                    with open(title_path, 'r') as f:
                        lines = f.readlines()
                        title = lines[0].strip() if lines else "Untitled"
                        target = lines[1].strip().lower() if len(lines) > 1 else "all"
                        if target == 'specific' and len(lines) >= 4:
                            class_name = normalize_value(lines[2])
                            section = normalize_value(lines[3])
                        print(f"Circular {file}: title={title}, target={target}, class={class_name}, section={section}")
                except Exception as e:
                    print(f"Error reading metadata from {title_path}: {e}")
                    continue

            # Filter circulars based on filter_type
            include_circular = False
            if filter_type == 'all':
                if target == "all" or (
                    target == "specific" and
                    student_class and student_section and
                    class_name == student_class and section == student_section
                ):
                    include_circular = True
            elif filter_type == 'specific':
                if target == "specific" and student_class and student_section and class_name == student_class and section == student_section:
                    include_circular = True

            if include_circular:
                try:
                    created_at = datetime.fromtimestamp(os.path.getctime(full_path)).strftime('%Y-%m-%d %H:%M:%S')
                    image_url = f"/media/circulars/{file}"  # Consistent path
                    print(f"Included circular: {file}, image_url: {image_url}, full_path: {full_path}")
                    display_target = "All" if target == "all" else f"Class: {class_name.capitalize()}, Section: {section.capitalize()}"
                    circulars.append({
                        'title': title,
                        'image_url': image_url,
                        'date': created_at,
                        'target': display_target
                    })
                except Exception as e:
                    print(f"Error processing file {file}: {e}")

    # Sort by newest first
    circulars = sorted(circulars, key=lambda x: x['date'], reverse=True)
    print(f"Total circulars displayed: {len(circulars)}")

    return render(request, 'users/parent_student_circular.html', {
        'circulars': circulars,
        'student_class': student_class,
        'student_section': student_section,
        'filter_type': filter_type,
        'error_message': error_message
    })

def parent_study_materials(request):
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access the parent student portal.")
        return redirect("/parent_login/")

    user_id = request.session["user_id"]
    student_class = None
    student_section = None

    # Fetch student's class and section
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT class, section FROM student_page1 WHERE user_id = %s",
                [user_id]
            )
            result = cursor.fetchone()
            if result:
                student_class, student_section = result
            else:
                messages.error(request, "No class or section found for your account. Please contact the admin.")
                return redirect("/parent-study-materials/")
    except Exception as e:
        messages.error(request, f"Error fetching class/section: {str(e)}")
        return redirect("/parent-study-materials/")

    try:
        with connection.cursor() as cursor:
            query = """
                SELECT title, file_path, upload_date, class, section
                FROM study_materials
                WHERE class = %s AND section = %s
                ORDER BY upload_date DESC
            """
            cursor.execute(query, [student_class, student_section])
            study_materials = [
                {
                    "title": r[0],
                    "file_path": r[1],
                    "upload_date": r[2],
                    "class": r[3],
                    "section": r[4]
                } for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving study materials: {str(e)}")
        study_materials = []

    return render(request, "users/parent_study_materials.html", {
        "study_materials": study_materials,
        "media_url": settings.MEDIA_URL,
        "student_class": student_class,
        "student_section": student_section
    })



def parent_homework(request):
    if "user_id" not in request.session:
        messages.error(request, "Please log in to access the parent student portal.")
        return redirect("/parent_login/")

    user_id = request.session["user_id"]
    student_class = None
    student_section = None

    # Fetch student's class and section
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT class, section FROM student_page1 WHERE user_id = %s",
                [user_id]
            )
            result = cursor.fetchone()
            if result:
                student_class, student_section = result
            else:
                messages.error(request, "No class or section found for your account. Please contact the admin.")
                return redirect("/parent-homework/")
    except Exception as e:
        messages.error(request, f"Error fetching class/section: {str(e)}")
        return redirect("/parent-homework/")

    if request.method == "POST":
        title = request.POST.get("title")
        submission_date = request.POST.get("submission_date")
        uploaded_file = request.FILES.get("file")

        if not all([title, submission_date, uploaded_file]):
            messages.error(request, "All fields are required.")
            return redirect("/parent-homework/")

        # Validate file is a PDF
        validator = FileExtensionValidator(allowed_extensions=['pdf'])
        try:
            validator(uploaded_file)
        except ValidationError:
            messages.error(request, "Only PDF files are allowed.")
            return redirect("/parent-homework/")

        try:
            # Save file with unique filename
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'Uploads'))
            filename = f"{uuid.uuid4().hex}_{uploaded_file.name}"
            filename = fs.save(filename, uploaded_file)
            file_path = f"Uploads/{filename}"

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO homework (user_id, title, submission_date, file_path, class, section)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    [user_id, title, submission_date, file_path, student_class, student_section]
                )
            messages.success(request, "Homework submitted successfully!")
        except Exception as e:
            messages.error(request, f"Error submitting homework: {str(e)}")
        return redirect("/parent-homework/")

    # Fetch student's submitted homework
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT h.title, h.submission_date, h.file_path
                FROM homework h
                WHERE h.user_id = %s
                ORDER BY h.submission_date DESC
                """,
                [user_id]
            )
            homework_list = [
                {"title": r[0], "submission_date": r[1], "file_path": r[2]}
                for r in cursor.fetchall()
            ]
    except Exception as e:
        messages.error(request, f"Error retrieving homework: {str(e)}")
        homework_list = []

    # Fetch teacher-uploaded homework for this student's class/section
    teacher_homework = []
    HOMEWORK_DIR = os.path.join(settings.MEDIA_ROOT, 'teacher_homework')
    
    if os.path.exists(HOMEWORK_DIR) and student_class and student_section:
        # Normalize student's class and section for comparison
        normalized_student_class = normalize_value(student_class)
        normalized_student_section = normalize_value(student_section)
        
        for file in os.listdir(HOMEWORK_DIR):
            # Skip metadata files
            if file.endswith('.txt'):
                continue
            
            # Check for valid file extensions
            if not any(file.lower().endswith(ext) for ext in ['.pdf', '.docx', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.gif', '.webp']):
                continue
            
            full_path = os.path.join(HOMEWORK_DIR, file)
            if not os.path.exists(full_path):
                continue

            metadata_file = f"{file}.txt"
            metadata_path = os.path.join(HOMEWORK_DIR, metadata_file)
            
            # Read metadata
            if os.path.exists(metadata_path):
                try:
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                        
                        if len(lines) < 6:
                            continue
                        
                        title = lines[0].strip() if len(lines) > 0 else "Untitled"
                        description = lines[1].strip() if len(lines) > 1 else ""
                        subject = lines[2].strip() if len(lines) > 2 else ""
                        due_date = lines[3].strip() if len(lines) > 3 else ""
                        target_type = lines[4].strip() if len(lines) > 4 else "all"
                        teacher_id = lines[5].strip() if len(lines) > 5 else ""
                        
                        # Check if homework is for this student
                        homework_for_student = False
                        
                        if target_type == 'all':
                            # Homework is for all classes
                            homework_for_student = True
                        elif target_type == 'specific' and len(lines) >= 8:
                            # Check if homework matches student's class and section
                            hw_class = normalize_value(lines[6].strip())
                            hw_section = normalize_value(lines[7].strip())
                            
                            if hw_class == normalized_student_class and hw_section == normalized_student_section:
                                homework_for_student = True
                        
                        # If homework is for this student, add to list
                        if homework_for_student:
                            file_ext = os.path.splitext(file)[1].lower()
                            upload_date = datetime.fromtimestamp(os.path.getctime(full_path)).strftime('%B %d, %Y at %I:%M %p')
                            
                            # Format due date if exists
                            formatted_due_date = None
                            if due_date:
                                try:
                                    formatted_due_date = datetime.strptime(due_date, '%Y-%m-%d').strftime('%B %d, %Y')
                                except:
                                    formatted_due_date = due_date
                            
                            teacher_homework.append({
                                'title': title,
                                'description': description if description else None,
                                'subject': subject if subject else None,
                                'due_date': formatted_due_date,
                                'file_path': file,
                                'file_type': file_ext,
                                'date': upload_date,
                            })
                except Exception as e:
                    print(f"Error reading metadata from {metadata_path}: {e}")
                    continue
    
    # Sort teacher homework by newest first
    teacher_homework = sorted(teacher_homework, key=lambda x: x['date'], reverse=True)

    return render(request, "users/parent_homework.html", {
        "homework_list": homework_list,
        "teacher_homework": teacher_homework,
        "student_class": student_class,
        "student_section": student_section
    })



def parent_student_timetable(request):
    if 'user_id' not in request.session:
        messages.error(request, 'Please log in to access the parent student portal.')
        return redirect('/parent_login/')
    
    user_id = request.session['user_id']
    with connection.cursor() as cursor:
        # Fetch class and section for the student
        cursor.execute("SELECT class, section FROM student_page1 WHERE user_id = %s", [user_id])
        student = cursor.fetchone()
        if not student:
            messages.error(request, 'Student class information not found.')
            return redirect('parent_student_timetable')
        
        class_name, section = student
        if not class_name:
            messages.error(request, 'Invalid class information for the student.')
            return redirect('parent_student_timetable')
        
        # Construct class_id
        class_id = f"{class_name}{section}" if section else class_name
        
        # Check if timetable entries exist
        cursor.execute("SELECT COUNT(*) FROM timetable WHERE class_id = %s", [class_id])
        timetable_count = cursor.fetchone()[0]
        if timetable_count == 0:
            messages.warning(request, f'No timetable entries found for class {class_id}.')
            return render(request, 'users/parent_student_timetable.html', {
                'timetable_data': [], 'class_id': class_id
            })
        
        # Fetch timetable with teacher details
        query = """
            SELECT t.id, t.class_id, t.subject, tch.name, t.day_of_week, 
                   t.start_time, t.end_time, t.room
            FROM timetable t
            JOIN teachers tch ON t.teacher_id = tch.id
            WHERE t.class_id = %s
            ORDER BY FIELD(t.day_of_week, 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday')
        """
        cursor.execute(query, [class_id])
        timetables = [
            {
                'id': row[0],
                'class_id': row[1],
                'subject': row[2],
                'name': row[3],
                'day_of_week': row[4],
                'start_time': row[5],
                'end_time': row[6],
                'room': row[7]
            } for row in cursor.fetchall()
        ]
    
    # Organize by day as a list of (day, entries)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    timetable_data = []
    for day in days:
        entries = [entry for entry in timetables if entry['day_of_week'] == day]
        timetable_data.append((day, entries))
    
    return render(request, 'users/parent_student_timetable.html', {
        'timetable_data': timetable_data, 'class_id': class_id
    })


def student_progress_card(request):
    if 'user_id' not in request.session:
        messages.error(request, 'Please log in to access the student portal.')
        return redirect('/login/')
    
    user_id = request.session['user_id']
    with connection.cursor() as cursor:
        # Fetch class and section for the student
        cursor.execute("SELECT class, section FROM student_page1 WHERE user_id = %s", [user_id])
        student = cursor.fetchone()
        if not student:
            messages.error(request, 'Student class information not found.')
            return redirect('student_progress_card')
        
        class_name, section = student
        if not class_name:
            messages.error(request, 'Invalid class information for the student.')
            return redirect('student_progress_card')
        
        # Placeholder for progress card data (since results are not yet available)
        progress_card_data = []

    return render(request, 'users/student_progress_card.html', {
        'progress_card_data': progress_card_data,
        'class_name': class_name,
        'section': section
    })


from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection

def parent_student_progress_card(request):
    if 'user_id' not in request.session:
        messages.error(request, 'Please log in to access the parent student portal.')
        return redirect('/login/')
    
    user_id = request.session['user_id']
    
    with connection.cursor() as cursor:
        # === Fetch Student Details (with admission_number fallback) ===
        try:
            cursor.execute("""
                SELECT 
                    s.name, 
                    s.roll_number, 
                    s.class, 
                    s.section, 
                    COALESCE(s.admission_number, ''), 
                    p.image_path
                FROM student_page1 s 
                LEFT JOIN profile_pics p ON s.user_id = p.user_id
                WHERE s.user_id = %s
            """, [user_id])
            student_row = cursor.fetchone()
        except Exception as e:
            if "admission_number" in str(e):
                cursor.execute("""
                    SELECT 
                        s.name, 
                        s.roll_number, 
                        s.class, 
                        s.section, 
                        p.image_path
                    FROM student_page1 s 
                    LEFT JOIN profile_pics p ON s.user_id = p.user_id
                    WHERE s.user_id = %s
                """, [user_id])
                student_row = cursor.fetchone()
                admission_number = ''
            else:
                raise e
        
        if not student_row:
            messages.error(request, 'Student information not found.')
            return redirect('parent_dashboard')
        
        if len(student_row) == 6:
            name, roll_number, class_name, section, admission_number, image_path = student_row
        else:
            name, roll_number, class_name, section, image_path = student_row
            admission_number = ''
        
        if not class_name or not section:
            messages.error(request, 'Invalid class or section information.')
            return redirect('parent_dashboard')
        
        # === Fetch ALL Subjects and Marks (LEFT JOIN so all subjects appear even if no marks) ===
        cursor.execute("""
            SELECT 
                ss.name AS subject_name,
                COALESCE(m.marks, 0) AS marks,
                ss.max_marks,
                COALESCE(m.grade, 'E') AS grade
            FROM school_subjects ss
            LEFT JOIN school_marks m 
                ON m.subject_id = ss.id 
                AND m.student_id = %s
            WHERE ss.class = %s
            ORDER BY ss.name
        """, [user_id, class_name])
        
        marks_rows = cursor.fetchall()
        marks = [
            {
                'subject': row[0],
                'marks': row[1],
                'max_marks': row[2],
                'grade': row[3]
            }
            for row in marks_rows
        ]
        
        # === Fetch Class Teacher's Signature ===
        cursor.execute("""
            SELECT ts.signature
            FROM teachers t
            JOIN teacher_signature ts ON ts.teacher_id = t.id
            WHERE t.class_teacher_of = %s
            LIMIT 1
        """, [f"{class_name}-{section}"])
        
        teacher_sig_row = cursor.fetchone()
        teacher_signature = teacher_sig_row[0] if teacher_sig_row else None
        
        # === Fetch Principal Signature (global) ===
        cursor.execute("SELECT signature FROM principal_signature LIMIT 1")
        principal_sig_row = cursor.fetchone()
        principal_signature = principal_sig_row[0] if principal_sig_row else None
        
        # === Calculate Performance Summary ===
        has_marks = any(m['marks'] > 0 for m in marks)  # True if at least one subject has marks
        total_marks = sum(m['marks'] for m in marks)
        total_max_marks = sum(m['max_marks'] for m in marks)
        
        if total_max_marks > 0:
            percentage = round((total_marks / total_max_marks) * 100, 2)
        else:
            percentage = 0.0
        
        # Overall Grade
        if percentage >= 80:
            overall_grade = 'A'
        elif percentage >= 60:
            overall_grade = 'B'
        elif percentage >= 40:
            overall_grade = 'C'
        elif percentage >= 33:
            overall_grade = 'D'
        else:
            overall_grade = 'E'
        
        # Pass/Fail Status (must pass minimum 33% in ALL subjects)
        passed = all(m['marks'] >= (0.33 * m['max_marks']) for m in marks) if marks else False
        status = 'Pass' if passed else 'Fail'
        
        if not marks:
            overall_grade = 'N/A'
            status = 'Pending'
        
        # === Final Context ===
        context = {
            'student': {
                'name': name,
                'roll_number': roll_number,
                'class': class_name,
                'section': section,
                'admission_number': admission_number,
                'image_path': image_path,  # Will be used as /media/{{ image_path }} in template
            },
            'marks': marks,
            'total_marks': total_marks,
            'total_max_marks': total_max_marks,
            'percentage': percentage,
            'overall_grade': overall_grade,
            'status': status,
            'teacher_signature': teacher_signature,
            'principal_signature': principal_signature,
            'has_marks': has_marks,
            'class_name': class_name,
            'section': section or 'N/A',
        }
    
    return render(request, 'users/parent_student_progress_card.html', context)


    


from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.db import connection
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import json
import os
from io import BytesIO

def render_qr_scan(request):
    if "admin_id" not in request.session:
        return redirect("/admin_login/")
    return render(request, 'users/qr_scan.html')

def scan_qr_code(request):
    if "admin_id" not in request.session:
        return JsonResponse({"error": "Admin not authenticated."}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            qr_url = data.get('url', '')
            
            # Extract admission number from URL
            admission_number = qr_url.split('/')[-2] if qr_url.endswith('/') else qr_url.split('/')[-1]
            
            # Remove any query parameters or fragments
            admission_number = admission_number.split('?')[0].split('#')[0]

            with connection.cursor() as cursor:
                # Fetch student data with better error handling
                cursor.execute("""
                    SELECT 
                        sp1.user_id, sp1.name, sp1.admission_number, sp1.class, sp1.section, 
                        sp1.roll_number, sp1.emis, 
                        sp2.gender, sp2.community, sp2.tamil_name, sp2.dob, 
                        sp2.nationality, sp2.blood_group, sp2.mother_tongue, 
                        sp2.caste, sp2.religion, sp2.place_of_birth, sp2.aadhaar, 
                        sp2.disability, sp2.id_mark1, sp2.id_mark2, sp2.current_class, 
                        sp2.admission_class, sp2.admission_year, sp2.admission_date,
                        sp3.email, sp3.address, sp3.contact, sp3.alt_contact, 
                        sp3.country, sp3.state, sp3.city, sp3.pincode, sp3.status, 
                        sp3.house, sp3.teacher_ward, sp3.rte, sp3.sports_quota, 
                        sp3.prev_school, sp3.prev_board,
                        sp4.father_name, sp4.father_name_tamil, sp4.mother_name, 
                        sp4.mother_name_tamil, sp4.father_contact, sp4.mother_contact, 
                        sp4.father_email, sp4.mother_email, sp4.father_qualification, 
                        sp4.mother_qualification, sp4.father_occupation, 
                        sp4.mother_occupation, sp4.father_income, sp4.mother_income, 
                        sp4.guardian_name, sp4.guardian_contact, sp4.guardian_email, 
                        sp4.child_living, sp4.rights_on_child, sp4.med_blood_group, 
                        sp4.diseases, sp4.allergies, sp4.medicines, sp4.hospital, 
                        sp4.doctor
                    FROM 
                        student_page1 sp1
                    LEFT JOIN 
                        student_page2 sp2 ON sp1.user_id = sp2.user_id
                    LEFT JOIN 
                        student_page3 sp3 ON sp1.user_id = sp3.user_id
                    LEFT JOIN 
                        student_page4 sp4 ON sp1.user_id = sp4.user_id
                    WHERE 
                        sp1.admission_number = %s
                """, [admission_number])
                
                student_data = cursor.fetchone()

                if not student_data:
                    return JsonResponse({"error": f"Student not found with admission number: {admission_number}"}, status=404)

                user_id = student_data[0]
                profile_picture = None
                
                # Fetch profile picture with better error handling
                try:
                    cursor.execute("SELECT image_path FROM profile_pics WHERE user_id = %s", [user_id])
                    profile_picture_result = cursor.fetchone()
                    if profile_picture_result and profile_picture_result[0]:
                        # For HTML display, use MEDIA_URL; for PDF, use full path
                        profile_picture_web = f"{settings.MEDIA_URL}{profile_picture_result[0]}"
                        profile_picture_path = os.path.join(settings.MEDIA_ROOT, profile_picture_result[0])
                    else:
                        profile_picture_web = None
                        profile_picture_path = None
                except Exception as e:
                    print(f"Error fetching profile picture: {str(e)}")
                    profile_picture_web = None
                    profile_picture_path = None

                # Helper function to safely convert values
                def safe_str(value):
                    if value is None:
                        return "N/A"
                    return str(value)

                response_data = {
                    "user_id": user_id,
                    "profile_picture": profile_picture_web,  # For HTML display
                    "profile_picture_path": profile_picture_path,  # For PDF generation
                    "name": safe_str(student_data[1]),
                    "admission_number": safe_str(student_data[2]),
                    "class": safe_str(student_data[3]),
                    "section": safe_str(student_data[4]),
                    "roll_number": safe_str(student_data[5]),
                    "emis": safe_str(student_data[6]),
                    "gender": safe_str(student_data[7]),
                    "community": safe_str(student_data[8]),
                    "tamil_name": safe_str(student_data[9]),
                    "dob": safe_str(student_data[10]) if student_data[10] else "N/A",
                    "nationality": safe_str(student_data[11]),
                    "blood_group": safe_str(student_data[12]),
                    "mother_tongue": safe_str(student_data[13]),
                    "caste": safe_str(student_data[14]),
                    "religion": safe_str(student_data[15]),
                    "place_of_birth": safe_str(student_data[16]),
                    "aadhaar": safe_str(student_data[17]),
                    "disability": safe_str(student_data[18]),
                    "id_mark1": safe_str(student_data[19]),
                    "id_mark2": safe_str(student_data[20]),
                    "current_class": safe_str(student_data[21]),
                    "admission_class": safe_str(student_data[22]),
                    "admission_year": safe_str(student_data[23]),
                    "admission_date": safe_str(student_data[24]) if student_data[24] else "N/A",
                    "email": safe_str(student_data[25]),
                    "address": safe_str(student_data[26]),
                    "contact": safe_str(student_data[27]),
                    "alt_contact": safe_str(student_data[28]),
                    "country": safe_str(student_data[29]),
                    "state": safe_str(student_data[30]),
                    "city": safe_str(student_data[31]),
                    "pincode": safe_str(student_data[32]),
                    "status": safe_str(student_data[33]),
                    "house": safe_str(student_data[34]),
                    "teacher_ward": safe_str(student_data[35]),
                    "rte": safe_str(student_data[36]),
                    "sports_quota": safe_str(student_data[37]),
                    "prev_school": safe_str(student_data[38]),
                    "prev_board": safe_str(student_data[39]),
                    "father_name": safe_str(student_data[40]),
                    "father_name_tamil": safe_str(student_data[41]),
                    "mother_name": safe_str(student_data[42]),
                    "mother_name_tamil": safe_str(student_data[43]),
                    "father_contact": safe_str(student_data[44]),
                    "mother_contact": safe_str(student_data[45]),
                    "father_email": safe_str(student_data[46]),
                    "mother_email": safe_str(student_data[47]),
                    "father_qualification": safe_str(student_data[48]),
                    "mother_qualification": safe_str(student_data[49]),
                    "father_occupation": safe_str(student_data[50]),
                    "mother_occupation": safe_str(student_data[51]),
                    "father_income": safe_str(student_data[52]),
                    "mother_income": safe_str(student_data[53]),
                    "guardian_name": safe_str(student_data[54]),
                    "guardian_contact": safe_str(student_data[55]),
                    "guardian_email": safe_str(student_data[56]),
                    "child_living": safe_str(student_data[57]),
                    "rights_on_child": safe_str(student_data[58]),
                    "med_blood_group": safe_str(student_data[59]),
                    "diseases": safe_str(student_data[60]),
                    "allergies": safe_str(student_data[61]),
                    "medicines": safe_str(student_data[62]),
                    "hospital": safe_str(student_data[63]),
                    "doctor": safe_str(student_data[64])
                }

                # Store student data in session for PDF generation
                request.session['student_data'] = response_data

                return JsonResponse(response_data)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error processing request: {str(e)}"}, status=500)
    
    return JsonResponse({"error": "POST request required with QR code URL."}, status=400)

def download_student_pdf(request):
    if "admin_id" not in request.session:
        return JsonResponse({"error": "Admin not authenticated."}, status=401)
    
    if 'student_data' not in request.session:
        return JsonResponse({"error": "No student data found. Please scan QR code first."}, status=400)
    
    try:
        student_data = request.session['student_data']
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#0052cc')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#0052cc'),
            backColor=colors.HexColor('#f0f8ff'),
            borderPadding=8
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            textColor=colors.black
        )
        
        # Add title
        title = Paragraph("Student Information Report", title_style)
        elements.append(title)
        
        subtitle = Paragraph("Manavargal School Management System", normal_style)
        elements.append(subtitle)
        
        from datetime import datetime
        date_para = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y')}", normal_style)
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Add profile picture if available
        profile_picture = student_data.get('profile_picture_path')
        if profile_picture and os.path.exists(profile_picture):
            try:
                img = Image(profile_picture, width=2*inch, height=2*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 20))
            except:
                pass
        
        # Basic Information Section
        elements.append(Paragraph("Basic Information", heading_style))
        basic_data = [
            ['Name:', student_data.get('name', 'N/A')],
            ['Admission Number:', student_data.get('admission_number', 'N/A')],
            ['Class:', f"{student_data.get('class', 'N/A')}-{student_data.get('section', 'N/A')}"],
            ['Roll Number:', student_data.get('roll_number', 'N/A')],
            ['EMIS:', student_data.get('emis', 'N/A')],
            ['Email:', student_data.get('email', 'N/A')]
        ]
        basic_table = Table(basic_data, colWidths=[2*inch, 4*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(basic_table)
        elements.append(Spacer(1, 15))
        
        # Personal Details Section
        elements.append(Paragraph("Personal Details", heading_style))
        personal_data = [
            ['Gender:', student_data.get('gender', 'N/A')],
            ['Date of Birth:', student_data.get('dob', 'N/A')],
            ['Tamil Name:', student_data.get('tamil_name', 'N/A')],
            ['Nationality:', student_data.get('nationality', 'N/A')],
            ['Blood Group:', student_data.get('blood_group', 'N/A')],
            ['Mother Tongue:', student_data.get('mother_tongue', 'N/A')],
            ['Community:', student_data.get('community', 'N/A')],
            ['Caste:', student_data.get('caste', 'N/A')],
            ['Religion:', student_data.get('religion', 'N/A')],
            ['Place of Birth:', student_data.get('place_of_birth', 'N/A')],
            ['Aadhaar:', student_data.get('aadhaar', 'N/A')],
            ['Disability:', student_data.get('disability', 'N/A')]
        ]
        personal_table = Table(personal_data, colWidths=[2*inch, 4*inch])
        personal_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(personal_table)
        elements.append(Spacer(1, 15))
        
        # Academic Information Section
        elements.append(Paragraph("Academic Information", heading_style))
        academic_data = [
            ['Current Class:', student_data.get('current_class', 'N/A')],
            ['Admission Class:', student_data.get('admission_class', 'N/A')],
            ['Admission Year:', student_data.get('admission_year', 'N/A')],
            ['Admission Date:', student_data.get('admission_date', 'N/A')],
            ['House:', student_data.get('house', 'N/A')],
            ['Teacher Ward:', student_data.get('teacher_ward', 'N/A')],
            ['RTE:', student_data.get('rte', 'N/A')],
            ['Sports Quota:', student_data.get('sports_quota', 'N/A')],
            ['Previous School:', student_data.get('prev_school', 'N/A')],
            ['Previous Board:', student_data.get('prev_board', 'N/A')]
        ]
        academic_table = Table(academic_data, colWidths=[2*inch, 4*inch])
        academic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(academic_table)
        elements.append(Spacer(1, 15))
        
        # Contact Information Section
        elements.append(Paragraph("Contact Information", heading_style))
        contact_data = [
            ['Address:', student_data.get('address', 'N/A')],
            ['Contact:', student_data.get('contact', 'N/A')],
            ['Alternate Contact:', student_data.get('alt_contact', 'N/A')],
            ['Country:', student_data.get('country', 'N/A')],
            ['State:', student_data.get('state', 'N/A')],
            ['City:', student_data.get('city', 'N/A')],
            ['Pincode:', student_data.get('pincode', 'N/A')],
            ['Status:', student_data.get('status', 'N/A')]
        ]
        contact_table = Table(contact_data, colWidths=[2*inch, 4*inch])
        contact_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(contact_table)
        elements.append(Spacer(1, 15))
        
        # Father's Information Section
        elements.append(Paragraph("Father's Information", heading_style))
        father_data = [
            ['Father\'s Name:', student_data.get('father_name', 'N/A')],
            ['Father\'s Tamil Name:', student_data.get('father_name_tamil', 'N/A')],
            ['Father\'s Contact:', student_data.get('father_contact', 'N/A')],
            ['Father\'s Email:', student_data.get('father_email', 'N/A')],
            ['Father\'s Qualification:', student_data.get('father_qualification', 'N/A')],
            ['Father\'s Occupation:', student_data.get('father_occupation', 'N/A')],
            ['Father\'s Income:', student_data.get('father_income', 'N/A')]
        ]
        father_table = Table(father_data, colWidths=[2*inch, 4*inch])
        father_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(father_table)
        elements.append(Spacer(1, 15))
        
        # Mother's Information Section
        elements.append(Paragraph("Mother's Information", heading_style))
        mother_data = [
            ['Mother\'s Name:', student_data.get('mother_name', 'N/A')],
            ['Mother\'s Tamil Name:', student_data.get('mother_name_tamil', 'N/A')],
            ['Mother\'s Contact:', student_data.get('mother_contact', 'N/A')],
            ['Mother\'s Email:', student_data.get('mother_email', 'N/A')],
            ['Mother\'s Qualification:', student_data.get('mother_qualification', 'N/A')],
            ['Mother\'s Occupation:', student_data.get('mother_occupation', 'N/A')],
            ['Mother\'s Income:', student_data.get('mother_income', 'N/A')]
        ]
        mother_table = Table(mother_data, colWidths=[2*inch, 4*inch])
        mother_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(mother_table)
        elements.append(Spacer(1, 15))
        
        # Guardian Information Section
        elements.append(Paragraph("Guardian Information", heading_style))
        guardian_data = [
            ['Guardian\'s Name:', student_data.get('guardian_name', 'N/A')],
            ['Guardian\'s Contact:', student_data.get('guardian_contact', 'N/A')],
            ['Guardian\'s Email:', student_data.get('guardian_email', 'N/A')],
            ['Child Living With:', student_data.get('child_living', 'N/A')],
            ['Rights on Child:', student_data.get('rights_on_child', 'N/A')]
        ]
        guardian_table = Table(guardian_data, colWidths=[2*inch, 4*inch])
        guardian_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(guardian_table)
        elements.append(Spacer(1, 15))
        
        # Medical Information Section
        elements.append(Paragraph("Medical Information", heading_style))
        medical_data = [
            ['Medical Blood Group:', student_data.get('med_blood_group', 'N/A')],
            ['Diseases:', student_data.get('diseases', 'N/A')],
            ['Allergies:', student_data.get('allergies', 'N/A')],
            ['Medicines:', student_data.get('medicines', 'N/A')],
            ['Hospital:', student_data.get('hospital', 'N/A')],
            ['Doctor:', student_data.get('doctor', 'N/A')]
        ]
        medical_table = Table(medical_data, colWidths=[2*inch, 4*inch])
        medical_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f0ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0052cc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(medical_table)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create filename as requested: <name> <class and section> details.pdf
        name = student_data.get('name', 'Student').replace(' ', '_')
        class_section = f"{student_data.get('class', 'N/A')}{student_data.get('section', 'N/A')}"
        filename = f"{name}_{class_section}_details.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        
        return response
        
    except Exception as e:
        return JsonResponse({"error": f"Error generating PDF: {str(e)}"}, status=500)



def landing_view(request):
    return render(request, 'users/landing.html')


def admin_master(request):
    # Check if admin is logged in via session
    admin_id = request.session.get('admin_id')
    if not admin_id:
        messages.error(request, 'Please login to access this page.')
        return redirect('admin_login')
    
    # Get admin name from session (fallback to email if name not set)
    admin_name = request.session.get('admin_name', request.session.get('admin_email', 'Admin'))
    return render(request, 'users/admin_master.html', {'admin_name': admin_name})



# Admin Add Exam Entry
def admin_exam_add(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    if request.method == 'POST':
        class_name = request.POST.get('class_id')
        section = request.POST.get('section')
        subject = request.POST.get('subject')
        exam_date = request.POST.get('exam_date')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        room = request.POST.get('room')
        invigilator_id = request.POST.get('invigilator_id')

        # Construct class_id
        class_id = f"{class_name}{section}" if section else class_name

        # Validate inputs
        if not class_name or not class_id:
            messages.error(request, 'Please select a valid class.')
            return redirect('admin_exam_add')
        if not subject or not invigilator_id or not exam_date or not start_time or not end_time:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('admin_exam_add')

        with connection.cursor() as cursor:
            # Check for conflicts
            cursor.execute("""
                SELECT id FROM exams 
                WHERE (class_id = %s OR invigilator_id = %s)
                AND exam_date = %s
                AND start_time <= %s AND end_time >= %s
            """, [class_id, invigilator_id, exam_date, end_time, start_time])
            conflict = cursor.fetchone()
            
            if conflict:
                messages.error(request, 'Scheduling conflict detected.')
                return redirect('admin_exam_add')
            
            # Insert into exams
            cursor.execute("""
                INSERT INTO exams (class_id, subject, exam_date, start_time, 
                                 end_time, room, invigilator_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, [class_id, subject, exam_date, start_time, end_time, room or None, invigilator_id])
        
        messages.success(request, 'Exam entry added successfully.')
        return redirect('admin_timetable')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT id, name FROM teachers")
        teachers = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    
    return render(request, 'users/admin_exam_add.html', {
        'classes': classes, 'sections': sections, 'teachers': teachers
    })

# Admin Exam Schedule Creation (similar to weekly timetable)
def admin_exam_schedule(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    # Example: Define exam periods or dates; store in session if dynamic
    if 'num_exam_days' not in request.session:
        request.session['num_exam_days'] = 5  # Default 5 days
    
    # Handle add/delete days (optional)
    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')
        current_days = request.session['num_exam_days']
        if action == 'add' and current_days < 10:
            request.session['num_exam_days'] = current_days + 1
        elif action == 'delete' and current_days > 1:
            request.session['num_exam_days'] = current_days - 1
        request.session.modified = True
        return redirect('admin_exam_schedule')
    
    num_days = request.session['num_exam_days']
    exam_days = list(range(1, num_days + 1))  # Or generate dates dynamically
    
    if request.method == 'POST' and 'create_schedule' in request.POST:
        class_name = request.POST.get('class')
        section = request.POST.get('section')
        class_id = f"{class_name}{section}" if section else class_name
        
        if not class_name or not class_id:
            messages.error(request, 'Please select a valid class.')
            return redirect('admin_exam_schedule')
        
        with connection.cursor() as cursor:
            for day in exam_days:
                subject = request.POST.get(f'subject_day_{day}')
                exam_date = request.POST.get(f'exam_date_{day}')
                start_time = request.POST.get(f'start_time_{day}')
                end_time = request.POST.get(f'end_time_{day}')
                room = request.POST.get(f'room_{day}')
                invigilator_id = request.POST.get(f'invigilator_{day}')
                
                if not (subject and invigilator_id and exam_date and start_time and end_time):
                    continue
                
                # Check conflicts (adapt query as needed)
                cursor.execute("""
                    SELECT id FROM exams 
                    WHERE (class_id = %s OR invigilator_id = %s)
                    AND exam_date = %s
                    AND start_time <= %s AND end_time >= %s
                """, [class_id, invigilator_id, exam_date, end_time, start_time])
                conflict = cursor.fetchone()
                
                if conflict:
                    messages.error(request, f'Conflict on day {day}.')
                    continue
                
                cursor.execute("""
                    INSERT INTO exams (class_id, subject, exam_date, start_time, 
                                     end_time, room, invigilator_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, [class_id, subject, exam_date, start_time, end_time, room or None, invigilator_id])
        
        messages.success(request, 'Exam schedule created successfully.')
        return redirect('admin_timetable')
    
    # Fetch data for form
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT id, name FROM teachers")
        teachers = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    
    return render(request, 'users/admin_exam_schedule.html', {
        'exam_days': exam_days, 'teachers': teachers, 
        'classes': classes, 'sections': sections, 'num_days': num_days
    })

def admin_exam_pdf_download(request):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    # Fetch all exams (you can add filters if needed, e.g., from GET params)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT e.class_id, e.subject, e.exam_date, e.start_time, 
                   e.end_time, COALESCE(e.room, 'N/A') as room, 
                   COALESCE(tch.name, 'N/A') as invigilator_name
            FROM exams e
            LEFT JOIN teachers tch ON e.invigilator_id = tch.id
            ORDER BY e.exam_date, e.start_time
        """)
        exams_data = cursor.fetchall()
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="exam_schedule.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=30,
        textColor=colors.green,
        alignment=1  # Center
    )
    p = Paragraph("Manavargal School - Exam Schedule", title_style)
    story.append(p)
    story.append(Spacer(1, 12))
    
    # Add date/time info
    date_style = styles['Normal']
    p = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style)
    story.append(p)
    story.append(Spacer(1, 20))
    
    # Table data
    if exams_data:
        table_data = [['Class', 'Subject', 'Date', 'Time', 'Room', 'Invigilator']]
        for row in exams_data:
            class_id = row[0]
            # Pretty print class (e.g., '10A' -> 'Class 10 - Section A')
            class_name = class_id[:-1] if len(class_id) > 1 and class_id[-1].isalpha() else class_id
            section = f" - Section {class_id[-1]}" if len(class_id) > 1 and class_id[-1].isalpha() else ''
            pretty_class = f"{class_name}{section}"
            
            time_slot = f"{row[3].strftime('%H:%M')} - {row[4].strftime('%H:%M')}"
            table_data.append([pretty_class, row[1], row[2].strftime('%Y-%m-%d'), time_slot, row[5], row[6]])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        story.append(table)
    else:
        p = Paragraph("No exams scheduled yet.", styles['Normal'])
        story.append(p)
    
    doc.build(story)
    return response


def admin_exam_edit(request, exam_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    # Fetch the exam
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, class_id, subject, exam_date, start_time, end_time, room, invigilator_id FROM exams WHERE id = %s", [exam_id])
        exam_row = cursor.fetchone()
        if not exam_row:
            messages.error(request, 'Exam not found.')
            return redirect('admin_timetable')
        
        exam = {
            'id': exam_row[0],
            'class_id': exam_row[1],
            'subject': exam_row[2],
            'exam_date': exam_row[3],
            'start_time': exam_row[4],
            'end_time': exam_row[5],
            'room': exam_row[6],
            'invigilator_id': exam_row[7],
        }
    
    # Parse class_id to current_class and current_section
    class_id = exam['class_id']
    if len(class_id) > 1 and class_id[-1].isalpha():
        current_class = class_id[:-1]
        current_section = class_id[-1]
    else:
        current_class = class_id
        current_section = ''
    
    if request.method == 'POST':
        class_name = request.POST.get('class_id')
        section = request.POST.get('section')
        subject = request.POST.get('subject')
        exam_date = request.POST.get('exam_date')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        room = request.POST.get('room')
        invigilator_id = request.POST.get('invigilator_id')

        # Construct class_id
        class_id_new = f"{class_name}{section}" if section else class_name

        # Validate inputs
        if not class_name or not class_id_new:
            messages.error(request, 'Please select a valid class.')
            return redirect('admin_exam_edit', exam_id=exam_id)
        if not subject or not invigilator_id or not exam_date or not start_time or not end_time:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('admin_exam_edit', exam_id=exam_id)

        with connection.cursor() as cursor:
            # Check for conflicts, excluding current exam
            cursor.execute("""
                SELECT id FROM exams 
                WHERE (class_id = %s OR invigilator_id = %s)
                AND exam_date = %s
                AND start_time <= %s AND end_time >= %s
                AND id != %s
            """, [class_id_new, invigilator_id, exam_date, end_time, start_time, exam_id])
            conflict = cursor.fetchone()
            
            if conflict:
                messages.error(request, 'Scheduling conflict detected.')
                return redirect('admin_exam_edit', exam_id=exam_id)
            
            # Update exams
            cursor.execute("""
                UPDATE exams SET class_id = %s, subject = %s, exam_date = %s, start_time = %s, 
                                 end_time = %s, room = %s, invigilator_id = %s
                WHERE id = %s
            """, [class_id_new, subject, exam_date, start_time, end_time, room or None, invigilator_id, exam_id])
        
        messages.success(request, 'Exam entry updated successfully.')
        return redirect('admin_timetable')
    
    # Fetch data for form
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT class FROM student_page1")
        classes = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT section FROM student_page1 WHERE section IS NOT NULL")
        sections = [row[0] for row in cursor.fetchall()]
        cursor.execute("SELECT id, name FROM teachers")
        teachers = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    
    return render(request, 'users/admin_exam_edit.html', {
        'classes': classes, 'sections': sections, 'teachers': teachers,
        'exam': exam, 'current_class': current_class, 'current_section': current_section
    })

# Add this view to your views.py

# Updated admin_exam_delete view - remove the delete from non-existent table
def admin_exam_delete(request, exam_id):
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    # Fetch the exam for confirmation
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT e.id, e.class_id, e.subject, e.exam_date, e.start_time, e.end_time, 
                   COALESCE(e.room, 'N/A') as room, COALESCE(t.name, 'N/A') as invigilator_name
            FROM exams e
            LEFT JOIN teachers t ON e.invigilator_id = t.id
            WHERE e.id = %s
        """, [exam_id])
        exam = cursor.fetchone()
        
        if not exam:
            messages.error(request, 'Exam not found.')
            return redirect('admin_timetable')
    
    if request.method == 'POST':
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM exams WHERE id = %s", [exam_id])
            # Removed the DELETE from exam_students since the table doesn't exist
        
        messages.success(request, 'Exam entry deleted successfully.')
        return redirect('admin_timetable')
    
    return render(request, 'users/admin_exam_delete.html', {
        'exam': {
            'id': exam[0],
            'class_id': exam[1],
            'subject': exam[2],
            'exam_date': exam[3],
            'start_time': exam[4],
            'end_time': exam[5],
            'room': exam[6],
            'invigilator_name': exam[7],
        }
    })


def admin_timetable_bulk_delete(request):
    """Delete multiple timetable entries at once"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.error(request, 'No entries selected.')
            return redirect('admin_timetable')
        
        with connection.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(ids))
            cursor.execute(f"DELETE FROM timetable WHERE id IN ({placeholders})", ids)
        
        messages.success(request, f'{len(ids)} timetable entries deleted successfully.')
    
    return redirect('admin_timetable')

from datetime import timedelta
from django.db import IntegrityError

def admin_timetable_bulk_copy(request):
    """Copy selected timetable entries to next week"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    ids = request.GET.get('ids', '').split(',')
    if not ids or ids == ['']:
        messages.error(request, 'No entries selected.')
        return redirect('admin_timetable')
    
    with connection.cursor() as cursor:
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            SELECT class_id, subject, teacher_id, day_of_week, 
                   start_time, end_time, room, week_start_date
            FROM timetable 
            WHERE id IN ({placeholders})
        """, ids)
        
        entries = cursor.fetchall()
        copied_count = 0
        
        for entry in entries:
            week_start = entry[7] + timedelta(weeks=1)  # Next week
            week_end = week_start + timedelta(days=6)
            
            try:
                cursor.execute("""
                    INSERT INTO timetable (class_id, subject, teacher_id, day_of_week, 
                                         start_time, end_time, room, week_start_date, week_end_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [entry[0], entry[1], entry[2], entry[3], entry[4], entry[5], 
                      entry[6], week_start, week_end])
                copied_count += 1
            except IntegrityError:
                continue
        
        messages.success(request, f'{copied_count} entries copied to next week.')
    
    return redirect('admin_timetable')


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

def admin_timetable_export_excel(request):
    """Export timetable to Excel"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    ids = request.GET.get('ids', '')
    
    with connection.cursor() as cursor:
        if ids:
            id_list = ids.split(',')
            placeholders = ','.join(['%s'] * len(id_list))
            query = f"""
                SELECT t.class_id, t.subject, tch.name, t.day_of_week, 
                       t.start_time, t.end_time, t.room, t.week_start_date
                FROM timetable t
                LEFT JOIN teachers tch ON t.teacher_id = tch.id
                WHERE t.id IN ({placeholders})
                ORDER BY t.class_id, t.week_start_date, t.day_of_week, t.start_time
            """
            cursor.execute(query, id_list)
        else:
            cursor.execute("""
                SELECT t.class_id, t.subject, tch.name, t.day_of_week, 
                       t.start_time, t.end_time, t.room, t.week_start_date
                FROM timetable t
                LEFT JOIN teachers tch ON t.teacher_id = tch.id
                ORDER BY t.class_id, t.week_start_date, t.day_of_week, t.start_time
            """)
        
        entries = cursor.fetchall()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Headers
    headers = ['Class', 'Subject', 'Teacher', 'Day', 'Start Time', 'End Time', 'Room', 'Week Start']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="00a676", end_color="00a676", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    for entry in entries:
        row = [
            entry[0],  # Class
            entry[1],  # Subject
            entry[2] or 'N/A',  # Teacher
            entry[3],  # Day
            entry[4].strftime('%H:%M') if entry[4] else '',  # Start time
            entry[5].strftime('%H:%M') if entry[5] else '',  # End time
            entry[6] or 'N/A',  # Room
            entry[7].strftime('%Y-%m-%d') if entry[7] else ''  # Week start
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=timetable_export.xlsx'
    
    wb.save(response)
    return response

def admin_exam_bulk_delete(request):
    """Delete multiple exam entries at once"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.error(request, 'No exams selected.')
            return redirect('admin_timetable')
        
        with connection.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(ids))
            cursor.execute(f"DELETE FROM exams WHERE id IN ({placeholders})", ids)
        
        messages.success(request, f'{len(ids)} exam(s) deleted successfully.')
    
    return redirect('admin_timetable')



def admin_exam_export_excel(request):
    """Export exams to Excel"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    ids = request.GET.get('ids', '')
    
    with connection.cursor() as cursor:
        if ids:
            id_list = ids.split(',')
            placeholders = ','.join(['%s'] * len(id_list))
            query = f"""
                SELECT e.class_id, e.subject, e.exam_date, e.start_time, 
                       e.end_time, e.room, COALESCE(tch.name, 'N/A') as invigilator_name
                FROM exams e
                LEFT JOIN teachers tch ON e.invigilator_id = tch.id
                WHERE e.id IN ({placeholders})
                ORDER BY e.class_id, e.exam_date, e.start_time
            """
            cursor.execute(query, id_list)
        else:
            cursor.execute("""
                SELECT e.class_id, e.subject, e.exam_date, e.start_time, 
                       e.end_time, e.room, COALESCE(tch.name, 'N/A') as invigilator_name
                FROM exams e
                LEFT JOIN teachers tch ON e.invigilator_id = tch.id
                ORDER BY e.class_id, e.exam_date, e.start_time
            """)
        
        entries = cursor.fetchall()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Schedule"
    
    # Headers
    headers = ['Class', 'Subject', 'Exam Date', 'Start Time', 'End Time', 'Room', 'Invigilator']
    ws.append(headers)
    
    # Style headers (orange theme for exams)
    header_fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    for entry in entries:
        row = [
            entry[0],  # Class
            entry[1],  # Subject
            entry[2].strftime('%Y-%m-%d') if entry[2] else '',  # Exam date
            entry[3].strftime('%H:%M') if entry[3] else '',  # Start time
            entry[4].strftime('%H:%M') if entry[4] else '',  # End time
            entry[5] or 'N/A',  # Room
            entry[6]  # Invigilator
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=exam_schedule.xlsx'
    
    wb.save(response)
    return response



from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from django.utils import timezone

def admin_exam_print(request):
    """Generate printable PDF of exam schedule"""
    if not request.session.get('admin_id'):
        messages.error(request, 'You must be logged in to access this page.')
        return redirect('admin_login')
    
    ids = request.GET.get('ids', '')
    
    with connection.cursor() as cursor:
        if ids:
            id_list = ids.split(',')
            placeholders = ','.join(['%s'] * len(id_list))
            query = f"""
                SELECT e.class_id, e.subject, e.exam_date, e.start_time, 
                       e.end_time, e.room, COALESCE(tch.name, 'N/A') as invigilator_name
                FROM exams e
                LEFT JOIN teachers tch ON e.invigilator_id = tch.id
                WHERE e.id IN ({placeholders})
                ORDER BY e.exam_date, e.class_id, e.start_time
            """
            cursor.execute(query, id_list)
        else:
            cursor.execute("""
                SELECT e.class_id, e.subject, e.exam_date, e.start_time, 
                       e.end_time, e.room, COALESCE(tch.name, 'N/A') as invigilator_name
                FROM exams e
                LEFT JOIN teachers tch ON e.invigilator_id = tch.id
                ORDER BY e.exam_date, e.class_id, e.start_time
            """)
        
        exams_data = cursor.fetchall()
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="exam_schedule.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#f59e0b'),
        alignment=1  # Center
    )
    title = Paragraph("Manavargal School - Exam Schedule", title_style)
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    # Date/time info
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        textColor=colors.grey
    )
    subtitle = Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %H:%M')}", subtitle_style)
    story.append(subtitle)
    story.append(Spacer(1, 0.5*inch))
    
    # Table data
    if exams_data:
        table_data = [['Class', 'Subject', 'Date', 'Time', 'Room', 'Invigilator']]
        
        for row in exams_data:
            time_slot = f"{row[3].strftime('%H:%M')}-{row[4].strftime('%H:%M')}" if row[3] and row[4] else 'N/A'
            exam_date = row[2].strftime('%d/%m/%Y') if row[2] else 'N/A'
            
            table_data.append([
                row[0],  # Class
                row[1],  # Subject
                exam_date,  # Date
                time_slot,  # Time
                row[5] or 'N/A',  # Room
                row[6]  # Invigilator
            ])
        
        # Create table
        table = Table(table_data, colWidths=[0.8*inch, 1.8*inch, 1.2*inch, 1.2*inch, 0.8*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        story.append(table)
    else:
        no_data = Paragraph("No exams scheduled.", styles['Normal'])
        story.append(no_data)
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer = Paragraph(
        "Note: Please verify all details before the exam date. Contact administration for any discrepancies.",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey)
    )
    story.append(footer)
    
    doc.build(story)
    return response



import os
import urllib.parse

from reportlab.pdfgen import canvas

from django.conf import settings
from django.db import connection
from django.http import JsonResponse, FileResponse, Http404
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

def generate_student_pdf(admission_number, name):
    """
    Generates a student PDF on demand using ReportLab.
    Safe for Windows & Linux.
    """

    pdf_dir = os.path.join(settings.MEDIA_ROOT, "student_pdfs")
    os.makedirs(pdf_dir, exist_ok=True)

    file_path = os.path.join(pdf_dir, f"{admission_number}.pdf")

    c = canvas.Canvas(file_path)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(180, 780, "STUDENT INFORMATION")

    c.setFont("Helvetica", 12)
    c.drawString(100, 740, f"Name : {name}")
    c.drawString(100, 720, f"Admission No : {admission_number}")

    c.setFont("Helvetica", 10)
    c.drawString(100, 680, "This document is auto-generated by the school system.")
    c.drawString(100, 665, "Issued by School Administration.")

    c.save()
    return file_path


def admin_send_student_pdf(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT user_id, name, admission_number
            FROM student_page1
            ORDER BY name
        """)
        students = cursor.fetchall()

    return render(request, "users/send_pdf_whatsapp.html", {
        "students": students
    })


def fetch_students_by_class_section(request):
    student_class = request.GET.get("class")
    section = request.GET.get("section")

    if not student_class or not section:
        return JsonResponse({"students": []})

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                s1.name,
                s1.admission_number,
                s3.contact
            FROM student_page1 s1
            LEFT JOIN student_page3 s3 ON s1.user_id = s3.user_id
            WHERE s1.class = %s AND s1.section = %s
            ORDER BY s1.name
        """, [student_class, section])

        rows = cursor.fetchall()

    students = [
        {
            "name": r[0],
            "admission_number": r[1],
            "contact": r[2] or ""
        }
        for r in rows
    ]

    return JsonResponse({"students": students})


@csrf_exempt
def generate_whatsapp_link(request):
    try:
        mobile = request.POST.get("mobile", "").strip()
        name = request.POST.get("name", "").strip()
        admission_number = request.POST.get("admission_number", "").strip()

        if not mobile or not name or not admission_number:
            return JsonResponse({"error": "Missing required fields"}, status=400)

        # 🔥 Generate PDF on demand
        generate_student_pdf(admission_number, name)

        pdf_url = request.build_absolute_uri(
            f"/media/student_pdfs/{admission_number}.pdf"
        )

        message = (
            "Hello,\n\n"
            "Please find the student document below:\n\n"
            f"Name: {name}\n"
            f"Admission No: {admission_number}\n\n"
            f"Download PDF:\n{pdf_url}\n\n"
            "Regards,\nSchool Administration"
        )

        whatsapp_url = (
            "https://wa.me/"
            + mobile
            + "?text="
            + urllib.parse.quote(message)
        )

        return JsonResponse({"whatsapp_url": whatsapp_url})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def serve_student_pdf(request, filename):
    file_path = os.path.join(
        settings.MEDIA_ROOT,
        "student_pdfs",
        filename
    )

    if not os.path.exists(file_path):
        raise Http404("PDF not found")

    return FileResponse(
        open(file_path, "rb"),
        content_type="application/pdf"
    )
